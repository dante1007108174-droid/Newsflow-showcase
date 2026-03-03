import json
import re
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]

# Keep this file ASCII-only by using unicode escapes for Chinese paths.
INSTANT_QUERY_DIR = "\u751f\u6210\u5c42\u6a21\u578b\u9009\u578b\u2014\u5373\u65f6\u67e5\u8be2"
SRC_NAME = "\u5373\u65f6\u67e5\u8be2-\u5185\u5bb9\u751f\u6210\u6a21\u578b\u6d4b\u8bd5.xlsx"

# Write to a new file to avoid Windows Excel file locks.
OUT_NAME = "\u751f\u6210\u5c42\u6a21\u578b\u9009\u578b_\u5373\u65f6\u67e5\u8be2_V1_1.xlsx"

SRC = ROOT / "docs" / INSTANT_QUERY_DIR / SRC_NAME
OUT = ROOT / "docs" / INSTANT_QUERY_DIR / OUT_NAME


def safe_json_load(text: object):
    if not isinstance(text, str):
        return None, "not_str"
    raw = text.strip()
    has_fence = "```" in raw
    cleaned = re.sub(r"^```json\s*", "", raw, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```\s*$", "", cleaned)
    try:
        return json.loads(cleaned), ("fence" if has_fence else "")
    except Exception:
        return None, "parse_error"


def normalize_llm_payload(raw_text: object):
    """Best-effort extract {top_news, quick_news} JSON from raw payload."""

    obj, meta = safe_json_load(raw_text)
    notes = []
    if meta == "fence":
        notes.append("has_markdown_fence")

    def try_extract(o):
        if isinstance(o, dict) and isinstance(o.get("top_news"), list) and isinstance(o.get("quick_news"), list):
            return o
        # Common wrappers: {"output": "..."} or {"output": {...}}
        if isinstance(o, dict) and "output" in o:
            inner = o.get("output")
            if isinstance(inner, dict):
                notes.append("wrapped_output_dict")
                return try_extract(inner)
            if isinstance(inner, str):
                notes.append("wrapped_output_str")
                inner_obj, inner_meta = safe_json_load(inner)
                if inner_meta == "fence":
                    notes.append("inner_has_markdown_fence")
                return try_extract(inner_obj)
        if isinstance(o, str):
            inner_obj, inner_meta = safe_json_load(o)
            if inner_meta == "fence":
                notes.append("inner_has_markdown_fence")
            return try_extract(inner_obj)
        return None

    extracted = try_extract(obj)
    if extracted is None:
        notes.append("missing_top_quick" if meta != "parse_error" else "json_parse_error")
    return extracted, notes


def parse_dt(s: object):
    if not isinstance(s, str):
        return None
    s = s.strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None


def parse_seconds(s: object):
    """Parse values like '56s', '1.14m', '72.3', '1m14s'."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    if not isinstance(s, str):
        return None
    raw = s.strip().lower()
    if not raw:
        return None

    m = re.fullmatch(r"(?P<m>\d+(?:\.\d+)?)m(?P<s>\d+(?:\.\d+)?)s", raw)
    if m:
        return float(m.group("m")) * 60.0 + float(m.group("s"))

    if raw.endswith("ms"):
        try:
            return float(raw[:-2]) / 1000.0
        except Exception:
            return None
    if raw.endswith("s"):
        try:
            return float(raw[:-1])
        except Exception:
            return None
    if raw.endswith("m"):
        try:
            return float(raw[:-1]) * 60.0
        except Exception:
            return None
    try:
        return float(raw)
    except Exception:
        return None


def extract_param(user_prompt: object, key: str):
    """Extract simple params from a possibly-truncated JSON string via regex.

    Returns:
      - None when the key is not found
      - Parsed value (could be empty string) when found
    """
    if not isinstance(user_prompt, str):
        return None

    # Avoid accidentally extracting nested keys inside news_list items.
    # Most top-level params appear before the huge "news_list" array.
    idx = user_prompt.find('"news_list"')
    search_text = user_prompt if idx == -1 else user_prompt[:idx]
    # Most inputs are JSON; the huge news_list can get truncated near Excel limits.
    # Use regex that searches only for the requested key.
    # Accept string, number, null/bool.
    pat = re.compile(rf'"{re.escape(key)}"\s*:\s*(?P<v>"(?:\\.|[^"])*"|-?\d+(?:\.\d+)?|null|true|false)')
    m = pat.search(search_text)
    if not m:
        return None
    v = m.group("v")
    if v == "null":
        return ""
    if v.startswith('"') and v.endswith('"'):
        try:
            s = json.loads(v)
        except Exception:
            s = v.strip('"')
        # Treat common placeholders as empty.
        if isinstance(s, str):
            ss = s.strip()
            if ss in ("", "\u7a7a", "\u4e0d\u9650", "\u65e0"):
                return ""
        return s
    return v


def bucket_by_rate(rate: float, table: list[tuple[float, int]]):
    for min_rate, score in table:
        if rate >= min_rate:
            return score
    return table[-1][1]


def score_json(parsed) -> int:
    return 10 if parsed is not None else 0


def score_structure(parsed) -> int:
    if not isinstance(parsed, dict):
        return 0
    top = parsed.get("top_news")
    quick = parsed.get("quick_news")
    if not isinstance(top, list) or not isinstance(quick, list):
        return 0

    top_req = ["title", "tag", "time", "summary", "source", "link"]
    q_req = ["title", "time", "source", "link"]

    def missing_ratio(items: list, req: list[str]) -> float:
        if not items:
            return 1.0
        miss = 0
        total = len(items) * len(req)
        for it in items:
            if not isinstance(it, dict):
                miss += len(req)
                continue
            for k in req:
                v = it.get(k)
                if v is None or (isinstance(v, str) and not v.strip()):
                    miss += 1
        return miss / total if total else 1.0

    r = max(missing_ratio(top, top_req), missing_ratio(quick, q_req))
    if r == 0:
        base = 10
    elif r <= 0.05:
        base = 8
    elif r <= 0.15:
        base = 6
    elif r <= 0.30:
        base = 4
    else:
        base = 2

    # Soft quantity sanity: instant query usually returns top<=5, quick<=10.
    qty_pen = 0
    if len(top) > 8:
        qty_pen += 1
    if len(quick) > 10:
        qty_pen += 1
    return max(0, base - qty_pen)


def topic_key(topic: str) -> str:
    t = (topic or "").strip().lower()
    if t == "ai" or "ai" in t:
        return "ai"
    if "\u79d1\u6280" in t:
        return "tech"
    if "\u8d22\u7ecf" in t:
        return "finance"
    return "tech"


def normalize_source_filter(sf: str) -> str:
    """Map common placeholders to empty (no filter)."""
    s = (sf or "").strip()
    if not s:
        return ""
    if s in ("\u7a7a", "\u4e0d\u9650", "\u5168\u90e8", "\u5168\u90e8\u6765\u6e90"):
        return ""
    # Some exports may store unreadable placeholder text; treat it as empty.
    if "\u4e0d\u9650" in s or "\u7a7a" in s:
        return ""
    return s


def score_relevance(parsed, topic: str) -> int:
    if not isinstance(parsed, dict):
        return 0
    top = parsed.get("top_news")
    if not isinstance(top, list) or not top:
        return 0

    tokens = {
        "ai": [
            "ai",
            "\u4eba\u5de5\u667a\u80fd",
            "\u5927\u6a21\u578b",
            "llm",
            "gpt",
            "openai",
            "anthropic",
            "deepseek",
            "gemini",
            "chatgpt",
            "agent",
            "\u667a\u80fd\u4f53",
            "\u591a\u6a21\u6001",
        ],
        "tech": [
            "\u82af\u7247",
            "\u534a\u5bfc\u4f53",
            "\u624b\u673a",
            "\u786c\u4ef6",
            "\u64cd\u4f5c\u7cfb\u7edf",
            "os",
            "iphone",
            "android",
            "\u534e\u4e3a",
            "\u82f9\u679c",
            "\u5c0f\u7c73",
            "\u4e09\u661f",
            "\u82f1\u4f1f\u8fbe",
            "nvidia",
            "amd",
            "intel",
            "gpu",
            "\u670d\u52a1\u5668",
            "\u7279\u65af\u62c9",
            "\u7535\u52a8\u8f66",
        ],
        "finance": [
            "\u80a1",
            "a\u80a1",
            "\u6e2f\u80a1",
            "\u7f8e\u80a1",
            "\u592e\u884c",
            "\u964d\u606f",
            "\u5229\u7387",
            "cpi",
            "gdp",
            "ipo",
            "\u878d\u8d44",
            "\u5e76\u8d2d",
            "\u8d22\u62a5",
            "\u8425\u6536",
            "\u5229\u6da6",
            "\u6c47\u7387",
            "\u7f8e\u5143",
            "\u4eba\u6c11\u5e01",
            "\u901a\u80c0",
            "\u57fa\u91d1",
            "\u503a",
            "\u9ec4\u91d1",
        ],
    }
    tset = tokens.get(topic, tokens["tech"])

    def hit(text: object) -> bool:
        s = str(text or "").lower()
        return any(tok in s for tok in tset)

    hits = 0
    total = 0
    for it in top:
        total += 1
        if not isinstance(it, dict):
            continue
        if hit(it.get("title")) or hit(it.get("summary")) or hit(it.get("tag")):
            hits += 1

    rate = hits / total if total else 0.0
    # Prefer a smooth scale to avoid threshold cliffs.
    return max(0, min(10, int(round(rate * 10))))


def score_instruction(
    parsed,
    expected_limit: int | None,
    source_filter: str | None,
    language: str | None,
    *,
    enforce_limit: bool,
    enforce_source: bool,
    enforce_language: bool,
) -> int:
    """Instruction adherence: focus on limit and source filter (if provided).

    Note: language adherence is hard to verify for JSON outputs, so we only lightly
    penalize if a language was explicitly requested but output looks non-compliant.
    """
    if not isinstance(parsed, dict):
        return 0
    top = parsed.get("top_news")
    if not isinstance(top, list):
        return 0

    # If the input didn't provide any constraint params, treat this as N/A.
    if not enforce_limit and not enforce_source and not enforce_language:
        return 5

    # limit
    limit_score = 10
    if enforce_limit and expected_limit is not None:
        diff = abs(len(top) - expected_limit)
        limit_score = 10 if diff == 0 else (5 if diff == 1 else 0)

    # source filter
    source_score = 10
    sf = (source_filter or "").strip()
    if enforce_source and sf:
        total = 0
        ok = 0
        for it in top:
            if not isinstance(it, dict):
                continue
            total += 1
            src = str(it.get("source") or "").strip()
            if src and sf in src:
                ok += 1
        if total == 0:
            source_score = 0
        else:
            rate = ok / total
            source_score = 10 if rate == 1.0 else (5 if rate >= 0.8 else 0)

    # language (best-effort): if language asks English, titles/summaries should contain many ASCII letters.
    lang_score = 10
    lang = (language or "").strip().lower()
    if enforce_language and lang in ("english", "\u82f1\u8bed"):
        sample = " ".join(
            [
                str(it.get("title") or "") + " " + str(it.get("summary") or "")
                for it in top[:3]
                if isinstance(it, dict)
            ]
        )
        ascii_letters = sum(ch.isascii() and ch.isalpha() for ch in sample)
        cjk = sum("\u4e00" <= ch <= "\u9fff" for ch in sample)
        if ascii_letters == 0 and cjk > 0:
            lang_score = 0

    return min(limit_score, source_score, lang_score)


def score_sorting_time(parsed) -> int:
    if not isinstance(parsed, dict):
        return 0
    top = parsed.get("top_news")
    if not isinstance(top, list) or len(top) < 2:
        return 8

    times = [parse_dt(it.get("time")) if isinstance(it, dict) else None for it in top]
    if all(t is None for t in times):
        return 0

    inversions = 0
    pairs = 0
    for i in range(len(times) - 1):
        a, b = times[i], times[i + 1]
        if a is None or b is None:
            continue
        pairs += 1
        if a < b:
            inversions += 1

    if pairs == 0:
        return 4
    inv_rate = inversions / pairs
    if inv_rate == 0:
        return 10
    if inv_rate <= 0.25:
        return 8
    if inv_rate <= 0.5:
        return 6
    if inv_rate <= 0.8:
        return 4
    return 0


def score_summary(parsed) -> int:
    if not isinstance(parsed, dict):
        return 0
    top = parsed.get("top_news")
    if not isinstance(top, list) or not top:
        return 0

    def ok(s: object) -> bool:
        s = str(s or "").strip()
        if not s:
            return False
        if "\u6682\u65e0" in s or "N/A" in s:
            return False
        # Instant query spec typically wants ~80-100 chars; allow some slack.
        if len(s) < 60 or len(s) > 160:
            return False
        if s[-1] not in "\u3002\uff01\uff1f.!?":
            return False
        return True

    ok_n = 0
    total = 0
    for it in top:
        if not isinstance(it, dict):
            continue
        total += 1
        if ok(it.get("summary")):
            ok_n += 1

    if total == 0:
        return 0
    rate = ok_n / total
    return bucket_by_rate(rate, [(0.8, 10), (0.6, 8), (0.4, 6), (0.2, 4), (0.0, 0)])


def score_freshness(parsed, current_time_str: str) -> int:
    if not isinstance(parsed, dict):
        return 0
    top = parsed.get("top_news")
    if not isinstance(top, list) or not top:
        return 0

    now = parse_dt(current_time_str)
    if now is None:
        return 5
    cutoff = now - timedelta(hours=24)

    total = 0
    ok_n = 0
    for it in top:
        if not isinstance(it, dict):
            continue
        total += 1
        t = parse_dt(it.get("time"))
        if t is None:
            continue
        if cutoff <= t <= now:
            ok_n += 1

    if total == 0:
        return 0
    rate = ok_n / total
    return bucket_by_rate(rate, [(1.0, 10), (0.9, 8), (0.75, 6), (0.5, 4), (0.0, 0)])


def cost_score(model: str) -> int:
    return 100 if model == "DeepSeek V3.2" else 60


def main():
    if not SRC.exists():
        raise SystemExit(f"Source workbook not found: {SRC}")

    wb_src = load_workbook(SRC)
    ws_src = wb_src["Sheet1"]

    wb = Workbook()
    ws = wb.active
    ws.title = "\u8bc4\u6d4b\u8bb0\u5f55"

    headers = [
        "Case ID",
        "keyword",
        "\u6a21\u578b",
        "Agent\u63d0\u793a\u8bcd",
        "\u7528\u6237\u8f93\u5165(User_prompt)",
        "\u5b9e\u9645\u8f93\u51fa",
        "JSON\u683c\u5f0f(15%)",
        "\u7ed3\u6784\u5b8c\u6574(15%)",
        "\u76f8\u5173\u6027(15%)",
        "\u6307\u4ee4\u9075\u5faa(15%)",
        "\u6392\u5e8f\u4e00\u81f4\u6027(20%)",
        "\u6458\u8981\u8d28\u91cf(10%)",
        "\u65f6\u6548\u6027(10%)",
        "\u8d28\u91cfRaw(10\u5206\u5236)",
        "\u54cd\u5e94\u65f6\u95f4(s)",
        "\u8d28\u91cf\u5f97\u5206(100)",
        "\u65f6\u6548\u5f97\u5206(100)",
        "\u6210\u672c\u5f97\u5206(100)",
        "\u7efc\u5408\u603b\u5206",
        "\u5ba1\u8ba1\u5907\u6ce8",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Collect test rows where session/turn are numeric.
    rows = []
    for r in range(2, ws_src.max_row + 1):
        session = ws_src.cell(r, 1).value
        turn = ws_src.cell(r, 2).value
        if not isinstance(session, (int, float)) or not isinstance(turn, (int, float)):
            continue
        topic = str(ws_src.cell(r, 3).value or "").strip()
        sys_prompt = ws_src.cell(r, 4).value
        user_prompt = ws_src.cell(r, 5).value
        rows.append((r, int(session), int(turn), topic, sys_prompt, user_prompt))

    # Build 2 rows per case: one per model.
    for src_row, session, turn, topic, sys_prompt, user_prompt in rows:
        tkey = topic_key(topic)
        topic_label = {"ai": "AI", "tech": "\u79d1\u6280", "finance": "\u8d22\u7ecf"}.get(tkey, "\u79d1\u6280")

        # Extract input params from (possibly truncated) JSON.
        keyword_val = extract_param(user_prompt, "keyword")
        language_val = extract_param(user_prompt, "language")
        limit_val = extract_param(user_prompt, "limit")
        source_val = extract_param(user_prompt, "source")
        current_time_val = extract_param(user_prompt, "current_time")

        keyword = str(keyword_val if keyword_val is not None else topic_label).strip()
        language = str(language_val if language_val is not None else "").strip()
        limit_raw = str(limit_val if limit_val is not None else "").strip()
        source_filter = normalize_source_filter(str(source_val if source_val is not None else "").strip())
        current_time = str(current_time_val if current_time_val is not None else "").strip()

        # This source workbook stores huge JSON payloads that may get truncated,
        # and keys like "limit" / "language" can be missing entirely when empty.
        # In that case, we treat them as not provided (N/A) for instruction scoring.
        enforce_limit = bool(limit_raw)
        enforce_language = bool(language)

        # Only enforce when a non-empty filter is explicitly provided.
        enforce_source = bool(source_filter)

        expected_limit = None
        if enforce_limit:
            # Spec says empty => default 5.
            expected_limit = 5
            try:
                if limit_raw:
                    expected_limit = int(float(limit_raw))
            except Exception:
                expected_limit = 5

        for model, out_col, time_col in (
            ("DeepSeek V3.2", 6, 7),
            ("\u8c46\u5305 1.8", 8, 9),
        ):
            output_raw = ws_src.cell(src_row, out_col).value
            resp_time_raw = ws_src.cell(src_row, time_col).value
            resp_s = parse_seconds(resp_time_raw)

            extracted, notes = normalize_llm_payload(output_raw)

            s_json = score_json(extracted)
            s_struct = score_structure(extracted)
            s_rel = score_relevance(extracted, tkey)
            s_instr = score_instruction(
                extracted,
                expected_limit,
                source_filter,
                language,
                enforce_limit=enforce_limit,
                enforce_source=enforce_source,
                enforce_language=enforce_language,
            )
            s_sort = score_sorting_time(extracted)
            s_sum = score_summary(extracted)
            s_fresh = score_freshness(extracted, current_time)

            audit_notes = []
            audit_notes.extend(notes)
            if language:
                audit_notes.append(f"lang={language}")
            if limit_raw:
                audit_notes.append(f"limit={limit_raw}")
            if source_filter:
                audit_notes.append(f"source={source_filter}")

            case_id = f"S{session}-T{turn}-{topic_label}"
            ws.append(
                [
                    case_id,
                    keyword or topic_label,
                    model,
                    sys_prompt,
                    user_prompt,
                    output_raw,
                    s_json,
                    s_struct,
                    s_rel,
                    s_instr,
                    s_sort,
                    s_sum,
                    s_fresh,
                    None,
                    resp_s,
                    None,
                    None,
                    cost_score(model),
                    None,
                    ";".join(audit_notes),
                ]
            )

            rr = ws.max_row
            # quality raw (0-10)
            ws.cell(rr, 14, f"=G{rr}*0.15+H{rr}*0.15+I{rr}*0.15+J{rr}*0.15+K{rr}*0.2+L{rr}*0.1+M{rr}*0.1")
            # quality score (0-100)
            ws.cell(rr, 16, f"=N{rr}*10")
            # latency score (0-100), cap at 60s
            ws.cell(rr, 17, f"=MAX(0,MIN(100,(1-O{rr}/60)*100))")
            # total
            ws.cell(rr, 19, f"=P{rr}*0.75+R{rr}*0.15+Q{rr}*0.10")

    # Column widths
    widths = {
        "A": 16,
        "B": 10,
        "C": 14,
        "D": 24,
        "E": 24,
        "F": 40,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 14,
        "L": 12,
        "M": 12,
        "N": 16,
        "O": 12,
        "P": 12,
        "Q": 12,
        "R": 12,
        "S": 12,
        "T": 20,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("\u6c47\u603b\u5bf9\u6bd4")
    ws2.append(["\u6a21\u578b", "\u8d28\u91cf\u5747\u5206(75%)", "\u6210\u672c\u5747\u5206(15%)", "\u65f6\u6548\u5747\u5206(10%)", "\u6700\u7ec8\u52a0\u6743\u603b\u5206", "\u63a8\u8350"])
    for c in ws2[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    models = ["DeepSeek V3.2", "\u8c46\u5305 1.8"]
    for i, m in enumerate(models, start=2):
        ws2.cell(i, 1, m)
        ws2.cell(i, 2, f"=AVERAGEIF(\u8bc4\u6d4b\u8bb0\u5f55!C:C,\"{m}\",\u8bc4\u6d4b\u8bb0\u5f55!P:P)")
        ws2.cell(i, 3, f"=AVERAGEIF(\u8bc4\u6d4b\u8bb0\u5f55!C:C,\"{m}\",\u8bc4\u6d4b\u8bb0\u5f55!R:R)")
        ws2.cell(i, 4, f"=AVERAGEIF(\u8bc4\u6d4b\u8bb0\u5f55!C:C,\"{m}\",\u8bc4\u6d4b\u8bb0\u5f55!Q:Q)")
        ws2.cell(i, 5, f"=B{i}*0.75+C{i}*0.15+D{i}*0.10")

    ws2.cell(2, 6, "=IF(E2>E3,\"YES\",\"\")")
    ws2.cell(3, 6, "=IF(E3>E2,\"YES\",\"\")")
    for col in "ABCDEF":
        ws2.column_dimensions[col].width = 20

    # Rubric sheet
    ws3 = wb.create_sheet("\u8bc4\u5206\u6807\u51c6")
    ws3.append(["\u7ef4\u5ea6", "\u6743\u91cd", "\u8bf4\u660e"])
    for c in ws3[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws3.append(["JSON\u683c\u5f0f", "15%", "\u80fd\u4ecepayload\u4e2d\u89e3\u6790\u51fa{top_news,quick_news}\u7ed3\u6784=10\uff1b\u5426\u5219=0"])
    ws3.append([
        "\u7ed3\u6784\u5b8c\u6574",
        "15%",
        "top/quick\u5fc5\u987b\u5b58\u5728\u4e14\u4e3a\u6570\u7ec4\uff1b\u5fc5\u586b\u5b57\u6bb5\u7f3a\u5931\u7387\u8d8a\u4f4e\u5206\u8d8a\u9ad8\uff0c\u8d85\u91cf\u8fd4\u56de\u4f1a\u6263\u5206",
    ])
    ws3.append([
        "\u76f8\u5173\u6027",
        "15%",
        "\u7528\u4e3b\u9898\u5173\u952e\u8bcd\u8868\u5339\u914d\u6807\u9898/\u6458\u8981/Tag\uff0c\u547d\u4e2d\u7387\u53d8\u6362\u4e3a0-10\u8fde\u7eed\u5f97\u5206",
    ])
    ws3.append([
        "\u6307\u4ee4\u9075\u5faa",
        "15%",
        "\u68c0\u67e5limit\u6570\u91cf\u3001source\u4ec5\u4fdd\u7559\u6307\u5b9a\u6765\u6e90(\u5982\u4f20\u5165)\uff0c\u82f1\u6587\u8bed\u8a00\u8981\u6c42\u4e3a\u82f1\u6587(\u5982\u4f20\u5165)",
    ])
    ws3.append([
        "\u6392\u5e8f\u4e00\u81f4\u6027",
        "20%",
        "\u4f5c\u4e3a\u81ea\u6d3d\u6027\u68c0\u67e5\uff0c\u7b80\u5316\u4e3atop_news\u65f6\u95f4\u662f\u5426\u5927\u4f53\u65b0->\u65e7(\u4e0d\u662f\u5168\u91cf\u70ed\u5ea6\u6392\u540d\u7684\u4e25\u683c\u9a8c\u8bc1)",
    ])
    ws3.append([
        "\u6458\u8981\u8d28\u91cf",
        "10%",
        "\u6458\u8981\u975e\u7a7a\uff0c\u957f\u5ea6\u7ea6\u4e2d\uff0c\u65e0\u201c\u6682\u65e0\u201d\u7b49\u7a7a\u8bdd\uff0c\u672b\u5c3e\u6807\u70b9\u6b63\u5e38\uff1b\u6309\u5408\u683c\u6bd4\u4f8b\u6253\u5206",
    ])
    ws3.append([
        "\u65f6\u6548\u6027",
        "10%",
        "\u65b0\u95fb\u65f6\u95f4\u662f\u5426\u5728current_time\u524d24h\u5185\uff1b\u6309\u5408\u683c\u6bd4\u4f8b\u6253\u5206(\u82e5\u65e0current_time\u5219\u7ed9\u4e2d\u6027\u5206)",
    ])
    ws3.append([
        "\u603b\u6743\u91cd",
        "-",
        "\u603b\u5206=\u8d28\u91cf(75%)+\u6210\u672c(15%)+\u65f6\u6548(10%)\uff1b\u6210\u672c\uff1aDeepSeek=100, \u8c46\u5305=60\uff1b\u5b8c\u6210\u65f6\u95f4\u5f97\u5206\uff1aMAX(0,MIN(100,(1-\u79d2/60)*100))",
    ])

    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 110
    for r in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=3):
        for c in r:
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    wb.save(OUT)


if __name__ == "__main__":
    main()
