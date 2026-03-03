#!/usr/bin/env python3
"""Update badcase合集.xlsx with individual test results for each case"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
import sys

def main():
    # Test results
    test_results = {
        # Fixed cases
        "GEN-05": {"status": "已修复", "result": "来源不存在时正确提示\"暂无可用的新闻数据\""},
        "GEN-15": {"status": "已修复", "result": "全部过期时正确提示\"24小时内暂无「AI」相关新闻\""},
        "GEN-17": {"status": "已修复", "result": "正确聚类，显示\"3家媒体报道|跨2类媒体\""},
        "GEN-24": {"status": "已修复", "result": "英文格式正确，包含Hotspot标记"},
        "GEN-27": {"status": "已修复", "result": "空数据集时正确提示\"暂无可用的新闻数据\""},
        "GEN-30": {"status": "已修复", "result": "英文格式正确，输出2条新闻"},
        
        # Still failing cases
        "GEN-07": {"status": "待修复", "result": "只识别一个来源（华尔街见闻），漏了虎嗅"},
        "GEN-14": {"status": "待修复", "result": "时间解析问题，无法正确识别24小时内新闻"},
        "GEN-16": {"status": "待修复", "result": "时间解析问题，新鲜新闻被过滤"},
        "GEN-25": {"status": "待修复", "result": "来源名被翻译（华尔街见闻→TechCrunch）"},
        "GEN-26": {"status": "待修复", "result": "空keyword时返回\"暂无\"而不是通用新闻"},
        "GEN-29": {"status": "待修复", "result": "来源匹配失败，即使数据存在也返回\"暂无\""},
    }
    
    input_file = '/d/daily-ai-news/docs/badcase合集.xlsx'
    output_file = '/d/daily-ai-news/docs/badcase合集_v3.xlsx'
    
    try:
        # Read existing file
        df = pd.read_excel(input_file, sheet_name=0)
        print(f"Read {len(df)} rows from original file")
        print(f"Columns: {list(df.columns)}")
        
        # Add new columns for detailed tracking
        if '重测状态' not in df.columns:
            df['重测状态'] = ''
        if '重测时间' not in df.columns:
            df['重测时间'] = ''
        if '测试结果' not in df.columns:
            df['测试结果'] = ''
        
        # Update each row
        updated_count = 0
        for idx, row in df.iterrows():
            case_id = str(row.get('ID', ''))
            
            # Check if this row contains any GEN IDs
            for gen_id in test_results.keys():
                if gen_id in case_id:
                    result = test_results[gen_id]
                    
                    # Update the 问题状态 if it's a single case row
                    if case_id == gen_id or case_id.startswith(gen_id + ':'):
                        df.at[idx, '问题状态'] = result['status']
                    
                    # Add detailed test info
                    df.at[idx, '重测状态'] = result['status']
                    df.at[idx, '重测时间'] = datetime.now().strftime('%Y-%m-%d %H:%M')
                    df.at[idx, '测试结果'] = result['result']
                    updated_count += 1
                    break
        
        # Save to new file
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            
            # Get workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            # Format header
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"\n✅ Updated {updated_count} rows")
        print(f"✅ Saved to: {output_file}")
        
        # Print summary
        print("\n=== Update Summary ===")
        fixed_cases = [k for k, v in test_results.items() if v['status'] == '已修复']
        fail_cases = [k for k, v in test_results.items() if v['status'] == '待修复']
        
        print(f"\n已修复 ({len(fixed_cases)} 个):")
        for case in fixed_cases:
            print(f"  ✓ {case}: {test_results[case]['result']}")
        
        print(f"\n待修复 ({len(fail_cases)} 个):")
        for case in fail_cases:
            print(f"  ✗ {case}: {test_results[case]['result']}")
        
        return 0
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    sys.exit(main())
