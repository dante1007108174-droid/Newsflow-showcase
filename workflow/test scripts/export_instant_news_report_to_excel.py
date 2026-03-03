"""Export instant_news_generation_node workflow suite + latest run report to Excel.

Inputs:
- tests/instant_news_suite.json
- tests/reports/instant_news_report.json

Output:
- docs/instant_news_workflow_data_test.xlsx

This is a deterministic export step (no network calls).

版式特点（对齐 subscribe_using 风格）：
- 深灰色表头 + 白色字体 + 首行冻结
- 条件格式：PASS=浅绿, FAIL=浅红, SUGGEST_REVIEW=浅黄
- 列宽自适应（限制最大70）
- 关键列自动换行
- 全中文表达
- 失败原因用自然语言描述（带修复建议）
- 包含完整的测评维度：硬性规则 + 质量维度评分
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]

SUITE_PATH = ROOT / "tests" / "instant_news_suite.json"
REPORT_PATH = ROOT / "tests" / "reports" / "instant_news_report.json"
GPT_SCORES_PATH = ROOT / "tests" / "reports" / "instant_news_gpt_scores.json"
OUTPUT_PATH = ROOT / "docs" / "instant_news_workflow_data_test.xlsx"


def _load_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _fmt(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return str(value)
    return json.dumps(value, ensure_ascii=False)


def _write_table(ws, headers: List[str], rows: List[List[Any]], freeze: bool = True) -> None:
    """写入表格并应用基础样式"""
    header_fill = PatternFill("solid", fgColor="1F2937")  # 深灰色
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(vertical="top")
    
    ws.append(headers)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
    
    for r in rows:
        ws.append(r)
    
    if freeze:
        ws.freeze_panes = "A2"
    
    # 列宽自适应（限制最大70）
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for r in rows:
            if col - 1 >= len(r):
                continue
            s = str(r[col - 1] if r[col - 1] is not None else "")
            if len(s) > max_len:
                max_len = len(s)
        width = min(max(10, max_len + 2), 70)
        ws.column_dimensions[get_column_letter(col)].width = width


def _translate_check_name(check_name: str) -> str:
    """检查项名称中文化"""
    mapping = {
        "non_empty": "输出非空",
        "format": "格式正确",
        "structure": "结构完整",
        "limit_max": "条数限制",
        "source_only": "来源约束",
        "language_en": "英文输出",
        "timeliness": "时效性",
        "raw_query_constraints": "原始查询约束",
        "empty_or_friendly": "空结果友好提示",
        "exception": "无异常",
    }
    return mapping.get(check_name, check_name)


def _get_hard_rule_status(checks: List[Dict], rule_name: str) -> str:
    """获取硬性规则的检查结果"""
    for check in checks:
        if check.get("name") == rule_name:
            if check.get("passed"):
                return "通过"
            else:
                return "失败"
    return "未检查"


def _load_gpt_scores() -> Dict[str, Any]:
    """加载GPT评分结果"""
    if GPT_SCORES_PATH.exists():
        return _load_json(GPT_SCORES_PATH).get("scores", {})
    return {}


def _get_quality_score(r: Dict[str, Any], gpt_scores: Dict[str, Any]) -> tuple[str, str, str, str]:
    """
    获取质量维度评分
    返回: (去重评分, 排序评分, 相关性评分, 评分来源)
    """
    case_id = r.get("id", "")
    suite_case = r.get("suite_case", {})
    expect = suite_case.get("expect", {})
    review_items = expect.get("review_items", [])
    status = r.get("status", "")
    
    if status != "SUGGEST_REVIEW":
        # 非待评审状态，显示为'-'或'N/A'
        return ("-", "-", "-", "")
    
    # 检查是否有GPT评分结果
    if case_id in gpt_scores:
        score_data = gpt_scores[case_id]
        dedup = str(score_data.get("dedup", "-"))
        ranking = str(score_data.get("ranking", "-"))
        relevance = str(score_data.get("relevance", "-"))
        source = "GPT评分"
        return (dedup, ranking, relevance, source)
    
    # 待评审状态，显示具体维度
    dedup_score = "待评审" if "dedup" in review_items else "-"
    ranking_score = "待评审" if "ranking" in review_items else "-"
    filter_score = "待评审" if "filter_precision" in review_items else "-"
    
    return (dedup_score, ranking_score, filter_score, "")


def _generate_natural_failure_reason(case_result: Dict[str, Any]) -> tuple[str, str]:
    """
    生成自然语言的失败原因和修复建议
    返回: (失败原因, 修复建议)
    """
    status = case_result.get("status", "")
    failure_reason = case_result.get("failure_reason", "")
    suggestion = case_result.get("suggestion", "")
    checks = case_result.get("checks", [])
    
    if status == "PASS":
        return "通过", "无需修复"
    
    if status == "SUGGEST_REVIEW":
        return "需人工评审质量维度（去重/排序/相关性）", "使用独立评审或人工抽样复核"
    
    # 处理异常类失败
    if "exception" in failure_reason.lower():
        if "SSLError" in failure_reason or "SSL" in failure_reason:
            return (
                "网络连接失败：SSL证书验证错误，无法连接到 Coze API 服务器",
                "1) 检查网络连接是否正常；2) 确认 COZE_API_TOKEN 有效；3) 若使用代理，检查代理配置；4) 可尝试重新运行测试"
            )
        if "timeout" in failure_reason.lower():
            return (
                "请求超时：工作流执行时间超过预期",
                "1) 检查工作流是否正常响应；2) 增加超时时间设置；3) 联系 Coze 技术支持"
            )
        return (
            f"运行时异常：{failure_reason[:100]}",
            suggestion or "检查工作流配置和参数，查看 Coze 控制台日志"
        )
    
    # 处理具体检查项失败
    failed_checks = [c for c in checks if not c.get("passed", True)]
    reasons = []
    suggestions = []
    
    for check in failed_checks:
        check_name = check.get("name", "")
        check_reason = check.get("reason", "")
        check_suggestion = check.get("suggestion", "")
        
        if check_name == "empty_or_friendly":
            reasons.append("当指定来源无匹配新闻时，模型未给出友好的空结果提示，而是返回了其他来源的新闻")
            suggestions.append("在 Prompt 中明确添加指令：当 source 参数有值但数据中没有匹配来源时，必须输出「暂无来自[source]的相关新闻」提示，不要返回其他来源的内容")
        elif check_name == "source_only":
            reasons.append("输出来源不符合约束：包含了非指定来源的新闻")
            suggestions.append("强化 Prompt：严格按 source 参数过滤，只返回指定来源的新闻")
        elif check_name == "limit_max":
            reasons.append(f"输出条数超限：{check_reason}")
            suggestions.append(check_suggestion or "强化 Prompt 中的条数限制指令")
        elif check_name == "language_en":
            reasons.append("语言约束未满足：未输出英文内容")
            suggestions.append("在 Prompt 中添加明确的语言指令，如「请用英文输出」")
        elif check_name == "timeliness":
            reasons.append("时效性约束未满足：包含了超过24小时的新闻")
            suggestions.append("强化 Prompt：只保留24小时内的新闻，过滤旧内容")
        elif check_name == "raw_query_constraints":
            reasons.append("原始查询约束未满足：未按 raw_query 中的要求过滤")
            suggestions.append("提升 raw_query 理解优先级，使其覆盖槽位参数")
        elif check_name == "format":
            reasons.append("输出格式不正确")
            suggestions.append("检查 Prompt 中的格式要求是否清晰，确保输出符合预期格式")
        elif check_name == "structure":
            reasons.append("输出结构不完整")
            suggestions.append("确保 Prompt 要求输出完整的字段（标题、时间、摘要、来源等）")
        elif check_name == "non_empty":
            reasons.append("输出为空")
            suggestions.append("检查输入数据是否正常，Prompt 是否正确触发")
        else:
            reasons.append(f"检查项「{_translate_check_name(check_name)}」未通过：{check_reason}")
            suggestions.append(check_suggestion or "检查相关配置")
    
    if not reasons:
        reasons.append(failure_reason or "未知错误")
    if not suggestions:
        suggestions.append(suggestion or "检查测试配置和 Prompt 设计")
    
    return "；".join(reasons), "；".join(suggestions)


def main() -> int:
    if not SUITE_PATH.exists():
        raise SystemExit(f"缺少测试套件 JSON: {SUITE_PATH}")
    if not REPORT_PATH.exists():
        raise SystemExit(f"缺少测试报告 JSON: {REPORT_PATH}")
    
    suite = _load_json(SUITE_PATH)
    report = _load_json(REPORT_PATH)
    gpt_scores = _load_gpt_scores()  # 加载GPT评分结果
    
    suite_cases_by_id: Dict[str, Dict[str, Any]] = {}
    for c in suite.get("cases", []) or []:
        cid = _fmt(c.get("id"))
        if not cid:
            continue
        suite_cases_by_id[cid] = c
    
    wb = Workbook()
    
    # 定义样式
    pass_fill = PatternFill("solid", fgColor="DCFCE7")  # 浅绿色
    fail_fill = PatternFill("solid", fgColor="FEE2E2")  # 浅红色
    review_fill = PatternFill("solid", fgColor="FEF9C3")  # 浅黄色
    wrap = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== Sheet 1: 测评维度说明 ==========
    ws_guide = wb.active
    ws_guide.title = "测评维度说明"
    
    guide_content = [
        ["", ""],
        ["一、硬性规则检查（自动化）", ""],
        ["", ""],
        ["检查项", "说明"],
        ["输出非空", "模型输出不能为空"],
        ["格式正确", "输出格式符合预期（Markdown格式、结构化）"],
        ["结构完整", "包含标题、时间、摘要、来源等必要字段"],
        ["条数限制", "输出条数不超过 limit 参数指定值"],
        ["来源约束", "只返回 source 参数指定来源的新闻"],
        ["英文输出", "language=English 时输出英文内容"],
        ["时效性", "只保留24小时内的新闻（相对于 current_time）"],
        ["原始查询约束", "优先遵循 raw_query 中的约束（来源、条数、语言）"],
        ["空结果友好提示", "无匹配时给出友好提示，不返回无关内容"],
        ["", ""],
        ["评分策略：", "以上检查项全部通过 = 硬性规则通过"],
        ["", ""],
        ["二、质量维度评估（人工/GPT评审）", ""],
        ["", ""],
        ["维度", "说明"],
        ["去重质量", "相同事件是否被正确合并，无重复报道"],
        ["排序质量", "热点/重要新闻是否排在前面"],
        ["相关性", "返回的新闻是否与查询主题高度相关"],
        ["", ""],
        ["评分策略：", "0-10分制，≥7分视为通过"],
        ["评审方式：", "GPT自动评分（优先）或人工抽样评审"],
        ["", ""],
        ["三、综合判定", ""],
        ["", ""],
        ["判定结果", "条件"],
        ["通过", "硬性规则全部通过 AND 质量维度全部≥7分"],
        ["失败", "任一硬性规则未通过"],
        ["待评审", "硬性规则通过但质量维度需人工确认"],
    ]
    
    for row in guide_content:
        ws_guide.append(row)
    
    # 设置标题样式
    ws_guide['A2'].font = Font(bold=True, size=12, color="1F2937")
    ws_guide['A17'].font = Font(bold=True, size=12, color="1F2937")
    ws_guide['A27'].font = Font(bold=True, size=12, color="1F2937")
    
    # 设置表头样式
    header_rows = [4, 19, 28]
    for row_idx in header_rows:
        for col in ['A', 'B']:
            cell = ws_guide[f'{col}{row_idx}']
            cell.fill = PatternFill("solid", fgColor="374151")
            cell.font = Font(color="FFFFFF", bold=True)
    
    ws_guide.column_dimensions["A"].width = 20
    ws_guide.column_dimensions["B"].width = 70
    
    # ========== Sheet 2: 测试结果总览 ==========
    ws_overview = wb.create_sheet("测试结果总览")
    
    suite_meta = suite.get("suite") or {}
    run_meta = report.get("run") or {}
    summary = report.get("summary") or {}
    
    # 构建总览行
    overview_rows: List[List[Any]] = []
    for r in report.get("results", []) or []:
        case_id = _fmt(r.get("id"))
        suite_case = suite_cases_by_id.get(case_id, {})
        r["suite_case"] = suite_case  # 保存引用供后续使用
        
        # 用例名称：优先用中文，否则用英文
        case_name_zh = _fmt(suite_case.get("name_zh"))
        case_name = case_name_zh or _fmt(r.get("name")) or _fmt(suite_case.get("name"))
        
        # 测试目的
        purpose = _fmt(suite_case.get("purpose"))
        if not purpose:
            # 自动生成目的描述
            params = r.get("params", {})
            dataset = r.get("dataset", "")
            if "limit" in str(suite_case.get("name", "")).lower():
                purpose = f"验证条数限制功能（limit={params.get('limit', 'N/A')}）"
            elif "source" in str(suite_case.get("name", "")).lower():
                purpose = f"验证来源过滤功能（source={params.get('source', 'N/A')}）"
            elif "language" in str(suite_case.get("name", "")).lower():
                purpose = "验证语言约束功能"
            elif "raw_query" in str(suite_case.get("name", "")).lower():
                purpose = "验证自然语言查询理解"
            elif "timeliness" in str(suite_case.get("name", "")).lower():
                purpose = "验证24小时时效性过滤"
            elif "dedup" in str(suite_case.get("name", "")).lower():
                purpose = "验证去重能力"
            elif "ranking" in str(suite_case.get("name", "")).lower():
                purpose = "验证排序能力"
            else:
                purpose = "基础功能验证"
        
        # 测试数据
        dataset_info = _fmt(r.get("dataset"))
        
        # 关键参数
        params = r.get("params", {})
        key_params = []
        if params.get("keyword"):
            key_params.append(f"关键词={params.get('keyword')}")
        if params.get("source"):
            key_params.append(f"来源={params.get('source')}")
        if params.get("limit"):
            key_params.append(f"条数={params.get('limit')}")
        if params.get("language"):
            key_params.append(f"语言={params.get('language')}")
        if params.get("raw_query"):
            key_params.append(f"查询=「{params.get('raw_query')[:20]}...」" if len(params.get("raw_query", "")) > 20 else f"查询=「{params.get('raw_query')}」")
        key_params_str = "，".join(key_params) if key_params else "默认参数"
        
        # 结果状态
        status = r.get("status", "")
        if status == "PASS":
            result_label = "✓ 通过"
            problem_status = "正常"
        elif status == "FAIL":
            result_label = "✗ 失败"
            problem_status = "待修复"
        elif status == "SUGGEST_REVIEW":
            result_label = "? 待评审"
            problem_status = "需人工复核"
        else:
            result_label = status
            problem_status = "未知"
        
        # 自然语言的失败原因和修复建议
        failure_reason, fix_suggestion = _generate_natural_failure_reason(r)
        
        # 响应时间
        duration_ms = r.get("duration_ms", 0)
        
        # 硬性规则检查结果
        checks = r.get("checks", [])
        hard_rules_status = []
        for rule in ["non_empty", "format", "structure", "limit_max", "source_only", 
                     "language_en", "timeliness", "raw_query_constraints", "empty_or_friendly"]:
            status_text = _get_hard_rule_status(checks, rule)
            hard_rules_status.append(status_text)
        
        # 硬性规则通过数
        hard_passed = sum(1 for s in hard_rules_status if s == "通过")
        hard_total = len([s for s in hard_rules_status if s != "未检查"])
        hard_score = f"{hard_passed}/{hard_total}" if hard_total > 0 else "-"
        
        # 质量维度评分
        dedup_score, ranking_score, filter_score, score_source = _get_quality_score(r, gpt_scores)
        
        # 综合得分（预留，当前只统计硬性规则）
        overall_score = hard_score
        
        overview_rows.append([
            case_id,
            case_name,
            purpose,
            dataset_info,
            key_params_str,
            result_label,
            problem_status,
            hard_score,  # 硬性规则得分
            dedup_score,  # 去重评分
            ranking_score,  # 排序评分
            filter_score,  # 相关性评分
            overall_score,  # 综合得分
            score_source,  # 评分来源（GPT评分/待评审）
            failure_reason,
            fix_suggestion,
            duration_ms,
            round(duration_ms / 1000.0, 2) if duration_ms else 0,
        ])
    
    headers = [
        "用例ID", "用例名称", "测试目的", "测试数据", "关键参数",
        "测试结果", "问题状态", "硬性规则得分",
        "去重评分", "排序评分", "相关性评分", "综合得分", "评分来源",
        "失败原因", "修复建议", "响应时间(ms)", "响应时间(s)"
    ]
    
    _write_table(ws_overview, headers, overview_rows)
    
    # 应用条件格式和样式
    for row in ws_overview.iter_rows(min_row=2):
        # 失败原因和修复建议列自动换行
        row[12].alignment = wrap  # 失败原因
        row[13].alignment = wrap  # 修复建议
        
        # 根据结果着色
        result_cell = row[5]  # 测试结果列
        if "通过" in str(result_cell.value):
            result_cell.fill = pass_fill
            row[6].fill = pass_fill  # 问题状态
        elif "失败" in str(result_cell.value):
            result_cell.fill = fail_fill
            row[6].fill = fail_fill
        elif "待评审" in str(result_cell.value):
            result_cell.fill = review_fill
            row[6].fill = review_fill
    
    # 固定列宽
    ws_overview.column_dimensions["A"].width = 10
    ws_overview.column_dimensions["B"].width = 28
    ws_overview.column_dimensions["C"].width = 35
    ws_overview.column_dimensions["D"].width = 14
    ws_overview.column_dimensions["E"].width = 40
    ws_overview.column_dimensions["F"].width = 10
    ws_overview.column_dimensions["G"].width = 12
    ws_overview.column_dimensions["H"].width = 14  # 硬性规则得分
    ws_overview.column_dimensions["I"].width = 10  # 去重评分
    ws_overview.column_dimensions["J"].width = 10  # 排序评分
    ws_overview.column_dimensions["K"].width = 10  # 相关性评分
    ws_overview.column_dimensions["L"].width = 12  # 综合得分
    ws_overview.column_dimensions["M"].width = 12  # 评分来源
    ws_overview.column_dimensions["N"].width = 50  # 失败原因
    ws_overview.column_dimensions["O"].width = 50  # 修复建议
    ws_overview.column_dimensions["P"].width = 14
    ws_overview.column_dimensions["Q"].width = 12
    
    # ========== Sheet 3: 硬性规则详情 ==========
    ws_checks = wb.create_sheet("硬性规则详情")
    
    check_rows: List[List[Any]] = []
    for r in report.get("results", []) or []:
        case_id = _fmt(r.get("id"))
        suite_case = suite_cases_by_id.get(case_id, {})
        case_name = _fmt(suite_case.get("name_zh")) or _fmt(r.get("name")) or _fmt(suite_case.get("name"))
        
        checks = r.get("checks", [])
        
        # 为每个硬性规则检查项创建一行
        for check in checks:
            check_name = check.get("name", "")
            if check_name == "exception":
                continue  # 异常单独处理
            
            passed = check.get("passed", True)
            reason = check.get("reason", "")
            suggestion = check.get("suggestion", "")
            
            check_rows.append([
                case_id,
                case_name,
                _translate_check_name(check_name),
                "通过" if passed else "失败",
                reason if not passed else "",
                suggestion if not passed else "",
            ])
    
    if check_rows:
        _write_table(
            ws_checks,
            ["用例ID", "用例名称", "检查项", "结果", "失败原因", "修复建议"],
            check_rows,
        )
        
        # 设置样式
        for row in ws_checks.iter_rows(min_row=2):
            result_cell = row[3]
            if result_cell.value == "通过":
                result_cell.fill = pass_fill
            elif result_cell.value == "失败":
                result_cell.fill = fail_fill
            row[4].alignment = wrap
            row[5].alignment = wrap
        
        ws_checks.column_dimensions["A"].width = 10
        ws_checks.column_dimensions["B"].width = 30
        ws_checks.column_dimensions["C"].width = 18
        ws_checks.column_dimensions["D"].width = 10
        ws_checks.column_dimensions["E"].width = 50
        ws_checks.column_dimensions["F"].width = 50
    else:
        ws_checks.append(["暂无硬性规则检查记录"])
    
    # ========== Sheet 4: 失败用例详情 ==========
    ws_failures = wb.create_sheet("失败用例详情")
    
    failure_rows: List[List[Any]] = []
    for r in report.get("results", []) or []:
        status = r.get("status", "")
        if status != "FAIL":
            continue
        
        case_id = _fmt(r.get("id"))
        suite_case = suite_cases_by_id.get(case_id, {})
        case_name = _fmt(suite_case.get("name_zh")) or _fmt(r.get("name")) or _fmt(suite_case.get("name"))
        
        failure_reason, fix_suggestion = _generate_natural_failure_reason(r)
        
        # 获取详细检查信息
        checks = r.get("checks", [])
        failed_checks_str = "，".join([
            f"{_translate_check_name(c.get('name', ''))}"
            for c in checks if not c.get("passed", True)
        ])
        
        # 硬性规则得分
        hard_rules_status = []
        for rule in ["non_empty", "format", "structure", "limit_max", "source_only", 
                     "language_en", "timeliness", "raw_query_constraints", "empty_or_friendly"]:
            status_text = _get_hard_rule_status(checks, rule)
            hard_rules_status.append(status_text)
        hard_passed = sum(1 for s in hard_rules_status if s == "通过")
        hard_total = len([s for s in hard_rules_status if s != "未检查"])
        hard_score = f"{hard_passed}/{hard_total}"
        
        params = r.get("params", {})
        params_str = json.dumps(params, ensure_ascii=False, indent=2)
        
        output_preview = _fmt(r.get("output", ""))[:500] + "..." if len(_fmt(r.get("output", ""))) > 500 else _fmt(r.get("output", ""))
        
        failure_rows.append([
            case_id,
            case_name,
            hard_score,
            failed_checks_str,
            failure_reason,
            fix_suggestion,
            "待修复",
            params_str,
            output_preview,
        ])
    
    if failure_rows:
        _write_table(
            ws_failures,
            [
                "用例ID", "用例名称", "硬性规则得分", "失败检查项",
                "失败原因（自然语言）", "修复建议", "修复状态",
                "请求参数", "输出预览"
            ],
            failure_rows,
        )
        
        # 设置样式
        for row in ws_failures.iter_rows(min_row=2):
            row[4].alignment = wrap  # 失败原因
            row[5].alignment = wrap  # 修复建议
            row[7].alignment = wrap  # 请求参数
            row[8].alignment = wrap  # 输出预览
        
        ws_failures.column_dimensions["A"].width = 10
        ws_failures.column_dimensions["B"].width = 28
        ws_failures.column_dimensions["C"].width = 14
        ws_failures.column_dimensions["D"].width = 20
        ws_failures.column_dimensions["E"].width = 50
        ws_failures.column_dimensions["F"].width = 50
        ws_failures.column_dimensions["G"].width = 12
        ws_failures.column_dimensions["H"].width = 40
        ws_failures.column_dimensions["I"].width = 50
    else:
        ws_failures.append(["暂无失败用例"])
    
    # ========== Sheet 5: 待评审用例（含质量维度） ==========
    ws_review = wb.create_sheet("待评审用例")
    
    review_rows: List[List[Any]] = []
    for r in report.get("results", []) or []:
        status = r.get("status", "")
        if status != "SUGGEST_REVIEW":
            continue
        
        case_id = _fmt(r.get("id"))
        suite_case = suite_cases_by_id.get(case_id, {})
        case_name = _fmt(suite_case.get("name_zh")) or _fmt(r.get("name")) or _fmt(suite_case.get("name"))
        
        expect = suite_case.get("expect", {})
        review_items = expect.get("review_items", [])
        
        # 硬性规则得分（待评审用例硬性规则已通过）
        hard_score = "9/9"
        
        # 检查是否有GPT评分
        if case_id in gpt_scores:
            score_data = gpt_scores[case_id]
            dedup_status = f"{score_data.get('dedup', '-')}分"
            ranking_status = f"{score_data.get('ranking', '-')}分"
            filter_status = f"{score_data.get('relevance', '-')}分"
            review_state = "GPT评分" if score_data.get("passed") else "GPT评分-未通过"
            review_method = score_data.get("reasoning", "")[:200]
        else:
            # 待评审维度
            dedup_status = "待评审" if "dedup" in review_items else "-"
            ranking_status = "待评审" if "ranking" in review_items else "-"
            filter_status = "待评审" if "filter_precision" in review_items else "-"
            review_state = "待评审"
            review_method = "人工抽样检查3-5条输出，按0-10分评分，≥7分通过"
        
        # 评审标准
        review_criteria = []
        if "dedup" in review_items:
            review_criteria.append("去重：相同事件应合并为1条")
        if "ranking" in review_items:
            review_criteria.append("排序：高热度/重要新闻排前面")
        if "filter_precision" in review_items:
            review_criteria.append("相关性：与查询主题高度相关")
        
        output_preview = _fmt(r.get("output", ""))[:800] + "..." if len(_fmt(r.get("output", ""))) > 800 else _fmt(r.get("output", ""))
        
        review_rows.append([
            case_id,
            case_name,
            hard_score,
            dedup_status,
            ranking_status,
            filter_status,
            "；".join(review_criteria),
            review_method,
            review_state,
            output_preview,
        ])
    
    if review_rows:
        _write_table(
            ws_review,
            [
                "用例ID", "用例名称", "硬性规则得分",
                "去重评分", "排序评分", "相关性评分",
                "评审标准", "评审方法", "评审状态",
                "输出内容"
            ],
            review_rows,
        )
        
        for row in ws_review.iter_rows(min_row=2):
            row[6].alignment = wrap  # 评审标准
            row[7].alignment = wrap  # 评审方法
            row[9].alignment = wrap  # 输出内容
        
        ws_review.column_dimensions["A"].width = 10
        ws_review.column_dimensions["B"].width = 28
        ws_review.column_dimensions["C"].width = 14
        ws_review.column_dimensions["D"].width = 10
        ws_review.column_dimensions["E"].width = 10
        ws_review.column_dimensions["F"].width = 10
        ws_review.column_dimensions["G"].width = 45
        ws_review.column_dimensions["H"].width = 40
        ws_review.column_dimensions["I"].width = 10
        ws_review.column_dimensions["J"].width = 70
    else:
        ws_review.append(["暂无待评审用例"])
    
    # ========== Sheet 6: 测试摘要 ==========
    ws_summary = wb.create_sheet("测试摘要")
    
    summary_rows = [
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["工作流名称", suite_meta.get("name", "instant_news_generation_node")],
        ["工作流ID", suite_meta.get("workflow_id", "")],
        ["套件版本", suite_meta.get("version", "")],
        ["运行ID", run_meta.get("run_id", "")],
        ["运行时间", run_meta.get("timestamp", "")],
        ["Coze地址", run_meta.get("base_url", "")],
        ["", ""],
        ["统计项", "数值"],
        ["总用例数", summary.get("total", "")],
        ["通过数", summary.get("passed", "")],
        ["失败数", summary.get("failed", "")],
        ["待评审数", summary.get("suggested_review", "")],
        ["硬性规则通过率", f"{summary.get('passed', 0)}/{summary.get('total', 0)}"],
    ]
    
    for row in summary_rows:
        ws_summary.append(row)
    
    # 设置样式
    ws_summary['A9'].fill = PatternFill("solid", fgColor="374151")
    ws_summary['A9'].font = Font(color="FFFFFF", bold=True)
    ws_summary['B9'].fill = PatternFill("solid", fgColor="374151")
    ws_summary['B9'].font = Font(color="FFFFFF", bold=True)
    
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 50
    
    # ========== Sheet 7: 测试数据 ==========
    ws_data = wb.create_sheet("测试数据")
    
    datasets = report.get("datasets", {})
    slim_12 = datasets.get("slim_12", [])
    
    data_rows: List[List[Any]] = []
    for i, item in enumerate(slim_12, 1):
        data_rows.append([
            i,
            item.get("title", ""),
            item.get("source", ""),
            item.get("time_code", ""),
            item.get("link", ""),
            item.get("description", "")[:200] + "..." if len(item.get("description", "")) > 200 else item.get("description", ""),
        ])
    
    if data_rows:
        _write_table(
            ws_data,
            ["序号", "标题", "来源", "时间", "链接", "摘要"],
            data_rows,
        )
        
        for row in ws_data.iter_rows(min_row=2):
            row[1].alignment = wrap  # 标题
            row[5].alignment = wrap  # 摘要
        
        ws_data.column_dimensions["A"].width = 6
        ws_data.column_dimensions["B"].width = 50
        ws_data.column_dimensions["C"].width = 12
        ws_data.column_dimensions["D"].width = 16
        ws_data.column_dimensions["E"].width = 50
        ws_data.column_dimensions["F"].width = 60
    
    # ========== Sheet 8: Mock数据 ==========
    ws_mock = wb.create_sheet("Mock数据")
    
    mock_meta = datasets.get("mock_meta", [])
    mock_rows: List[List[Any]] = []
    for meta in mock_meta:
        mock_rows.append([
            meta.get("dataset", ""),
            meta.get("purpose", ""),
            meta.get("count", ""),
            meta.get("current_time", ""),
        ])
    
    if mock_rows:
        _write_table(
            ws_mock,
            ["数据集名称", "用途", "条数", "基准时间"],
            mock_rows,
        )
    
    # 保存文件
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUTPUT_PATH)
        print(f"[OK] Excel 报告已生成: {OUTPUT_PATH}")
    except PermissionError:
        run_id = _fmt(run_meta.get("run_id")) or datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = OUTPUT_PATH.with_name(f"instant_news_workflow_data_test_{run_id}.xlsx")
        wb.save(fallback)
        print(f"[WARN] 无法写入 {OUTPUT_PATH}（文件可能正在打开）")
        print(f"[OK] 已保存到: {fallback}")
    
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
