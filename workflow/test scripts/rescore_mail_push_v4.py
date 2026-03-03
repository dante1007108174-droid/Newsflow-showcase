import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "docs" / "生成层模型选型邮件推送" / "生成层模型选型_邮件推送_V3.xlsx"
# Writing to a new file avoids Windows file locks if V4 is open in Excel.
OUT = ROOT / "docs" / "生成层模型选型邮件推送" / "生成层模型选型_邮件推送_V4_中文.xlsx"


def safe_json_load(text: object):
    if not isinstance(text, str):
        return None, "not_str"
    raw = text.strip()
    has_fence = "```" in raw
    cleaned = re.sub(r"^```json\s*", "", raw, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```\s*$", "", cleaned)
    try:
        return json.loads(cleaned), ("fence" if has_fence else "")
    except Exception:
        return None, "parse_error"


def normalize_llm_payload(raw_text: object):
    """Best-effort extract {top_news, quick_news} JSON from raw payload.

    In this project, the LLM node output is sometimes wrapped like:
      {"output": "{...json...}"}
    or even multiple layers.
    """

    obj, meta = safe_json_load(raw_text)
    notes = []
    if meta == "fence":
        notes.append("has_markdown_fence")

    def try_extract(o):
        if isinstance(o, dict) and isinstance(o.get("top_news"), list) and isinstance(o.get("quick_news"), list):
            return o
        # Coze may wrap as {"output": "..."}
        if isinstance(o, dict) and "output" in o:
            inner = o.get("output")
            # output could be dict
            if isinstance(inner, dict):
                notes.append("wrapped_output_dict")
                return try_extract(inner)
            # output could be JSON string
            if isinstance(inner, str):
                notes.append("wrapped_output_str")
                inner_obj, inner_meta = safe_json_load(inner)
                if inner_meta == "fence":
                    notes.append("inner_has_markdown_fence")
                return try_extract(inner_obj)
        # Rare: JSON string as the value
        if isinstance(o, str):
            inner_obj, inner_meta = safe_json_load(o)
            if inner_meta == "fence":
                notes.append("inner_has_markdown_fence")
            return try_extract(inner_obj)
        return None

    extracted = try_extract(obj)
    if extracted is None:
        if meta == "parse_error":
            notes.append("json_parse_error")
        else:
            notes.append("missing_top_quick")
    return extracted, notes


def parse_dt(s: object):
    if not isinstance(s, str):
        return None
    s = s.strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None


def bucket_by_rate(rate: float, table: list[tuple[float, int]]):
    for min_rate, score in table:
        if rate >= min_rate:
            return score
    return table[-1][1]


def score_json(parsed, meta: str) -> int:
    # This dimension is about whether we can recover valid JSON for downstream use.
    return 10 if parsed is not None else 0


def score_structure(parsed) -> int:
    if not isinstance(parsed, dict):
        return 0
    top = parsed.get("top_news")
    quick = parsed.get("quick_news")
    if not isinstance(top, list) or not isinstance(quick, list):
        return 0

    top_req = ["title", "tag", "time", "summary", "source", "link"]
    q_req = ["title", "time", "source", "link"]

    def missing_ratio(items: list, req: list[str]) -> float:
        if not items:
            return 0.0
        miss = 0
        total = len(items) * len(req)
        for it in items:
            if not isinstance(it, dict):
                miss += len(req)
                continue
            for k in req:
                v = it.get(k)
                if v is None or (isinstance(v, str) and not v.strip()):
                    miss += 1
        return miss / total if total else 0.0

    r = max(missing_ratio(top, top_req), missing_ratio(quick, q_req))

    if r == 0:
        base = 10
    elif r <= 0.05:
        base = 8
    elif r <= 0.15:
        base = 6
    elif r <= 0.30:
        base = 4
    else:
        base = 2

    qty_pen = 0
    if len(top) > 8:
        qty_pen += 1
    if len(quick) > 10:
        qty_pen += 1

    return max(0, base - qty_pen)


def score_relevance(parsed, topic: str) -> int:
    if not isinstance(parsed, dict):
        return 0
    top = parsed.get("top_news")
    if not isinstance(top, list) or not top:
        return 4

    tokens = {
        "ai": [
            "ai",
            "人工智能",
            "大模型",
            "llm",
            "gpt",
            "openai",
            "anthropic",
            "deepseek",
            "gemini",
            "chatgpt",
            "agent",
            "智能体",
            "模型",
            "多模态",
        ],
        "tech": [
            "芯片",
            "半导体",
            "手机",
            "硬件",
            "操作系统",
            "os",
            "iphone",
            "android",
            "华为",
            "苹果",
            "小米",
            "三星",
            "英伟达",
            "nvidia",
            "amd",
            "intel",
            "gpu",
            "服务器",
            "电动车",
            "特斯拉",
        ],
        "finance": [
            "股",
            "a股",
            "港股",
            "美股",
            "央行",
            "降息",
            "利率",
            "cpi",
            "gdp",
            "ipo",
            "融资",
            "并购",
            "财报",
            "营收",
            "利润",
            "汇率",
            "美元",
            "人民币",
            "通胀",
            "基金",
            "债",
        ],
    }
    tset = tokens.get(topic, tokens["tech"])

    def hit(text: object) -> bool:
        s = str(text or "").lower()
        return any(tok in s for tok in tset)

    hits = 0
    total = 0
    for it in top:
        total += 1
        if not isinstance(it, dict):
            continue
        if hit(it.get("title")) or hit(it.get("summary")) or hit(it.get("tag")):
            hits += 1

    rate = hits / total if total else 0.0
    return bucket_by_rate(rate, [(0.9, 10), (0.75, 8), (0.6, 6), (0.4, 4), (0.0, 0)])


def score_sorting(parsed) -> int:
    if not isinstance(parsed, dict):
        return 0
    top = parsed.get("top_news")
    if not isinstance(top, list) or len(top) < 2:
        return 8

    times = [parse_dt(it.get("time")) if isinstance(it, dict) else None for it in top]
    if all(t is None for t in times):
        return 0

    inversions = 0
    pairs = 0
    for i in range(len(times) - 1):
        a, b = times[i], times[i + 1]
        if a is None or b is None:
            continue
        pairs += 1
        if a < b:
            inversions += 1

    if pairs == 0:
        return 4
    inv_rate = inversions / pairs

    if inv_rate == 0:
        return 10
    if inv_rate <= 0.25:
        return 8
    if inv_rate <= 0.5:
        return 6
    if inv_rate <= 0.8:
        return 4
    return 0


def score_tag(parsed) -> int:
    if not isinstance(parsed, dict):
        return 0
    top = parsed.get("top_news")
    if not isinstance(top, list) or not top:
        return 5

    banned_exact = {"ai", "人工智能", "科技", "技术", "财经", "经济"}

    bad = 0.0
    total = 0
    for it in top:
        total += 1
        if not isinstance(it, dict):
            bad += 1
            continue
        tag = str(it.get("tag") or "").strip()
        if not tag:
            bad += 1
            continue
        if tag.lower() in banned_exact:
            bad += 1
            continue
        if len(tag) < 2 or len(tag) > 10:
            bad += 0.5

    good_rate = 1 - (bad / total if total else 1)
    return bucket_by_rate(good_rate, [(0.9, 10), (0.75, 8), (0.6, 6), (0.4, 4), (0.0, 0)])


def score_summary(parsed) -> int:
    if not isinstance(parsed, dict):
        return 0
    top = parsed.get("top_news")
    if not isinstance(top, list) or not top:
        return 5

    def ok(s: object) -> bool:
        s = str(s or "").strip()
        if not s:
            return False
        if "暂无" in s or "N/A" in s:
            return False
        if len(s) < 80 or len(s) > 180:
            return False
        if s[-1] not in "。！？.!?":
            return False
        return True

    ok_n = 0
    total = 0
    for it in top:
        total += 1
        if not isinstance(it, dict):
            continue
        if ok(it.get("summary")):
            ok_n += 1

    rate = ok_n / total if total else 0.0
    return bucket_by_rate(rate, [(0.8, 10), (0.6, 8), (0.4, 6), (0.2, 4), (0.0, 0)])


def cost_score(model: str) -> int:
    return 100 if model == "DeepSeek V3.2" else 60


def topic_for_row(row_index: int) -> str:
    # V3 里 keyword 列出现乱码，按行号分组（每主题两行：DeepSeek+豆包）
    if row_index in (2, 3):
        return "ai"
    if row_index in (4, 5):
        return "tech"
    if row_index in (6, 7):
        return "finance"
    return "tech"


def main():
    wb_src = load_workbook(SRC)
    ws_src = wb_src[wb_src.sheetnames[0]]

    wb = Workbook()
    ws = wb.active
    ws.title = "评测记录"

    headers = [
        "Case ID",
        "keyword",
        "模型",
        "Agent提示词",
        "用户输入",
        "实际输出",
        "JSON格式(20%)",
        "结构完整(20%)",
        "相关性(20%)",
        "排序合理(20%)",
        "Tag质量(10%)",
        "摘要质量(10%)",
        "质量Raw(10分制)",
        "响应时间(s)",
        "质量得分(100)",
        "时效得分(100)",
        "成本得分(100)",
        "综合总分",
        "审计备注",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in range(2, ws_src.max_row + 1):
        case_id = ws_src.cell(row=r, column=1).value

        topic = topic_for_row(r)
        topic_label = {"ai": "AI", "tech": "科技", "finance": "财经"}.get(topic, "科技")

        model_raw = str(ws_src.cell(row=r, column=3).value or "")
        model = "DeepSeek V3.2" if "DeepSeek" in model_raw else "豆包 1.8"

        sys_prompt = ws_src.cell(row=r, column=4).value
        user_prompt = ws_src.cell(row=r, column=5).value
        output_raw = ws_src.cell(row=r, column=6).value
        resp_time = ws_src.cell(row=r, column=14).value

        extracted, notes = normalize_llm_payload(output_raw)

        s_json = score_json(extracted, "")
        s_struct = score_structure(extracted)
        s_rel = score_relevance(extracted, topic)
        s_sort = score_sorting(extracted)
        s_tag = score_tag(extracted)
        s_sum = score_summary(extracted)

        ws.append(
            [
                case_id,
                topic_label,
                model,
                sys_prompt,
                user_prompt,
                output_raw,
                s_json,
                s_struct,
                s_rel,
                s_sort,
                s_tag,
                s_sum,
                None,
                resp_time,
                None,
                None,
                cost_score(model),
                None,
                ";".join(notes),
            ]
        )

        rr = ws.max_row
        ws.cell(rr, 13, f"=G{rr}*0.2+H{rr}*0.2+I{rr}*0.2+J{rr}*0.2+K{rr}*0.1+L{rr}*0.1")
        ws.cell(rr, 15, f"=M{rr}*10")
        ws.cell(rr, 16, f"=MAX(0,MIN(100,(1-N{rr}/120)*100))")
        ws.cell(rr, 18, f"=O{rr}*0.7+Q{rr}*0.2+P{rr}*0.1")

    # widths
    widths = {
        "A": 10,
        "B": 8,
        "C": 14,
        "D": 18,
        "E": 18,
        "F": 35,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 14,
        "N": 12,
        "O": 12,
        "P": 12,
        "Q": 12,
        "R": 12,
        "S": 18,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # 汇总
    ws2 = wb.create_sheet("汇总对比")
    ws2.append(["模型", "质量均分(70%)", "成本均分(20%)", "时效均分(10%)", "最终加权总分", "推荐"])
    for c in ws2[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, m in enumerate(["DeepSeek V3.2", "豆包 1.8"], start=2):
        ws2.cell(i, 1, m)
        ws2.cell(i, 2, f"=AVERAGEIF(评测记录!C:C,\"{m}\",评测记录!O:O)")
        ws2.cell(i, 3, f"=AVERAGEIF(评测记录!C:C,\"{m}\",评测记录!Q:Q)")
        ws2.cell(i, 4, f"=AVERAGEIF(评测记录!C:C,\"{m}\",评测记录!P:P)")
        ws2.cell(i, 5, f"=B{i}*0.7+C{i}*0.2+D{i}*0.1")
    ws2.cell(2, 6, "=IF(E2>E3,\"YES\",\"\")")
    ws2.cell(3, 6, "=IF(E3>E2,\"YES\",\"\")")
    for col in "ABCDEF":
        ws2.column_dimensions[col].width = 18

    # 评分标准
    ws3 = wb.create_sheet("评分标准")
    ws3.append(["维度", "权重", "说明"])
    for c in ws3[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.append(["JSON格式", "20%", "能从payload中提取出{top_news,quick_news}结构=10；否则=0"])
    ws3.append([
        "结构完整",
        "20%",
        "top/quick必须存在且为数组；字段缺失率<=5%=8；<=15%=6；<=30%=4；否则=2；top>8或quick>10每项-1",
    ])
    ws3.append([
        "相关性",
        "20%",
        "用主题关键词表匹配标题/摘要/Tag，按命中率分桶：>=0.9=10, >=0.75=8, >=0.6=6, >=0.4=4, else=0",
    ])
    ws3.append([
        "排序合理",
        "20%",
        "只审计top_news时间是否按新->旧：无逆序=10；<=25%逆序=8；<=50%=6；<=80%=4；否则=0",
    ])
    ws3.append([
        "Tag质量",
        "10%",
        "Tag非空；禁用词(完全匹配)：AI/人工智能/科技/技术/财经/经济；长度2-10；按好Tag比例分桶",
    ])
    ws3.append([
        "摘要质量",
        "10%",
        "summary长度80-180且末尾标点；含“暂无”不合格；按合格比例分桶",
    ])
    ws3.append([
        "总权重",
        "-",
        "总分=质量(70%)+成本(20%)+时效(10%)；成本：DeepSeek=100，豆包=60；时效：MAX(0,MIN(100,(1-秒/120)*100))",
    ])
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 100
    for r in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=3):
        for c in r:
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    wb.save(OUT)


if __name__ == "__main__":
    main()
