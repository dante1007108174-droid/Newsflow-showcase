"""Export badcases grouped by similar root cause.

Reads:
  docs/badcase合集.xlsx (Sheet1)

Writes:
  docs/badcase_按根因合并.xlsx

Notes:
- Does NOT modify the original workbook.
- Groups "similar" root causes via a small rule-based classifier.
"""

from __future__ import annotations

import sys
from collections import defaultdict
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]


def _safe(s: Any) -> str:
    return str(s or "").strip()


def _classify_root(root: str) -> str:
    s = _safe(root)
    s_l = s.lower()

    # Instant news / LLM behavior
    if "来源" in s or "source" in s_l:
        if "过滤" in s or "指定" in s:
            return "来源过滤/来源约束"
        if "字段" in s:
            return "来源字段/格式约束"
        return "来源相关"

    if "格式" in s or "链接" in s or "时间" in s:
        return "输出格式模板"

    if "结构" in s or "热点" in s:
        return "输出结构模板"

    # Subscription/backend
    if "supabase" in s_l or "rpc" in s_l:
        return "Supabase/RPC"

    if "白名单" in s or "主题" in s:
        return "主题/白名单校验"

    return s or "(未填写)"


def _worst_status(statuses: List[str]) -> str:
    # Order: pending > solved > empty
    s = [x for x in (_safe(v) for v in statuses) if x]
    if any("待" in x or "🔴" in x for x in s):
        return "🔴 待修复"
    if any("已解决" in x or "✅" in x for x in s):
        return "✅ 已解决"
    return s[0] if s else ""


def _merge_lines(items: List[str], *, max_items: int = 12) -> str:
    uniq: List[str] = []
    for it in items:
        it = _safe(it)
        if not it:
            continue
        if it not in uniq:
            uniq.append(it)
    if len(uniq) > max_items:
        uniq = uniq[:max_items] + [f"...（省略{len(uniq) - max_items}条）"]
    return "\n".join(uniq)


def _extract_raw_query(desc: str) -> str:
    """Extract user input from description like '用户输入「xxx」'."""
    import re
    m = re.search(r'用户输入「([^」]+)」', desc)
    return m.group(1) if m else ""


def _extract_failed_rounds(desc: str) -> str:
    """Extract failure count like '失败2/3'."""
    import re
    m = re.search(r'失败(\d+/\d+)', desc)
    return m.group(1) if m else ""


def _extract_main_issue(desc: str) -> str:
    """Extract main issue after '主要问题：'."""
    import re
    m = re.search(r'主要问题：([^；]+)', desc)
    return m.group(1) if m else ""


def _generate_natural_desc(items: List[Dict[str, str]], group_key: str) -> str:
    """Generate a natural language summary for the group."""
    if not items:
        return ""
    
    # Collect unique inputs and issues
    inputs: List[str] = []
    issues: List[str] = []
    
    for item in items:
        desc = item.get("desc", "")
        inp = _extract_raw_query(desc)
        issue = _extract_main_issue(desc)
        if inp and inp not in inputs:
            inputs.append(inp)
        if issue and issue not in issues:
            issues.append(issue)
    
    # Build natural description based on group type
    if "来源过滤" in group_key or "来源约束" in group_key:
        base = "用户指定特定来源时，模型未能正确过滤，返回了非指定来源的内容。"
        if inputs:
            examples = "；".join(inputs[:3])
            base += f"例如：{examples}。"
        if issues:
            base += f"主要表现为：{'；'.join(issues[:2])}。"
        return base
    
    if "格式" in group_key or "字段" in group_key:
        base = "模型输出格式不稳定，有时缺少必要的时间或来源链接字段。"
        if inputs:
            examples = "；".join(inputs[:3])
            base += f"例如：{examples}。"
        if issues:
            base += f"主要问题：{'；'.join(issues[:2])}。"
        return base
    
    if "结构" in group_key:
        base = "模型输出结构不完整，缺少标题或【热点N】条目格式。"
        if inputs:
            examples = "；".join(inputs[:2])
            base += f"例如：{examples}。"
        return base
    
    # Default: simple concatenation of inputs and issues
    parts: List[str] = []
    if inputs:
        parts.append(f"涉及输入：{'；'.join(inputs[:3])}")
    if issues:
        parts.append(f"主要问题：{'；'.join(issues[:2])}")
    
    return "。".join(parts) if parts else "详见原始记录"


def _copy_header_style(src_ws, dst_ws, src_row: int, dst_row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        src = src_ws.cell(src_row, c)
        dst = dst_ws.cell(dst_row, c)
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def main() -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    src_path = ROOT / "docs" / "badcase合集.xlsx"
    if not src_path.exists():
        raise SystemExit(f"Not found: {src_path}")

    wb = load_workbook(src_path)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    # Header mapping (expects the current template naming).
    header_row = 1
    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = _safe(ws.cell(header_row, c).value)
        if v:
            headers[v] = c

    required = ["ID", "问题名称", "问题描述", "根因", "影响", "优化方式", "问题状态"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise SystemExit(f"Missing columns in Sheet1 header: {missing}")

    # Read all rows.
    items: List[Dict[str, str]] = []
    for r in range(2, ws.max_row + 1):
        rid = _safe(ws.cell(r, headers["ID"]).value)
        if not rid:
            continue
        items.append(
            {
                "id": rid,
                "name": _safe(ws.cell(r, headers["问题名称"]).value),
                "desc": _safe(ws.cell(r, headers["问题描述"]).value),
                "root": _safe(ws.cell(r, headers["根因"]).value),
                "impact": _safe(ws.cell(r, headers["影响"]).value),
                "fix": _safe(ws.cell(r, headers["优化方式"]).value),
                "status": _safe(ws.cell(r, headers["问题状态"]).value),
            }
        )

    # Group by classified root.
    groups: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for it in items:
        key = _classify_root(it.get("root", ""))
        groups[key].append(it)

    # Build output workbook.
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "按根因合并"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    out_headers = [
        "根因分类",
        "用例数",
        "ID列表",
        "问题名称列表",
        "问题描述(合并)",
        "影响(合并)",
        "优化方式(合并)",
        "问题状态(最差)",
    ]
    out_ws.append(out_headers)
    for c in out_ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
    out_ws.freeze_panes = "A2"

    # Deterministic order: pending groups first, then by size desc.
    def group_sort_key(k: str) -> Tuple[int, int, str]:
        g = groups[k]
        worst = _worst_status([x.get("status", "") for x in g])
        pending_first = 0 if ("待" in worst or "🔴" in worst) else 1
        return (pending_first, -len(g), k)

    for key in sorted(groups.keys(), key=group_sort_key):
        g = groups[key]
        ids = [x["id"] for x in g]
        names = [x.get("name", "") for x in g]
        worst = _worst_status([x.get("status", "") for x in g])

        row = [
            key,
            len(g),
            ", ".join(ids),
            _merge_lines(names, max_items=20),
            _generate_natural_desc(g, key),
            _merge_lines([x.get("impact", "") for x in g], max_items=20),
            _merge_lines([x.get("fix", "") for x in g], max_items=20),
            worst,
        ]
        out_ws.append(row)

    # Styling
    for row in out_ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    # Zebra rows
    zebra = PatternFill("solid", fgColor="F8FAFC")
    for i, row in enumerate(out_ws.iter_rows(min_row=2), start=2):
        if i % 2 == 0:
            for cell in row:
                if cell.fill is None or cell.fill.fill_type is None:
                    cell.fill = zebra

    # Status fills
    pass_fill = PatternFill("solid", fgColor="DCFCE7")
    fail_fill = PatternFill("solid", fgColor="FEE2E2")
    for row in out_ws.iter_rows(min_row=2):
        cell = row[7]
        v = _safe(cell.value)
        if "已解决" in v or "✅" in v:
            cell.fill = pass_fill
        elif "待" in v or "🔴" in v:
            cell.fill = fail_fill

    # Column widths
    widths = {
        "A": 18,
        "B": 10,
        "C": 28,
        "D": 30,
        "E": 60,
        "F": 40,
        "G": 40,
        "H": 14,
    }
    for col, w in widths.items():
        out_ws.column_dimensions[col].width = w

    dst_path = ROOT / "docs" / "badcase_按根因合并.xlsx"
    try:
        out_wb.save(dst_path)
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = dst_path.with_name(f"badcase_按根因合并_autosave_{ts}.xlsx")
        out_wb.save(fallback)
        dst_path = fallback

    print(f"Wrote: {dst_path}")
    print(f"Groups: {len(groups)}, Rows: {len(items)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
