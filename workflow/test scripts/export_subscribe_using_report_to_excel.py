"""Export subscribe_using workflow suite + latest run report to Excel.

Inputs:
- tests/subscribe_using_suite.json
- tests/reports/subscribe_using_report.json

Output:
- docs/subscribe_using_workflow_data_test.xlsx

This is a deterministic export step (no network calls).
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]

SUITE_PATH = ROOT / "tests" / "subscribe_using_suite.json"
REPORT_PATH = ROOT / "tests" / "reports" / "subscribe_using_report.json"
OUTPUT_PATH = ROOT / "docs" / "subscribe_using_workflow_data_test.xlsx"


def _load_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _fmt(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return str(value)
    return json.dumps(value, ensure_ascii=False)


def _write_table(ws, headers: List[str], rows: List[List[Any]]) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(vertical="top")
    ws.append(headers)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align

    for r in rows:
        ws.append(r)

    ws.freeze_panes = "A2"

    # Basic autosize with caps.
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for r in rows:
            if col - 1 >= len(r):
                continue
            s = str(r[col - 1] if r[col - 1] is not None else "")
            if len(s) > max_len:
                max_len = len(s)
        width = min(max(10, max_len + 2), 70)
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width


def main() -> int:
    if not SUITE_PATH.exists():
        raise SystemExit(f"Missing suite JSON: {SUITE_PATH}")
    if not REPORT_PATH.exists():
        raise SystemExit(f"Missing report JSON: {REPORT_PATH}")

    suite = _load_json(SUITE_PATH)
    report = _load_json(REPORT_PATH)

    suite_cases_by_id: Dict[str, Dict[str, Any]] = {}
    for c in suite.get("cases", []) or []:
        cid = _fmt(c.get("id"))
        if not cid:
            continue
        suite_cases_by_id[cid] = c

    wb = Workbook()

    # Make the first tab a human-friendly overview.
    ws_overview = wb.active
    ws_overview.title = "测试结果总览"

    suite_meta = suite.get("suite") or {}
    run_meta = report.get("run") or {}
    summary = report.get("summary") or {}

    # Overview: one row per case.
    overview_rows: List[List[Any]] = []
    for r in report.get("results", []) or []:
        case_id = _fmt(r.get("id"))
        suite_case = suite_cases_by_id.get(case_id, {})

        name_zh = _fmt(suite_case.get("name_zh"))
        case_name = name_zh or _fmt(r.get("name")) or _fmt(suite_case.get("name"))
        purpose = _fmt(suite_case.get("purpose"))

        # Use suite-defined step actions for readability (stable even if run output varies).
        defined_actions: List[str] = []
        for s in suite_case.get("steps", []) or []:
            a = _fmt(s.get("action"))
            if a:
                defined_actions.append(a)
        action_flow = " -> ".join(defined_actions)

        steps = r.get("steps", []) or []
        total_steps = len(steps)
        failed_steps = 0
        total_duration_ms = 0
        first_failure = ""

        for s in steps:
            passed = bool(s.get("passed"))
            total_duration_ms += int(s.get("duration_ms") or 0)
            if not passed:
                failed_steps += 1
                if not first_failure:
                    errs = s.get("errors") or []
                    err0 = _fmt(errs[0]) if errs else ""
                    first_failure = f"step={_fmt(s.get('step_index'))} action={_fmt(s.get('action'))}: {err0}".strip()

        case_passed = bool(r.get("passed"))
        known_fail = bool(r.get("known_fail"))
        if (not case_passed) and (not known_fail):
            # Back-compat if report.json came from older runner.
            known_fail = bool(suite_case.get("known_issue"))

        issue_status = _fmt(r.get("issue_status")) or _fmt(suite_case.get("issue_status"))
        issue_note = _fmt(r.get("issue_note")) or _fmt(suite_case.get("issue_note"))

        if case_passed:
            result_label = "PASS"
            problem_status = "正常"
            problem_note = ""
        elif known_fail:
            result_label = "KNOWN_FAIL"
            problem_status = issue_status or "待修复"
            problem_note = issue_note
        else:
            result_label = "FAIL"
            problem_status = "待排查"
            problem_note = ""

        overview_rows.append(
            [
                case_id,
                case_name,
                purpose,
                action_flow,
                result_label,
                problem_status,
                problem_note,
                total_steps,
                failed_steps,
                total_duration_ms,
                round(total_duration_ms / 1000.0, 3),
                first_failure,
            ]
        )

    _write_table(
        ws_overview,
        [
            "case_id",
            "用例名称",
            "测试目的",
            "操作步骤",
            "结果",
            "问题状态",
            "问题备注",
            "总步骤",
            "失败步骤",
            "总耗时_ms",
            "总耗时_s",
            "首个失败原因",
        ],
        overview_rows,
    )

    # Result coloring (make it scannable).
    pass_fill = PatternFill("solid", fgColor="DCFCE7")  # light green
    fail_fill = PatternFill("solid", fgColor="FEE2E2")  # light red
    known_fill = PatternFill("solid", fgColor="FEF9C3")  # light yellow
    wrap = Alignment(wrap_text=True, vertical="top")

    # Wrap long columns in overview.
    for row in ws_overview.iter_rows(min_row=2):
        # purpose, action_flow, first_failure
        row[2].alignment = wrap
        row[3].alignment = wrap
        row[6].alignment = wrap  # problem_note
        row[11].alignment = wrap  # first_failure

        # result coloring
        result_cell = row[4]
        if result_cell.value == "PASS":
            result_cell.fill = pass_fill
        elif result_cell.value == "KNOWN_FAIL":
            result_cell.fill = known_fill
        elif result_cell.value == "FAIL":
            result_cell.fill = fail_fill

    # Keep overview columns readable.
    ws_overview.column_dimensions["B"].width = 24
    ws_overview.column_dimensions["C"].width = 38
    ws_overview.column_dimensions["D"].width = 28
    ws_overview.column_dimensions["G"].width = 48
    ws_overview.column_dimensions["L"].width = 60

    ws_summary = wb.create_sheet("summary")

    summary_rows = [
        ["generated_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["workflow_name", suite_meta.get("name", "subscribe_using")],
        ["workflow_id", suite_meta.get("workflow_id", "")],
        ["suite_version", suite_meta.get("version", "")],
        ["run_id", run_meta.get("run_id", "")],
        ["run_timestamp", run_meta.get("timestamp", "")],
        ["coze_base_url", run_meta.get("base_url", "")],
        ["total_cases", summary.get("total_cases", "")],
        ["passed_cases", summary.get("passed_cases", "")],
        ["failed_cases", summary.get("failed_cases", "")],
        ["known_failed_cases", summary.get("known_failed_cases", "")],
        ["suite_json", str(SUITE_PATH)],
        ["report_json", str(REPORT_PATH)],
    ]

    _write_table(ws_summary, ["key", "value"], summary_rows)
    ws_summary.column_dimensions["A"].width = 18
    ws_summary.column_dimensions["B"].width = 90

    ws_suite = wb.create_sheet("suite_steps")
    suite_rows: List[List[Any]] = []
    for case in suite.get("cases", []) or []:
        case_id = _fmt(case.get("id"))
        case_name = _fmt(case.get("name"))
        for i, step in enumerate(case.get("steps", []) or []):
            suite_rows.append(
                [
                    case_id,
                    case_name,
                    i,
                    _fmt(step.get("action")),
                    _fmt(step.get("ex_email")),
                    _fmt(step.get("new_email")),
                    _fmt(step.get("new_keyword")),
                    _fmt(step.get("expect")),
                ]
            )

    _write_table(
        ws_suite,
        ["case_id", "case_name", "step_index", "action", "ex_email", "new_email", "new_keyword", "expect_json"],
        suite_rows,
    )

    ws_run = wb.create_sheet("run_steps")
    run_rows: List[List[Any]] = []
    for r in report.get("results", []) or []:
        ctx = r.get("context") or {}
        case_id = _fmt(r.get("id"))
        case_name = _fmt(r.get("name"))
        case_passed = r.get("passed")
        for s in r.get("steps", []) or []:
            run_rows.append(
                [
                    case_id,
                    case_name,
                    _fmt(ctx.get("user_id")),
                    _fmt(ctx.get("email")),
                    _fmt(ctx.get("email2")),
                    bool(case_passed),
                    s.get("step_index"),
                    _fmt(s.get("action")),
                    _fmt(s.get("params")),
                    s.get("duration_ms"),
                    bool(s.get("passed")),
                    _fmt(s.get("errors")),
                    _fmt(s.get("output")),
                ]
            )

    _write_table(
        ws_run,
        [
            "case_id",
            "case_name",
            "ctx_user_id",
            "ctx_email",
            "ctx_email2",
            "case_passed",
            "step_index",
            "action",
            "params_json",
            "duration_ms",
            "step_passed",
            "errors_json",
            "output",
        ],
        run_rows,
    )

    # Wrap large text columns
    for row in ws_suite.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.alignment = wrap

    for row in ws_run.iter_rows(min_row=2):
        row[8].alignment = wrap   # params_json
        row[11].alignment = wrap  # errors_json
        row[12].alignment = wrap  # output

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUTPUT_PATH)
    except PermissionError:
        # Common on Windows when the workbook is open in Excel.
        run_id = _fmt(run_meta.get("run_id")) or datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = OUTPUT_PATH.with_name(f"subscribe_using_workflow_data_test_{run_id}.xlsx")
        wb.save(fallback)
        print(f"Permission denied writing {OUTPUT_PATH}; wrote fallback: {fallback}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
