"""Append current instant-news badcases into docs/badcase合集.xlsx.

Rules:
- Preserve existing formatting.
- Merge duplicates (same ID) by appending new run info.

Input:
- tests/reports/instant_news_stability_report.json (default) or STABILITY_REPORT
"""

from __future__ import annotations

import json
import os
import sys
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def _safe_text(s: Any) -> str:
    return str(s or "").strip()


def _load_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _find_last_data_row(ws, id_col: int) -> int:
    last = 1
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, id_col).value
        if v is not None and _safe_text(v) != "":
            last = r
    return last


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def _append_unique_text(existing: str, addition: str) -> str:
    existing = _safe_text(existing)
    addition = _safe_text(addition)
    if not addition:
        return existing
    if not existing:
        return addition

    # De-dupe by exact substring match; keep original style using Chinese semicolons.
    if addition in existing:
        return existing
    sep = "；" if not existing.endswith("；") else ""
    return f"{existing}{sep}{addition}"


def _simplify_issue(issue: str) -> str:
    parts: List[str] = []
    for raw in _safe_text(issue).split("；"):
        s = raw.strip()
        if not s:
            continue
        # Drop per-round prefixes like "R2:".
        if len(s) >= 3 and s[0] == "R" and s[2] == ":" and s[1].isdigit():
            s = s[3:].strip()
        if s and s not in parts:
            parts.append(s)
    return "；".join(parts)


def _classify_root_and_fix(issue: str) -> tuple[str, str, str]:
    s = _safe_text(issue)
    if "包含非指定来源" in s or "无法识别来源字段" in s:
        root = "来源过滤/来源字段约束不稳定（LLM会混入其他来源或格式不规范）"
        impact = "用户指定来源时仍会返回其他来源，来源过滤能力失去信任"
        fix = (
            "提示词加固：用户指定来源时，仅输出该来源；每条必须含“来源：[来源名](链接)”。"
            "若过滤后为空，则输出空状态；否则重写直到满足约束。"
            "必要时在RSS/代码节点先做来源硬过滤再交给LLM格式化。"
        )
        return root, impact, fix

    if "缺少时间或来源链接" in s or "缺少时间" in s or "link=" in s:
        root = "输出格式模板不稳定（时间/来源链接字段偶发缺失）"
        impact = "输出不可解析/不可点击，降低可用性与可信度"
        fix = "提示词加固：输出前自检每条是否包含时间与来源链接；不满足则重写。"
        return root, impact, fix

    if "缺少标题或条目" in s or "items=False" in s:
        root = "输出结构模板不稳定（缺少【热点N】条目或结构不完整）"
        impact = "条目结构缺失导致可读性差，且后续自动解析/评估失败"
        fix = "提示词加固：必须包含标题+至少N条【热点】条目；不满足则重写。"
        return root, impact, fix

    root = "LLM输出不稳定（指令遵循/模板约束未强制）"
    impact = "同一输入多次运行结果不一致，影响用户体验与测试通过率"
    fix = "提示词加固：增加输出自检与重写策略；必要时降低温度并收紧约束。"
    return root, impact, fix


def _build_badcase_rows(stability_report: Dict[str, Any]) -> List[Dict[str, str]]:
    run = stability_report.get("run") or {}
    run_id = _safe_text(run.get("run_id"))
    rounds = int(run.get("rounds") or 3)

    out: List[Dict[str, str]] = []
    for r in stability_report.get("results", []) or []:
        passed_rounds = int(r.get("passed_rounds") or 0)
        if passed_rounds >= rounds:
            continue

        cid = _safe_text(r.get("id"))
        scenario = _safe_text(r.get("scenario"))
        params = r.get("params") or {}
        raw_query = _safe_text(params.get("raw_query"))
        keyword = _safe_text(params.get("keyword"))
        round_result = _safe_text(r.get("round_result"))
        failed_rounds = int(r.get("failed_rounds") or 0)

        issue_raw = _safe_text(r.get("issue"))
        issue = _simplify_issue(issue_raw)

        name = f"即时新闻：{scenario}" if scenario else f"即时新闻：{cid}"
        root, impact, optimize = _classify_root_and_fix(issue_raw)
        status = "🔴 待修复"

        # Keep it readable: mimic existing rows' narrative style; no debug_url in description.
        desc_bits: List[str] = []
        if raw_query:
            desc_bits.append(f"用户输入「{raw_query}」")
        if round_result:
            desc_bits.append(f"三轮结果：{round_result}")
        if failed_rounds:
            desc_bits.append(f"失败{failed_rounds}/3")
        if issue:
            desc_bits.append(f"主要问题：{issue}")
        if run_id:
            desc_bits.append(f"run_id={run_id}")
        desc = "；".join(desc_bits)

        out.append(
            {
                "ID": f"NEWS-{cid}" if cid else "",
                "名称": name,
                "问题描述": desc,
                "根因": root,
                "影响": impact,
                "优化方式": optimize,
                "问题状态": status,
            }
        )
    return out


def main() -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    stability_path = Path(
        os.getenv("STABILITY_REPORT")
        or (ROOT / "tests" / "reports" / "instant_news_stability_report.json")
    )
    if not stability_path.exists():
        raise SystemExit(f"Stability report not found: {stability_path}")
    stability_report = _load_json(stability_path)

    xlsx_path = ROOT / "docs" / "badcase合集.xlsx"
    if not xlsx_path.exists():
        raise SystemExit(f"Badcase collection not found: {xlsx_path}")

    wb = load_workbook(xlsx_path)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    # Map headers.
    header_row = 1
    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = _safe_text(ws.cell(header_row, c).value)
        if v:
            headers[v] = c

    required = ["ID", "名称", "问题描述", "根因", "影响", "优化方式", "问题状态"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise SystemExit(f"Sheet header missing columns: {missing}")

    id_col = headers["ID"]
    max_col = max(headers.values())
    last_row = _find_last_data_row(ws, id_col)
    template_row = last_row

    # Index existing IDs to row.
    id_to_row: Dict[str, int] = {}
    for r in range(2, last_row + 1):
        v = _safe_text(ws.cell(r, id_col).value)
        if v:
            id_to_row[v] = r

    rows = _build_badcase_rows(stability_report)
    appended = 0
    updated = 0
    for row in rows:
        rid = row.get("ID") or ""
        if not rid:
            continue
        if rid in id_to_row:
            r = id_to_row[rid]
            # Overwrite with concise, latest description (and keep formatting).
            ws.cell(r, headers["名称"]).value = row.get("名称", "")
            ws.cell(r, headers["问题描述"]).value = row.get("问题描述", "")
            ws.cell(r, headers["根因"]).value = row.get("根因", "")
            ws.cell(r, headers["影响"]).value = row.get("影响", "")
            ws.cell(r, headers["优化方式"]).value = row.get("优化方式", "")
            ws.cell(r, headers["问题状态"]).value = "🔴 待修复"
            updated += 1
            continue

        dst = last_row + 1
        ws.insert_rows(dst)
        _copy_row_style(ws, template_row, dst, max_col)
        for key, col in headers.items():
            if key in row:
                ws.cell(dst, col).value = row[key]
        last_row = dst
        id_to_row[rid] = dst
        appended += 1

    try:
        wb.save(xlsx_path)
        saved_to = xlsx_path
        locked = False
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = xlsx_path.with_name(f"badcase合集_autosave_{ts}.xlsx")
        wb.save(fallback)
        saved_to = fallback
        locked = True

    print(f"Updated badcase collection: {saved_to}")
    print(f"Appended: {appended}, Updated: {updated}")
    if locked:
        print(f"Note: {xlsx_path} was locked; wrote fallback instead.")
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
