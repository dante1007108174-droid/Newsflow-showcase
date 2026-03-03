import json
import re
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]

INSTANT_QUERY_DIR = "\u751f\u6210\u5c42\u6a21\u578b\u9009\u578b\u2014\u5373\u65f6\u67e5\u8be2"
SRC_NAME = "\u5373\u65f6\u67e5\u8be2-\u5185\u5bb9\u751f\u6210\u6a21\u578b\u6d4b\u8bd5.xlsx"

# Always write a new file to avoid Windows Excel file locks.
OUT_NAME = "\u751f\u6210\u5c42\u6a21\u578b\u9009\u578b_\u5373\u65f6\u67e5\u8be2_\u5185\u5bb9\u751f\u6210_V2.xlsx"

SRC = ROOT / "docs" / INSTANT_QUERY_DIR / SRC_NAME
OUT = ROOT / "docs" / INSTANT_QUERY_DIR / OUT_NAME


FIRE = "\U0001F525"
NEWS = "\U0001F4F0"


def parse_dt(s: object):
    if not isinstance(s, str):
        return None
    s = s.strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None


def parse_seconds(s: object):
    """Parse values like '24s', '1.14m', '1m14s', '72.3'."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    if not isinstance(s, str):
        return None
    raw = s.strip().lower()
    if not raw:
        return None

    m = re.fullmatch(r"(?P<m>\d+(?:\.\d+)?)m(?P<s>\d+(?:\.\d+)?)s", raw)
    if m:
        return float(m.group("m")) * 60.0 + float(m.group("s"))

    if raw.endswith("ms"):
        try:
            return float(raw[:-2]) / 1000.0
        except Exception:
            return None
    if raw.endswith("s"):
        try:
            return float(raw[:-1])
        except Exception:
            return None
    if raw.endswith("m"):
        try:
            return float(raw[:-1]) * 60.0
        except Exception:
            return None
    try:
        return float(raw)
    except Exception:
        return None


def safe_json_load(text: object):
    if not isinstance(text, str):
        return None
    try:
        return json.loads(text)
    except Exception:
        return None


def extract_current_time(user_prompt: object):
    if not isinstance(user_prompt, str):
        return ""
    idx = user_prompt.find('"news_list"')
    head = user_prompt if idx == -1 else user_prompt[:idx]
    m = re.search(r'"current_time"\s*:\s*"([^"]+)"', head)
    return m.group(1).strip() if m else ""


def extract_keyword(user_prompt: object):
    if not isinstance(user_prompt, str):
        return ""
    idx = user_prompt.find('"news_list"')
    head = user_prompt if idx == -1 else user_prompt[:idx]
    m = re.search(r'"keyword"\s*:\s*"([^"]+)"', head)
    return m.group(1).strip() if m else ""


def normalize_language(lang: object) -> str:
    if lang is None:
        return ""
    s = str(lang).strip()
    if not s:
        return ""
    if s.lower() in ("english", "en"):
        return "english"
    if s in ("\u82f1\u8bed",):
        return "english"
    if s in ("\u4e2d\u6587",):
        return "chinese"
    # Fallback: treat unknown as raw.
    return s.lower()


def normalize_source(src: object) -> str:
    if src is None:
        return ""
    s = str(src).strip()
    if not s or s in ("\u7a7a", "\u4e0d\u9650", "\u5168\u90e8"):
        return ""
    return s


def normalize_limit(limit: object) -> int:
    if limit is None:
        return 5
    s = str(limit).strip()
    if not s:
        return 5
    try:
        return int(float(s))
    except Exception:
        return 5


def scenario_label(language: str, limit: int, src: str) -> str:
    if src:
        return "source"
    if language == "english":
        return "english"
    # limit is always a number; consider it a scenario only when explicitly set in source workbook.
    # We treat non-default limits as a scenario.
    if limit != 5:
        return "limit"
    return "baseline"


ITEM_START_RE = re.compile(
    r"^(?P<fire>\U0001F525{0,3})?\s*\*\*\s*\u3010(?P<kind>\u70ed\u70b9|Hotspot)\s*(?P<idx>\d+)\u3011(?P<title>.+?)\*\*\s*$"
)


def parse_markdown_items(text: object):
    if not isinstance(text, str):
        return [], ["output_not_str"]

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    items = []
    notes = []

    i = 0
    while i < len(lines):
        m = ITEM_START_RE.match(lines[i])
        if not m:
            i += 1
            continue
        fire_prefix = m.group("fire") or ""
        kind = m.group("kind")
        idx = int(m.group("idx"))
        title = m.group("title").strip()
        block = []
        j = i + 1
        while j < len(lines) and not ITEM_START_RE.match(lines[j]):
            block.append(lines[j])
            j += 1

        items.append(
            {
                "idx": idx,
                "kind": kind,
                "title": title,
                "fire_prefix": fire_prefix,
                "block": block,
            }
        )
        i = j

    if not items:
        notes.append("no_items_found")
    return items, notes


def parse_item_fields(item: dict):
    """Extract time/summary/source/heat from an item block."""
    block = "\n".join(item.get("block") or [])

    # Support both Chinese and English labels.
    time_m = re.search(r"(?:\u65f6\u95f4|Time)\s*[:\uff1a]\s*([0-9\-: ]{10,19})", block)
    summary_m = re.search(r"(?:\u6458\u8981|Summary)\s*[:\uff1a]\s*(.+)", block)

    # Heat: Chinese: '3家媒体报道 | 跨3类媒体'
    heat_m_cn = re.search(r"\u70ed\u5ea6\s*[:\uff1a]\s*(?P<n>\d+)\u5bb6\u5a92\u4f53\u62a5\u9053(?:\s*\|\s*\u8de8(?P<t>\d+)\u7c7b\u5a92\u4f53)?",
                        block)
    heat_m_en = re.search(
        r"Heat\s*[:\uff1a]\s*(?:Covered by\s*)?(?P<n>\d+)\s*(?:sources|source)(?:\s*\|\s*Across\s*(?P<t>\d+)\s*media\s*types)?",
        block,
        flags=re.IGNORECASE,
    )

    link_m = re.search(r"\[([^\]]+)\]\(([^)]+)\)", block)

    t = parse_dt(time_m.group(1).strip()) if time_m else None
    summary = summary_m.group(1).strip() if summary_m else ""
    src_name = link_m.group(1).strip() if link_m else ""
    src_link = link_m.group(2).strip() if link_m else ""

    report_n = None
    type_n = None
    if heat_m_cn:
        report_n = int(heat_m_cn.group("n"))
        type_n = int(heat_m_cn.group("t")) if heat_m_cn.group("t") else 1
    elif heat_m_en:
        report_n = int(heat_m_en.group("n"))
        type_n = int(heat_m_en.group("t")) if heat_m_en.group("t") else 1

    if report_n is None:
        # Heat line omitted => single source.
        report_n = 1
        type_n = 1

    return {
        "time": t,
        "summary": summary,
        "source_name": src_name,
        "source_link": src_link,
        "report_n": report_n,
        "type_n": type_n,
    }


def expected_fire(report_n: int, type_n: int) -> int:
    if report_n >= 3 and type_n >= 2:
        return 3
    if report_n >= 2 and type_n >= 2:
        return 2
    if report_n >= 2 and type_n == 1:
        return 1
    return 0


def score_format(output: object) -> int:
    if not isinstance(output, str) or not output.strip():
        return 0
    s = output.strip()
    has_header = (NEWS in s) or ("Hotspots" in s)
    has_item = ("\u3010\u70ed\u70b9" in s) or ("\u3010Hotspot" in s)
    has_link = re.search(r"\[[^\]]+\]\([^)]+\)", s) is not None
    looks_like_json = s.lstrip().startswith("{") and ("\"" in s and "}" in s)

    if looks_like_json:
        return 0
    if has_header and has_item and has_link:
        return 10
    if has_item:
        return 5
    return 0


def score_structure(items: list[dict], parsed_fields: list[dict]) -> int:
    if not items:
        return 0
    ok = 0
    total = 0
    for f in parsed_fields:
        total += 1
        if f.get("time") and f.get("summary") and f.get("source_name") and f.get("source_link"):
            ok += 1
    if total == 0:
        return 0
    rate = ok / total
    if rate >= 0.9:
        return 10
    if rate >= 0.6:
        return 5
    return 0


def score_instruction(items: list[dict], parsed_fields: list[dict], expected_limit: int, language: str, src_filter: str, output: object) -> int:
    if not items:
        return 0
    # limit
    diff = abs(len(items) - expected_limit)
    s_limit = 10 if diff == 0 else (5 if diff == 1 else 0)

    # language
    s_lang = 10
    if language == "english":
        if not isinstance(output, str):
            s_lang = 0
        else:
            s = output
            # Weak heuristic: if there are lots of CJK characters, it's likely not English.
            cjk = sum("\u4e00" <= ch <= "\u9fff" for ch in s)
            latin = sum(ch.isascii() and ch.isalpha() for ch in s)
            if latin == 0 and cjk > 0:
                s_lang = 0
            elif cjk > latin:
                s_lang = 5

    # source filter
    s_src = 10
    sf = (src_filter or "").strip()
    if sf:
        # English output may map names like 36\u6c2a => 36kr
        accept = {sf}
        if sf == "36\u6c2a":
            accept.add("36kr")
        total = 0
        ok = 0
        for f in parsed_fields:
            total += 1
            name = (f.get("source_name") or "").strip()
            if any(a in name for a in accept):
                ok += 1
        if total == 0:
            s_src = 0
        else:
            rate = ok / total
            if rate == 1.0:
                s_src = 10
            elif rate >= 0.8:
                s_src = 5
            else:
                s_src = 0

    return min(s_limit, s_lang, s_src)


def score_info_density(parsed_fields: list[dict], language: str) -> int:
    if not parsed_fields:
        return 0
    total = 0
    ok = 0
    for f in parsed_fields:
        total += 1
        s = (f.get("summary") or "").strip()
        if not s:
            continue
        n = len(s)
        if language == "english":
            # English summaries are often longer; keep a loose range.
            if 60 <= n <= 240:
                ok += 1
        else:
            if 80 <= n <= 120:
                ok += 1
    if total == 0:
        return 0
    rate = ok / total
    if rate >= 0.8:
        return 10
    if rate >= 0.4:
        return 5
    return 0


def score_summary_quality(parsed_fields: list[dict]) -> int:
    if not parsed_fields:
        return 0
    total = 0
    ok = 0
    for f in parsed_fields:
        total += 1
        s = (f.get("summary") or "").strip()
        if not s:
            continue
        if "\u6682\u65e0" in s or "N/A" in s:
            continue
        if s[-1] not in "\u3002\uff01\uff1f.!?":
            continue
        ok += 1
    rate = ok / total if total else 0.0
    if rate >= 0.9:
        return 10
    if rate >= 0.6:
        return 5
    return 0


def score_freshness(parsed_fields: list[dict], current_time_str: str) -> int:
    if not parsed_fields:
        return 0
    now = parse_dt(current_time_str)
    if now is None:
        return 5
    cutoff = now - timedelta(hours=24)
    total = 0
    ok = 0
    for f in parsed_fields:
        total += 1
        t = f.get("time")
        if t is None:
            continue
        if cutoff <= t <= now:
            ok += 1
    if total == 0:
        return 0
    rate = ok / total
    if rate == 1.0:
        return 10
    if rate >= 0.8:
        return 5
    return 0


def score_sorting(items: list[dict], parsed_fields: list[dict]) -> int:
    if not items or not parsed_fields or len(items) < 2:
        return 5

    tuples = []
    fire_mismatch = 0
    for it, f in zip(items, parsed_fields):
        rep = int(f.get("report_n") or 1)
        typ = int(f.get("type_n") or 1)
        t = f.get("time")
        # Missing time makes strict validation unreliable.
        if t is None:
            return 0
        tuples.append((rep, typ, t))

        fire_count = (it.get("fire_prefix") or "").count(FIRE)
        exp = expected_fire(rep, typ)
        if fire_count != exp:
            fire_mismatch += 1

    inversions = 0
    for i in range(len(tuples) - 1):
        a = tuples[i]
        b = tuples[i + 1]
        # Desired order: report desc, types desc, time desc
        if (a[0], a[1], a[2]) < (b[0], b[1], b[2]):
            inversions += 1

    base = 10 if inversions == 0 else (5 if inversions == 1 else 0)
    if fire_mismatch == 0:
        return base
    # Penalize one level if fire icons disagree with declared heat.
    return max(0, base - 5)


def cost_score(model: str) -> int:
    return 100 if model == "DeepSeek V3.2" else 60


def main():
    if not SRC.exists():
        raise SystemExit(f"Source workbook not found: {SRC}")

    wb_src = load_workbook(SRC)
    ws_src = wb_src["Sheet1"]

    wb = Workbook()
    ws = wb.active
    ws.title = "\u8bc4\u6d4b\u8bb0\u5f55"

    headers = [
        "Case ID",
        "scenario",
        "keyword",
        "\u6a21\u578b",
        "language",
        "limit",
        "source",
        "Agent\u63d0\u793a\u8bcd",
        "\u7528\u6237\u8f93\u5165(User_prompt)",
        "\u5b9e\u9645\u8f93\u51fa",
        "\u683c\u5f0f\u6b63\u786e(15%)",
        "\u7ed3\u6784\u5b8c\u6574(15%)",
        "\u6307\u4ee4\u9075\u5faa(15%)",
        "\u4fe1\u606f\u5bc6\u5ea6(15%)",
        "\u6392\u5e8f\u51c6\u786e\u6027(20%)",
        "\u6458\u8981\u8d28\u91cf(10%)",
        "\u65f6\u6548\u6027(10%)",
        "\u8d28\u91cfRaw(10\u5206\u5236)",
        "\u54cd\u5e94\u65f6\u95f4(s)",
        "\u8d28\u91cf\u5f97\u5206(100)",
        "\u65f6\u6548\u5f97\u5206(100)",
        "\u6210\u672c\u5f97\u5206(100)",
        "\u7efc\u5408\u603b\u5206",
        "\u5ba1\u8ba1\u5907\u6ce8",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Read test rows: session/turn numeric.
    for r in range(2, ws_src.max_row + 1):
        session = ws_src.cell(r, 1).value
        turn = ws_src.cell(r, 2).value
        if not isinstance(session, (int, float)) or not isinstance(turn, (int, float)):
            continue

        topic = str(ws_src.cell(r, 3).value or "").strip()
        lang = normalize_language(ws_src.cell(r, 4).value)
        limit_cell = ws_src.cell(r, 5).value
        source_filter = normalize_source(ws_src.cell(r, 6).value)
        sys_prompt = ws_src.cell(r, 7).value
        user_prompt = ws_src.cell(r, 8).value
        current_time = extract_current_time(user_prompt)
        keyword = extract_keyword(user_prompt) or topic

        # Determine expected limit.
        expected_limit = normalize_limit(limit_cell)
        scen = scenario_label(lang, expected_limit, source_filter)

        for model, out_col, time_col in (
            ("DeepSeek V3.2", 9, 10),
            ("\u8c46\u5305 1.8", 11, 12),
        ):
            output = ws_src.cell(r, out_col).value
            resp_s = parse_seconds(ws_src.cell(r, time_col).value)

            items, parse_notes = parse_markdown_items(output)
            fields = [parse_item_fields(it) for it in items]

            s_fmt = score_format(output)
            s_struct = score_structure(items, fields)
            s_instr = score_instruction(items, fields, expected_limit, lang, source_filter, output)
            s_dense = score_info_density(fields, lang)
            s_sort = score_sorting(items, fields)
            s_sum = score_summary_quality(fields)
            s_fresh = score_freshness(fields, current_time)

            audit = []
            if current_time:
                audit.append(f"current_time={current_time}")
            if parse_notes:
                audit.extend(parse_notes)
            if not items:
                audit.append("empty_output_or_parse_failed")

            case_id = f"S{int(session)}-T{int(turn)}"
            ws.append(
                [
                    case_id,
                    scen,
                    keyword,
                    model,
                    lang,
                    expected_limit,
                    source_filter,
                    sys_prompt,
                    user_prompt,
                    output,
                    s_fmt,
                    s_struct,
                    s_instr,
                    s_dense,
                    s_sort,
                    s_sum,
                    s_fresh,
                    None,
                    resp_s,
                    None,
                    None,
                    cost_score(model),
                    None,
                    ";".join(audit) if audit else "",
                ]
            )

            rr = ws.max_row
            # Quality raw (0-10)
            ws.cell(rr, 18, f"=K{rr}*0.15+L{rr}*0.15+M{rr}*0.15+N{rr}*0.15+O{rr}*0.2+P{rr}*0.1+Q{rr}*0.1")
            ws.cell(rr, 20, f"=R{rr}*10")
            ws.cell(rr, 21, f"=MAX(0,MIN(100,(1-S{rr}/60)*100))")
            ws.cell(rr, 23, f"=T{rr}*0.75+V{rr}*0.15+U{rr}*0.10")

    # Column widths
    widths = {
        "A": 10,
        "B": 10,
        "C": 10,
        "D": 14,
        "E": 10,
        "F": 6,
        "G": 10,
        "H": 22,
        "I": 22,
        "J": 45,
        "K": 12,
        "L": 12,
        "M": 12,
        "N": 12,
        "O": 14,
        "P": 12,
        "Q": 12,
        "R": 16,
        "S": 12,
        "T": 12,
        "U": 12,
        "V": 12,
        "W": 12,
        "X": 20,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("\u6c47\u603b\u5bf9\u6bd4")
    ws2.append(["\u6a21\u578b", "\u8d28\u91cf\u5747\u5206(75%)", "\u6210\u672c\u5747\u5206(15%)", "\u65f6\u6548\u5747\u5206(10%)", "\u6700\u7ec8\u52a0\u6743\u603b\u5206", "\u63a8\u8350"])
    for c in ws2[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    models = ["DeepSeek V3.2", "\u8c46\u5305 1.8"]
    for i, m in enumerate(models, start=2):
        ws2.cell(i, 1, m)
        ws2.cell(i, 2, f"=AVERAGEIF(\u8bc4\u6d4b\u8bb0\u5f55!D:D,\"{m}\",\u8bc4\u6d4b\u8bb0\u5f55!T:T)")
        ws2.cell(i, 3, f"=AVERAGEIF(\u8bc4\u6d4b\u8bb0\u5f55!D:D,\"{m}\",\u8bc4\u6d4b\u8bb0\u5f55!V:V)")
        ws2.cell(i, 4, f"=AVERAGEIF(\u8bc4\u6d4b\u8bb0\u5f55!D:D,\"{m}\",\u8bc4\u6d4b\u8bb0\u5f55!U:U)")
        ws2.cell(i, 5, f"=B{i}*0.75+C{i}*0.15+D{i}*0.10")
    ws2.cell(2, 6, "=IF(E2>E3,\"YES\",\"\")")
    ws2.cell(3, 6, "=IF(E3>E2,\"YES\",\"\")")
    for col in "ABCDEF":
        ws2.column_dimensions[col].width = 20

    # Rubric sheet
    ws3 = wb.create_sheet("\u8bc4\u5206\u6807\u51c6")
    ws3.append(["\u7ef4\u5ea6", "\u6743\u91cd", "\u6807\u51c6(10/5/0)"])
    for c in ws3[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.append(["\u683c\u5f0f\u6b63\u786e", "15%", "10=\u7a33\u5b9aMarkdown(\u6807\u9898+\u70ed\u70b9+\u94fe\u63a5)\uff1b5=\u4e3b\u4f53\u53ef\u8bfb\uff1b0=\u683c\u5f0f\u5d29\u574f/\u8f93\u51fa\u50cfJSON"])
    ws3.append(["\u7ed3\u6784\u5b8c\u6574", "15%", "10=\u5927\u90e8\u5206\u6761\u76ee\u90fd\u6709\u65f6\u95f4/\u6458\u8981/\u6765\u6e90\u94fe\u63a5\uff1b5=\u90e8\u5206\u7f3a\u5931\uff1b0=\u5927\u91cf\u7f3a\u5931"])
    ws3.append(["\u6307\u4ee4\u9075\u5faa", "15%", "10=limit\u6570\u91cf\u6b63\u786e+\u8bed\u8a00/\u4fe1\u6e90\u7ea6\u675f\u9075\u5faa\uff1b5=\u5c0f\u504f\u5dee\uff1b0=\u5ffd\u7565\u7ea6\u675f"])
    ws3.append(["\u4fe1\u606f\u5bc6\u5ea6", "15%", "10=\u6458\u898180-120\u5b57(\u82f1\u6587\u653e\u5bbd)\uff1b5=\u90e8\u5206\u8fc7\u957f/\u8fc7\u77ed\uff1b0=\u5927\u91cf\u5931\u63a7"])
    ws3.append(["\u6392\u5e8f\u51c6\u786e\u6027", "20%", "10=\u70ed\u5ea6\u6392\u5e8f\u81ea\u6d3d+\u706b\u82d7\u4e0e\u70ed\u5ea6\u6570\u5b57\u5339\u914d\uff1b5=\u5fae\u5c0f\u9519\u4f4d\uff1b0=\u660e\u663e\u4e71\u5e8f/\u76f8\u90bb\u4e24\u5904\u4ee5\u4e0a\u8fdd\u53cd\u89c4\u5219"])
    ws3.append(["\u6458\u8981\u8d28\u91cf", "10%", "10=\u6458\u8981\u7ed3\u675f\u6b63\u5e38\u3001\u65e0\u7a7a\u8bdd\u3001\u65e0\u2018\u6682\u65e0\u2019\uff1b5=\u5c11\u91cf\u95ee\u9898\uff1b0=\u5927\u91cf\u8d28\u91cf\u95ee\u9898"])
    ws3.append(["\u65f6\u6548\u6027", "10%", "10=\u516824h\u5185\uff1b5=\u5c11\u91cf\u8d85\u65f6/\u7f3a\u5931\uff1b0=\u5927\u91cf\u8d85\u65f6"])
    ws3.append(["\u603b\u5206", "-", "\u603b\u5206=\u8d28\u91cf(75%)+\u6210\u672c(15%)+\u65f6\u6548(10%)\uff1b\u6210\u672c\uff1aDeepSeek=100, \u8c46\u5305=60\uff1b\u5b8c\u6210\u65f6\u95f4\uff1aMAX(0,MIN(100,(1-\u79d2/60)*100))"])
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 110
    for r in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=3):
        for c in r:
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    wb.save(OUT)


if __name__ == "__main__":
    main()
