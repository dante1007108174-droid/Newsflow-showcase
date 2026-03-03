"""Update badcase合集.xlsx: keep first 4 data rows, merge the rest by root cause.

Preserves:
- Row 1 (header)
- Row 2-5 (U-04, U-NEG-02, D-03, GEN-05)

Replaces Row 6+ with root-cause-merged groups (natural language descriptions).
"""

from __future__ import annotations

import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]


def _safe(s: Any) -> str:
    return str(s or "").strip()


def _classify_root(root: str) -> str:
    s = _safe(root)
    s_l = s.lower()

    if "来源" in s or "source" in s_l:
        if "过滤" in s or "指定" in s:
            return "来源过滤/来源约束"
        if "字段" in s:
            return "来源字段/格式约束"
        return "来源相关"

    if "格式" in s or "链接" in s or "时间" in s:
        return "输出格式模板"

    if "结构" in s or "热点" in s:
        return "输出结构模板"

    if "supabase" in s_l or "rpc" in s_l:
        return "Supabase/RPC"

    if "白名单" in s or "主题" in s:
        return "主题/白名单校验"

    return s or "(未填写)"


def _extract_raw_query(desc: str) -> str:
    import re
    m = re.search(r'用户输入「([^」]+)」', desc)
    return m.group(1) if m else ""


def _extract_main_issue(desc: str) -> str:
    import re
    m = re.search(r'主要问题：([^；]+)', desc)
    return m.group(1) if m else ""


def _generate_natural_desc(items: List[Dict[str, str]], group_key: str) -> str:
    if not items:
        return ""

    inputs: List[str] = []
    issues: List[str] = []

    for item in items:
        desc = item.get("desc", "")
        inp = _extract_raw_query(desc)
        issue = _extract_main_issue(desc)
        if inp and inp not in inputs:
            inputs.append(inp)
        if issue and issue not in issues:
            issues.append(issue)

    if "来源过滤" in group_key or "来源约束" in group_key:
        base = "用户指定特定来源时，模型未能正确过滤，返回了非指定来源的内容。"
        if inputs:
            examples = "；".join(inputs[:3])
            base += f"例如：{examples}。"
        if issues:
            base += f"主要表现为：{'；'.join(issues[:2])}。"
        return base

    if "格式" in group_key or "字段" in group_key:
        base = "模型输出格式不稳定，有时缺少必要的时间或来源链接字段。"
        if inputs:
            examples = "；".join(inputs[:3])
            base += f"例如：{examples}。"
        if issues:
            base += f"主要问题：{'；'.join(issues[:2])}。"
        return base

    if "结构" in group_key:
        base = "模型输出结构不完整，缺少标题或【热点N】条目格式。"
        if inputs:
            examples = "；".join(inputs[:2])
            base += f"例如：{examples}。"
        return base

    parts: List[str] = []
    if inputs:
        parts.append(f"涉及输入：{'；'.join(inputs[:3])}")
    if issues:
        parts.append(f"主要问题：{'；'.join(issues[:2])}")

    return "。".join(parts) if parts else "详见原始记录"


def _worst_status(statuses: List[str]) -> str:
    s = [x for x in (_safe(v) for v in statuses) if x]
    if any("待" in x or "🔴" in x for x in s):
        return "🔴 待修复"
    if any("已解决" in x or "✅" in x for x in s):
        return "✅ 已解决"
    return s[0] if s else ""


def _merge_lines(items: List[str], *, max_items: int = 12) -> str:
    uniq: List[str] = []
    for it in items:
        it = _safe(it)
        if not it:
            continue
        if it not in uniq:
            uniq.append(it)
    if len(uniq) > max_items:
        uniq = uniq[:max_items] + [f"...（省略{len(uniq) - max_items}条）"]
    return "\n".join(uniq)


def main() -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    src_path = ROOT / "docs" / "badcase合集.xlsx"
    if not src_path.exists():
        raise SystemExit(f"Not found: {src_path}")

    wb = load_workbook(src_path)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    # Header mapping
    header_row = 1
    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = _safe(ws.cell(header_row, c).value)
        if v:
            headers[v] = c

    required = ["ID", "问题名称", "问题描述", "根因", "影响", "优化方式", "问题状态"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise SystemExit(f"Missing columns: {missing}")

    # Read all rows
    items: List[Dict[str, str]] = []
    for r in range(2, ws.max_row + 1):
        rid = _safe(ws.cell(r, headers["ID"]).value)
        if not rid:
            continue
        items.append(
            {
                "row": str(r),
                "id": rid,
                "name": _safe(ws.cell(r, headers["问题名称"]).value),
                "desc": _safe(ws.cell(r, headers["问题描述"]).value),
                "root": _safe(ws.cell(r, headers["根因"]).value),
                "impact": _safe(ws.cell(r, headers["影响"]).value),
                "fix": _safe(ws.cell(r, headers["优化方式"]).value),
                "status": _safe(ws.cell(r, headers["问题状态"]).value),
            }
        )

    # Split: first 4 data rows (rows 2-5 in original) vs rest
    first_4 = items[:4]  # U-04, U-NEG-02, D-03, GEN-05
    rest = items[4:]     # NEWS-Txx and others

    print(f"Preserving first 4: {[x['id'] for x in first_4]}")
    print(f"Merging rest: {[x['id'] for x in rest]}")

    # Group the rest by root cause
    groups: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for it in rest:
        key = _classify_root(it.get("root", ""))
        groups[key].append(it)

    # Build merged rows for the rest
    merged_rows: List[Dict[str, str]] = []

    # Sort groups: pending first, then by size desc
    def group_sort_key(k: str) -> Tuple[int, int, str]:
        g = groups[k]
        worst = _worst_status([x.get("status", "") for x in g])
        pending_first = 0 if ("待" in worst or "🔴" in worst) else 1
        return (pending_first, -len(g), k)

    for key in sorted(groups.keys(), key=group_sort_key):
        g = groups[key]
        ids = [x["id"] for x in g]
        names = [x.get("name", "") for x in g]
        worst = _worst_status([x.get("status", "") for x in g])

        # Generate merged description
        natural_desc = _generate_natural_desc(g, key)

        # Build merged impacts and fixes
        impacts = [x.get("impact", "") for x in g]
        fixes = [x.get("fix", "") for x in g]

        # Create merged ID and name
        merged_id = f"GROUP-{'-'.join(ids[:3])}{'...' if len(ids) > 3 else ''}"
        merged_name = f"【{key}】({len(g)}条)"

        merged_rows.append({
            "id": merged_id,
            "name": merged_name,
            "desc": natural_desc,
            "root": key,
            "impact": _merge_lines(impacts, max_items=20),
            "fix": _merge_lines(fixes, max_items=20),
            "status": worst,
        })

    # Clear rows after row 5 (keep header + first 4 data rows)
    # Delete from row 6 onwards
    if ws.max_row > 5:
        ws.delete_rows(6, ws.max_row - 5)

    # Append merged rows
    header_fill = PatternFill("solid", fgColor="FEF9C3")  # Light yellow for merged groups
    for row_data in merged_rows:
        r = ws.max_row + 1
        ws.cell(r, headers["ID"]).value = row_data["id"]
        ws.cell(r, headers["问题名称"]).value = row_data["name"]
        ws.cell(r, headers["问题描述"]).value = row_data["desc"]
        ws.cell(r, headers["根因"]).value = row_data["root"]
        ws.cell(r, headers["影响"]).value = row_data["impact"]
        ws.cell(r, headers["优化方式"]).value = row_data["fix"]
        ws.cell(r, headers["问题状态"]).value = row_data["status"]

        # Apply light yellow background to indicate merged group
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Style status column
    pass_fill = PatternFill("solid", fgColor="DCFCE7")
    fail_fill = PatternFill("solid", fgColor="FEE2E2")
    for r in range(6, ws.max_row + 1):
        cell = ws.cell(r, headers["问题状态"])
        v = _safe(cell.value)
        if "已解决" in v or "✅" in v:
            cell.fill = pass_fill
        elif "待" in v or "🔴" in v:
            cell.fill = fail_fill

    # Save
    try:
        wb.save(src_path)
        print(f"Updated: {src_path}")
        print(f"Total rows now: {ws.max_row}")
        print(f"Merged groups: {len(merged_rows)}")
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = src_path.with_name(f"badcase合集_merged_{ts}.xlsx")
        wb.save(fallback)
        print(f"Saved to fallback: {fallback}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
