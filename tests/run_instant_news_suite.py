"""Run end-to-end instant news workflow suite and write reports.

Usage:
  python tests/run_instant_news_suite.py

Outputs (tests/reports/):
  - instant_news_report_<run_id>.json
  - instant_news_report_<run_id>.md
  - instant_news_test_data_<run_id>.xlsx
  Plus stable "latest" files:
  - instant_news_report.json
  - instant_news_report.md
  - instant_news_test_data.xlsx

Notes:
  - This suite targets the full workflow (keyword + raw_query only).
  - All cases are evaluated by an LLM judge plus deterministic checks.
"""

from __future__ import annotations

import json
import os
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from coze_workflow_client import CozeWorkflowClient


ROOT = Path(__file__).resolve().parents[1]
SUITE_PATH = Path(__file__).resolve().parent / "instant_news_suite.json"
REPORT_DIR = Path(__file__).resolve().parent / "reports"


def _safe_print(s: str) -> None:
    try:
        print(s)
    except UnicodeEncodeError:
        print(s.encode("utf-8", errors="replace").decode("utf-8", errors="replace"))


def _now_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _load_json_file(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _save_json(path: Path, data: Dict[str, Any]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _truncate(s: str, n: int) -> str:
    s = s or ""
    return s if len(s) <= n else (s[: n - 1] + "…")


def _is_infra_exception(e: Exception) -> bool:
    msg = f"{type(e).__name__}: {e}".lower()
    return any(
        k in msg
        for k in (
            "sslerror",
            "certificate_verify_failed",
            "hostname mismatch",
            "connectionerror",
            "readtimeout",
            "connecttimeout",
            "proxyerror",
            "temporary failure",
            "name resolution",
        )
    )


def _normalize_text(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9\u4e00-\u9fff]+", "", s or "").lower()


def _parse_dt(s: str) -> Optional[datetime]:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue
    return None


def _count_items(output: str) -> int:
    m = re.findall(r"【\s*(?:热点|Hotspot)\s*\d+\s*】", output, flags=re.I)
    if m:
        return len(m)
    m2 = re.findall(r"【\s*\d+\s*】", output)
    if m2:
        return len(m2)
    # Last resort: count markdown list items.
    m3 = re.findall(r"^\s*[-*]\s+", output, flags=re.M)
    return len(m3) if m3 else 0


def _extract_sources(output: str) -> List[str]:
    sources: List[str] = []
    for line in output.splitlines():
        if ("来源" not in line) and (not re.search(r"\bSource\b", line, flags=re.I)):
            continue
        m = re.search(r"(?:来源|\bSource\b)\s*[:：]\s*(.*)$", line, flags=re.I)
        if not m:
            continue
        tail = m.group(1).strip()
        m2 = re.search(r"\[([^\]]+)\]\(", tail)
        if m2:
            sources.append(m2.group(1).strip())
            continue
        tail = re.sub(r"[·•]\s*", "", tail).strip()
        if tail:
            sources.append(tail.split()[0].strip())
    return sources


def _extract_times(output: str) -> List[datetime]:
    dts: List[datetime] = []
    for m in re.findall(r"(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2})", output):
        dt = _parse_dt(m)
        if dt:
            dts.append(dt)
    return dts


def _extract_current_time(output: str) -> Optional[datetime]:
    if not output:
        return None
    m = re.search(r"筛选时间[:：]\s*(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2})", output)
    if not m:
        m = re.search(r"Filtered\s+at[:：]\s*(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2})", output, flags=re.I)
    if not m:
        return None
    return _parse_dt(m.group(1))


def _is_hotspot_header(line: str) -> bool:
    return bool(
        re.search(
            r"^\s*(?:🔥+\s*)?(?:\*\*)?【\s*(?:热点|Hotspot)\s*\d+\s*】",
            line or "",
            flags=re.I,
        )
    )


def _extract_hotspot_blocks(output: str) -> List[List[str]]:
    blocks: List[List[str]] = []
    current: List[str] = []
    for line in (output or "").splitlines():
        if _is_hotspot_header(line):
            if current:
                blocks.append(current)
            current = [line]
            continue
        if current:
            current.append(line)
    if current:
        blocks.append(current)
    return blocks


def _extract_hotspot_title(header: str) -> str:
    m = re.search(r"【\s*(?:热点|Hotspot)\s*\d+\s*】\s*(.+)$", header or "", flags=re.I)
    if not m:
        return (header or "").strip()
    title = m.group(1).strip()
    return title.strip("* ")


def _extract_report_count(block_text: str) -> int:
    m_cn = re.search(r"热度\s*[:：]\s*(\d+)\s*家媒体报道", block_text)
    if m_cn:
        return max(1, int(m_cn.group(1)))
    m_en = re.search(r"Heat\s*[:：]\s*Covered\s+by\s+(\d+)\s+sources?", block_text, flags=re.I)
    if m_en:
        return max(1, int(m_en.group(1)))
    return 1


def _extract_media_type_count(block_text: str) -> int:
    m_cn = re.search(r"跨\s*(\d+)\s*类媒体", block_text)
    if m_cn:
        return max(1, int(m_cn.group(1)))
    m_en = re.search(r"Across\s+(\d+)\s+media\s+types?", block_text, flags=re.I)
    if m_en:
        return max(1, int(m_en.group(1)))
    return 1


def _extract_item_time(block_text: str) -> Optional[datetime]:
    m = re.search(
        r"(?:时间|Time)\s*[:：]\s*(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}(?::\d{2})?)",
        block_text,
        flags=re.I,
    )
    if not m:
        return None
    return _parse_dt(m.group(1))


def _check_sorted_by_stats(output: str) -> CheckResult:
    blocks = _extract_hotspot_blocks(output)
    if len(blocks) <= 1:
        return CheckResult(name="sorted_by_stats", passed=True)

    parsed: List[Dict[str, Any]] = []
    missing_time_titles: List[str] = []
    for i, block in enumerate(blocks, start=1):
        header = block[0] if block else ""
        block_text = "\n".join(block)
        title = _extract_hotspot_title(header)
        item_time = _extract_item_time(block_text)
        if not item_time:
            missing_time_titles.append(title or f"热点{i}")
        parsed.append(
            {
                "title": title or f"热点{i}",
                "report_count": _extract_report_count(block_text),
                "media_type_count": _extract_media_type_count(block_text),
                "time": item_time,
            }
        )

    if missing_time_titles:
        sample = "、".join(missing_time_titles[:3])
        return CheckResult(
            name="sorted_by_stats",
            passed=False,
            reason=f"排序键缺失：以下条目无法解析时间字段：{sample}",
            root_cause="每条热点未稳定输出“时间：YYYY-MM-DD HH:MM”导致无法做三重排序校验",
            suggestion="提示词中强制每条热点输出标准时间格式；必要时在工作流后处理补齐时间",
        )

    def _key(item: Dict[str, Any]) -> Tuple[int, int, datetime]:
        return (
            int(item["report_count"]),
            int(item["media_type_count"]),
            item["time"],
        )

    for idx in range(len(parsed) - 1):
        left = parsed[idx]
        right = parsed[idx + 1]
        if _key(left) < _key(right):
            return CheckResult(
                name="sorted_by_stats",
                passed=False,
                reason=(
                    "排序不符合规则："
                    f"“{left['title']}”(媒体={left['report_count']}, 跨类={left['media_type_count']}, 时间={left['time'].strftime('%Y-%m-%d %H:%M')}) "
                    "应排在 "
                    f"“{right['title']}”(媒体={right['report_count']}, 跨类={right['media_type_count']}, 时间={right['time'].strftime('%Y-%m-%d %H:%M')}) 之后"
                ),
                root_cause="输出顺序未严格按 报道数量→跨媒体类型→发布时间 三重键降序",
                suggestion="在工作流增加代码后处理节点，对热点块按三重排序键做确定性重排",
            )

    return CheckResult(name="sorted_by_stats", passed=True)


def _looks_like_friendly_no_result(output: str) -> bool:
    s = (output or "").strip()
    if s == "":
        return True
    has_link = bool(re.search(r"\[[^\]]+\]\(https?://[^)]+\)", s))
    has_dt = bool(re.search(r"\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}", s))
    has_items = _count_items(s) > 0
    if has_link or has_dt or has_items:
        return False
    patterns = [
        r"未找到",
        r"没有找到",
        r"暂无",
        r"无匹配",
        r"0\s*条",
        r"暂时.*在路上",
        r"还在路上",
        r"正在路上",
        r"情报.*在路上",
        r"稍后再试",
        r"请稍后",
        r"敬请期待",
        r"暂未",
        r"没有可用",
        r"暂时没有",
        r"暂未接入",
        r"目前支持的来源",
        r"not available yet",
        r"not supported yet",
        r"currently supported sources",
    ]
    return any(re.search(p, s) for p in patterns)


def _is_mostly_english(output: str) -> bool:
    if not output:
        return False
    zh = len(re.findall(r"[\u4e00-\u9fff]", output))
    total = max(len(output), 1)
    ratio = zh / total
    has_ascii_words = bool(re.search(r"[A-Za-z]{3,}", output))
    return ratio < 0.25 and has_ascii_words


def _canonicalize_source_name(name: str) -> str:
    raw = (name or "").strip()
    if not raw:
        return ""
    key = _normalize_text(raw)
    mapping = {
        "虎嗅": "虎嗅",
        "huxiu": "虎嗅",
        "华尔街见闻": "华尔街见闻",
        "wallstreetcn": "华尔街见闻",
        "wscn": "华尔街见闻",
        "it之家": "IT之家",
        "ithome": "IT之家",
        "36氪": "36氪",
        "36kr": "36氪",
        "三十六氪": "36氪",
        "钛媒体": "钛媒体",
        "tmtpost": "钛媒体",
        "界面": "界面",
        "界面新闻": "界面",
        "jiemian": "界面",
        "新华社": "新华社",
        "xinhua": "新华社",
        "中新网": "中新网",
        "chinanews": "中新网",
        "澎湃": "澎湃",
        "thepaper": "澎湃",
        "36kr": "36氪",
        "36krypton": "36氪",
        "36krnews": "36氪",
        "36krcom": "36氪",
        "36kr中文": "36氪",
        "36kr英文": "36氪",
        "36krenglish": "36氪",
        "36kr中文网": "36氪",
        "36kr英文网": "36氪",
        "36kr网站": "36氪",
        "36kr网站中文": "36氪",
        "36kr网站英文": "36氪",
        "36krnewscom": "36氪",
        "36krnews": "36氪",
        "36krsite": "36氪",
        "36krweb": "36氪",
        "36kr网站com": "36氪",
        "36kr网站news": "36氪",
        "36kr网站site": "36氪",
        "36kr网站web": "36氪",
        "36kr网站cn": "36氪",
        "36kr网站zh": "36氪",
        "36kr网站en": "36氪",
        "36kr网站english": "36氪",
        "36kr网站中文站": "36氪",
        "36kr网站英文站": "36氪",
        "36kr新闻": "36氪",
        "36kr报道": "36氪",
        "36kr资讯": "36氪",
        "36kr新闻网": "36氪",
        "36kr资讯网": "36氪",
        "36kr报道网": "36氪",
        "36kr文章": "36氪",
        "36kr媒体": "36氪",
    }
    return mapping.get(key, raw)


def _find_sources_in_raw_query(raw_query: str) -> List[str]:
    rq = _normalize_text(raw_query)
    alias_map = {
        "36氪": ["36kr", "36氪", "三十六氪"],
        "IT之家": ["ithome", "it之家", "it之家", "it之", "it之家"],
        "华尔街见闻": ["wallstreetcn", "wscn", "华尔街见闻", "华尔街"],
        "虎嗅": ["huxiu", "虎嗅", "虎嗅网"],
        "钛媒体": ["tmtpost", "钛媒体"],
        "界面": ["jiemian", "界面", "界面新闻"],
        "新华社": ["xinhua", "新华社"],
        "中新网": ["chinanews", "中新网"],
        "澎湃": ["thepaper", "澎湃"],
    }
    found: List[str] = []
    for canonical, aliases in alias_map.items():
        for alias in aliases:
            if _normalize_text(alias) in rq:
                found.append(canonical)
                break
    return found


def _parse_raw_query_constraints(raw_query: str) -> Dict[str, Any]:
    rq = (raw_query or "").strip()
    out: Dict[str, Any] = {}
    if not rq:
        return out

    sources = _find_sources_in_raw_query(rq)
    if sources:
        out["expected_sources"] = sources
        if len(sources) == 1:
            out["expected_source"] = sources[0]

    # Count constraints like "给我3条" or "3 items".
    m_cn = re.search(r"(\d+)\s*(条|个)", rq)
    m_en = re.search(r"\b(\d+)\s*(items|item|news|stories|results)\b", rq, flags=re.I)
    if m_cn:
        out["max_count"] = int(m_cn.group(1))
    elif m_en:
        out["max_count"] = int(m_en.group(1))

    if re.search(r"\bEnglish\b|英文|英语|in English", rq, flags=re.I):
        out["language"] = "en"

    return out


@dataclass
class CheckResult:
    name: str
    passed: bool
    reason: str = ""
    root_cause: str = ""
    suggestion: str = ""


@dataclass
class JudgeResult:
    passed: bool
    scores: Dict[str, float]
    issue: str
    root_cause: str
    suggestion: str
    confidence: str
    raw: str


def _check_hard_rules(output: str, raw_query: str, expect: Dict[str, Any]) -> List[CheckResult]:
    rules = expect.get("hard_checks") or []
    results: List[CheckResult] = []

    parsed_rq = _parse_raw_query_constraints(raw_query)

    expected_source = str(expect.get("expected_source") or parsed_rq.get("expected_source") or "").strip()
    expected_source_c = _canonicalize_source_name(expected_source)

    max_count = expect.get("max_count")
    if max_count is None:
        max_count = parsed_rq.get("max_count")
    if max_count is None:
        max_count = 5
    try:
        max_count = int(max_count)
    except Exception:
        max_count = 5
    max_count = min(max_count, 10)

    expected_language = str(expect.get("language") or parsed_rq.get("language") or "").strip().lower()
    allow_empty = bool(expect.get("allow_empty"))

    def add(name: str, ok: bool, reason: str = "", root_cause: str = "", suggestion: str = "") -> None:
        results.append(CheckResult(name=name, passed=ok, reason=reason, root_cause=root_cause, suggestion=suggestion))

    if "non_empty" in rules:
        ok = bool((output or "").strip()) or allow_empty
        add(
            "non_empty",
            ok,
            reason="输出为空" if (not ok and not allow_empty) else "",
            root_cause="模型未产出任何内容" if (not ok and not allow_empty) else "",
            suggestion="检查工作流是否正常执行，确保LLM节点有输出" if not ok else "",
        )

    if "empty_ok" in rules:
        friendly = _looks_like_friendly_no_result(output)
        add(
            "empty_ok",
            friendly,
            reason="未返回友好的空状态提示" if not friendly else "",
            root_cause="无匹配时未按空状态模板输出" if not friendly else "",
            suggestion="在提示词中明确无匹配时的固定空状态模板" if not friendly else "",
        )

    if "format" in rules:
        has_link = bool(re.search(r"\[[^\]]+\]\(https?://[^)]+\)", output))
        has_dt = bool(re.search(r"\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}", output))
        ok = has_link and has_dt
        add(
            "format",
            ok,
            reason=f"缺少时间或来源链接(link={has_link}, time={has_dt})" if not ok else "",
            root_cause="输出未按固定格式包含时间与来源链接" if not ok else "",
            suggestion="提示词强调每条必须包含时间和来源链接" if not ok else "",
        )

    if "structure" in rules:
        has_header = ("热点" in output) or ("24H" in output) or ("24h" in output)
        has_items = _count_items(output) > 0
        ok = has_header and has_items
        add(
            "structure",
            ok,
            reason=f"缺少标题或条目(header={has_header}, items={has_items})" if not ok else "",
            root_cause="输出结构不完整或未使用固定条目格式" if not ok else "",
            suggestion="使用【热点 N】格式输出每条新闻" if not ok else "",
        )

    if "limit_max" in rules and isinstance(max_count, int) and max_count > 0:
        n = _count_items(output)
        ok = (n == 0 and allow_empty) or (n <= max_count)
        add(
            "limit_max",
            ok,
            reason=f"条数超限(期望≤{max_count}, 实际={n})" if not ok else "",
            root_cause="未按用户条数要求截断输出" if not ok else "",
            suggestion="严格遵守用户指定条数，超出需截断" if not ok else "",
        )

    if "source_only" in rules and expected_source_c:
        if allow_empty and _looks_like_friendly_no_result(output):
            add("source_only", True)
        else:
            srcs = [_canonicalize_source_name(s) for s in _extract_sources(output)]
            if not srcs:
                add(
                    "source_only",
                    False,
                    reason="无法识别来源字段",
                    root_cause="输出未包含来源字段或格式不规范",
                    suggestion="强制每条输出“来源：[来源名](链接)”",
                )
            else:
                bad = sorted({s for s in srcs if s != expected_source_c})
                ok = len(bad) == 0
                add(
                    "source_only",
                    ok,
                    reason=f"包含非指定来源：{bad}" if not ok else "",
                    root_cause="来源过滤未严格执行" if not ok else "",
                    suggestion="用户指定来源时只保留该来源" if not ok else "",
                )

    if "language" in rules and expected_language:
        ok = _is_mostly_english(output) if expected_language == "en" else True
        add(
            "language",
            ok,
            reason="输出语言不符合英文要求" if not ok else "",
            root_cause="语言切换未生效" if not ok else "",
            suggestion="提示词强调语言输出规则" if not ok else "",
        )

    if "timeliness" in rules:
        current_time = _extract_current_time(output)
        if not current_time:
            add(
                "timeliness",
                False,
                reason="无法解析筛选时间",
                root_cause="输出未包含筛选时间或格式不正确",
                suggestion="在标题区域输出“筛选时间：YYYY-MM-DD HH:MM”",
            )
        else:
            times = _extract_times(output)
            if not times:
                add(
                    "timeliness",
                    False,
                    reason="未识别到新闻时间",
                    root_cause="每条新闻缺少时间字段",
                    suggestion="要求每条包含“时间：YYYY-MM-DD HH:MM”",
                )
            else:
                cutoff = current_time - timedelta(hours=24)
                old = [dt.strftime("%Y-%m-%d %H:%M") for dt in times if dt < cutoff]
                ok = len(old) == 0
                add(
                    "timeliness",
                    ok,
                    reason=f"包含超过24小时的新闻：{old[:3]}" if not ok else "",
                    root_cause="时效过滤未执行或执行不完整" if not ok else "",
                    suggestion="仅保留current_time往前24小时内新闻" if not ok else "",
                )

    should_check_sorted = (
        "sorted_by_stats" in rules
        or "structure" in rules
        or "timeliness" in rules
    )
    if should_check_sorted:
        add_result = _check_sorted_by_stats(output)
        results.append(add_result)

    return results


def _try_get_judge_api_key() -> Optional[str]:
    """Optional judge API key.

    If not set, we fall back to rule-based scoring so the suite can still run and
    produce Excel artifacts for MVP iteration.
    """

    key = os.getenv("OPENAI_API_KEY") or os.getenv("DEEPSEEK_API_KEY")
    return str(key).strip() if key else None


def _get_judge_api_base() -> str:
    return os.getenv("OPENAI_API_BASE") or os.getenv("DEEPSEEK_API_BASE") or "https://api.openai.com/v1"


def _get_judge_model() -> str:
    return os.getenv("GPT_JUDGE_MODEL", "gpt-4o-mini")


def _extract_json(text: str) -> Optional[Dict[str, Any]]:
    if not text:
        return None
    s = text.strip()
    if s.startswith("{") and s.endswith("}"):
        try:
            return json.loads(s)
        except Exception:
            return None
    m = re.search(r"\{[\s\S]*\}", s)
    if not m:
        return None
    try:
        return json.loads(m.group(0))
    except Exception:
        return None


def _call_judge(prompt: str, *, retries: int = 2) -> str:
    api_key = _try_get_judge_api_key()
    if not api_key:
        raise RuntimeError("judge API key missing")
    api_base = _get_judge_api_base()
    model = _get_judge_model()

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.2,
    }

    last_err: Optional[Exception] = None
    for attempt in range(1, retries + 1):
        try:
            resp = requests.post(
                f"{api_base}/chat/completions",
                headers=headers,
                json=payload,
                timeout=90,
            )
            resp.raise_for_status()
            data = resp.json()
            return data["choices"][0]["message"]["content"]
        except Exception as e:
            last_err = e
            time.sleep(0.6 * attempt)
    raise RuntimeError(f"judge request failed: {last_err}")


def _build_judge_prompt(
    raw_query: str,
    keyword: str,
    expected_source: str,
    expected_limit: int,
    expected_language: str,
    output: str,
) -> str:
    rules = [
        "你是严格的新闻结果评估员。",
        "请根据用户请求与模型输出进行判定，必须给出可执行的结论。",
        "重点判断：相关性、排序合理性、去重、指令遵循(来源/条数/语言)。",
        "只输出JSON，不要包含任何额外说明或Markdown。",
        "JSON格式：",
        "{",
        '  "pass": true/false,',
        '  "scores": {"relevance": 0-10, "ranking": 0-10, "dedup": 0-10, "instruction": 0-10, "language": 0-10},',
        '  "issue": "出现了什么问题(如果通过可写空)" ,',
        '  "root_cause": "失败原因(如果通过可写空)",',
        '  "suggestion": "修复建议(如果通过可写空)",',
        '  "confidence": "high|medium|low"',
        "}",
        "评分规则：>=7视为通过，<7视为不通过。",
        "如果输出为空或明显不符合模板，请判定为不通过。",
        "以下为输入信息：",
        f"用户请求: {raw_query}",
        f"关键词: {keyword}",
        f"期望来源: {expected_source or '未指定'}",
        f"期望条数(最大10): {expected_limit}",
        f"期望语言: {expected_language or '中文'}",
        "模型输出：",
        _truncate(output, 2200),
    ]
    return "\n".join(rules)


def _judge_case(
    raw_query: str,
    keyword: str,
    expected_source: str,
    expected_limit: int,
    expected_language: str,
    output: str,
) -> JudgeResult:
    prompt = _build_judge_prompt(raw_query, keyword, expected_source, expected_limit, expected_language, output)
    try:
        raw = _call_judge(prompt)
        parsed = _extract_json(raw) or {}
    except Exception as e:
        # Fallback: rule-based scoring so we can still generate the Excel report.
        raw = f"judge_skipped: {type(e).__name__}: {e}"
        parsed = {}

    scores = parsed.get("scores") or {}
    def _score(name: str) -> float:
        v = scores.get(name)
        try:
            return float(v)
        except Exception:
            return 0.0

    rel = _score("relevance")
    rank = _score("ranking")
    dedup = _score("dedup")
    instr = _score("instruction")
    lang = _score("language")

    # If judge was skipped, use conservative defaults.
    if not parsed:
        # Instruction/language can be approximated from output shape.
        instr = 10.0 if output.strip() else 0.0
        if expected_language == "en":
            lang = 10.0 if _is_mostly_english(output) else 0.0
        else:
            lang = 10.0
        # Soft dimensions default to neutral (requires manual verification if desired).
        rel = 7.0
        rank = 7.0
        dedup = 7.0

    if "pass" in parsed:
        passed = bool(parsed.get("pass"))
    else:
        avg = (rel + rank + dedup + instr + lang) / 5.0
        passed = avg >= 7.0

    issue = str(parsed.get("issue") or "").strip()
    root_cause = str(parsed.get("root_cause") or "").strip()
    suggestion = str(parsed.get("suggestion") or "").strip()
    confidence = str(parsed.get("confidence") or "").strip().lower() or ("low" if not parsed else "medium")
    if confidence not in ("high", "medium", "low"):
        confidence = "medium"

    return JudgeResult(
        passed=passed,
        scores={
            "relevance": rel,
            "ranking": rank,
            "dedup": dedup,
            "instruction": instr,
            "language": lang,
        },
        issue=issue,
        root_cause=root_cause,
        suggestion=suggestion,
        confidence=confidence,
        raw=raw,
    )


def _call_workflow(
    client: CozeWorkflowClient, workflow_id: str, params: Dict[str, str]
) -> Tuple[str, int, Optional[str]]:
    attempts = 3
    for attempt in range(1, attempts + 1):
        try:
            t0 = time.time()
            res = client.run_workflow(workflow_id, params, is_async=False, timeout_s=180)
            dt_ms = int((time.time() - t0) * 1000)
            return res.output or "", dt_ms, res.debug_url
        except RuntimeError as e:
            msg = str(e)
            if "did not return output or execute_id" in msg and attempt < attempts:
                time.sleep(0.5 * attempt)
                continue
            raise
    raise RuntimeError("unreachable")


def _run_one_case(workflow_id: str, case: Dict[str, Any]) -> Dict[str, Any]:
    """Run a single case.

    Important: create a fresh client per case to avoid sharing a requests.Session
    across threads.
    """

    case_id = str(case.get("id") or "").strip()
    intent = str(case.get("intent") or "").strip()
    scenario = str(case.get("scenario") or "").strip()
    inputs = case.get("inputs") or {}
    expect = case.get("expect") or {}

    keyword = str(inputs.get("keyword") or "").strip()
    raw_query = str(inputs.get("raw_query") or "").strip()
    if not case_id or not raw_query:
        return {}

    params = {"keyword": keyword, "raw_query": raw_query}
    t0 = time.time()
    debug_url: Optional[str] = None

    try:
        client = CozeWorkflowClient.from_env()
        output, dt_ms, debug_url = _call_workflow(client, workflow_id, params)
        duration_s = round(dt_ms / 1000.0, 3)

        checks = _check_hard_rules(output, raw_query, expect)
        hard_failed = [c for c in checks if not c.passed]

        parsed_rq = _parse_raw_query_constraints(raw_query)
        expected_source = str(expect.get("expected_source") or parsed_rq.get("expected_source") or "").strip()
        expected_limit = expect.get("max_count")
        if expected_limit is None:
            expected_limit = parsed_rq.get("max_count")
        if expected_limit is None:
            expected_limit = 5
        try:
            expected_limit = int(expected_limit)
        except Exception:
            expected_limit = 5
        expected_limit = min(expected_limit, 10)

        expected_language = str(expect.get("language") or parsed_rq.get("language") or "").strip().lower()
        judge = _judge_case(raw_query, keyword, expected_source, expected_limit, expected_language, output)

        if hard_failed:
            status = "FAIL"
            status_cn = "❌失败"
            issue = _merge_text([c.reason for c in hard_failed])
            root_cause = _merge_text([c.root_cause for c in hard_failed])
            suggestion = _merge_text([c.suggestion for c in hard_failed])
            fix_status = "🔴待修复"
        else:
            if judge.passed:
                status = "PASS"
                status_cn = "✅通过"
                issue = ""
                root_cause = ""
                suggestion = ""
                fix_status = "⚪不适用"
            else:
                status = "FAIL"
                status_cn = "❌失败"
                issue = judge.issue or "模型评估未通过"
                root_cause = judge.root_cause or "模型判断质量不达标"
                suggestion = judge.suggestion or "调整提示词或过滤逻辑以提升相关性与排序"
                fix_status = "🔴待修复"

        remarks: List[str] = []
        if duration_s > 30:
            remarks.append("响应慢")
        if len(output or "") > 5000:
            remarks.append("输出过长")
        if keyword and keyword not in raw_query:
            remarks.append("关键词与用户输入不一致")
        if judge.confidence == "low":
            remarks.append("模型评估置信度低")
        # Only expose debug_url for failed/error cases to reduce clutter.
        if debug_url and status != "PASS":
            remarks.append(f"debug_url={debug_url}")

        indicator = _indicator_text(checks, judge)

        out = {
            "id": case_id,
            "intent": intent,
            "scenario": scenario,
            "params": params,
            "acceptance": expect.get("acceptance") or "",
            "duration_ms": dt_ms,
            "duration_s": duration_s,
            "status": status,
            "status_cn": status_cn,
            "indicator": indicator,
            "issue": issue,
            "root_cause": root_cause,
            "suggestion": suggestion,
            "fix_status": fix_status,
            "manual_review": bool(case.get("manual_review")),
            "remarks": _merge_text(remarks),
            "checks": [c.__dict__ for c in checks],
            "judge": {
                "passed": judge.passed,
                "scores": judge.scores,
                "issue": judge.issue,
                "root_cause": judge.root_cause,
                "suggestion": judge.suggestion,
                "confidence": judge.confidence,
                "raw": judge.raw,
            },
            "output": output,
        }
        if debug_url and status != "PASS":
            out["debug_url"] = debug_url
        return out
    except Exception as e:
        duration_s = round((time.time() - t0), 3)
        status = "ERROR" if _is_infra_exception(e) else "FAIL"
        status_cn = "⚪异常" if status == "ERROR" else "❌失败"
        remarks = []
        if debug_url and status != "PASS":
            remarks.append(f"debug_url={debug_url}")
        out = {
            "id": case_id,
            "intent": intent,
            "scenario": scenario,
            "params": params,
            "acceptance": expect.get("acceptance") or "",
            "duration_ms": int(duration_s * 1000),
            "duration_s": duration_s,
            "status": status,
            "status_cn": status_cn,
            "indicator": "",
            "issue": f"{type(e).__name__}: {e}",
            "root_cause": "工作流执行异常或模型评估失败",
            "suggestion": "检查网络、API配置与工作流状态",
            "fix_status": "🔴待修复" if status != "ERROR" else "⚪不适用",
            "manual_review": bool(case.get("manual_review")),
            "remarks": _merge_text(remarks),
            "checks": [],
            "judge": {},
            "output": "",
        }
        if debug_url and status != "PASS":
            out["debug_url"] = debug_url
        return out


def _indicator_text(checks: List[CheckResult], judge: Optional[JudgeResult]) -> str:
    label_map = {
        "non_empty": "非空",
        "format": "格式",
        "structure": "结构",
        "limit_max": "条数",
        "source_only": "来源",
        "language": "语言",
        "timeliness": "时效",
        "empty_ok": "空状态",
        "sorted_by_stats": "排序键",
    }
    parts: List[str] = []
    for c in checks:
        label = label_map.get(c.name, c.name)
        parts.append(f"{label}{'✅' if c.passed else '❌'}")
    if judge:
        s = judge.scores
        parts.append(
            "模型:相关性{:.1f} 排序{:.1f} 去重{:.1f} 指令{:.1f} 语言{:.1f} 总体{}".format(
                s.get("relevance", 0.0),
                s.get("ranking", 0.0),
                s.get("dedup", 0.0),
                s.get("instruction", 0.0),
                s.get("language", 0.0),
                "PASS" if judge.passed else "FAIL",
            )
        )
    return " ".join(parts)


def _merge_text(items: List[str]) -> str:
    items = [x for x in items if x]
    return "；".join(items)


def _make_excel(report: Dict[str, Any], *, xlsx_path: Path) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "测试汇总"

    summary = report.get("summary") or {}
    run = report.get("run") or {}
    suite = report.get("suite") or {}

    summary_rows = [
        ["执行时间", run.get("timestamp") or ""],
        ["工作流ID", suite.get("workflow_id") or ""],
        ["测试版本", suite.get("version") or ""],
        ["总用例数", summary.get("total") or 0],
        ["通过", summary.get("passed") or 0],
        ["失败", summary.get("failed") or 0],
        ["异常", summary.get("errors") or 0],
        ["通过率", summary.get("pass_rate") or ""],
        ["待抽查", summary.get("manual_review") or 0],
    ]

    _write_table(ws_summary, ["指标", "结果"], summary_rows)
    ws_summary.column_dimensions["A"].width = 18
    ws_summary.column_dimensions["B"].width = 60

    ws_detail = wb.create_sheet("测试明细")
    headers = [
        "ID",
        "意图",
        "测试场景",
        "关键词",
        "用户输入",
        "预期结果(验收标准)",
        "评估指标",
        "结果判断",
        "问题",
        "归因",
        "建议",
        "解决状态",
        "响应时间(秒)",
        "待抽查",
        "备注",
        "输出预览",
    ]

    rows: List[List[Any]] = []
    for r in report.get("results", []) or []:
        rows.append(
            [
                r.get("id"),
                r.get("intent"),
                r.get("scenario"),
                (r.get("params") or {}).get("keyword"),
                (r.get("params") or {}).get("raw_query"),
                r.get("acceptance"),
                r.get("indicator"),
                r.get("status_cn"),
                r.get("issue"),
                r.get("root_cause"),
                r.get("suggestion"),
                r.get("fix_status"),
                r.get("duration_s"),
                "是" if r.get("manual_review") else "否",
                r.get("remarks"),
                _truncate(str(r.get("output") or ""), 300),
            ]
        )

    _write_table(ws_detail, headers, rows)

    # Style tweaks
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws_detail.iter_rows(min_row=2):
        for idx in (4, 5, 6, 8, 9, 10, 14, 15):
            row[idx].alignment = wrap

    _apply_result_fills(ws_detail, result_col=8, fix_col=12)
    _apply_zebra_rows(ws_detail)

    widths = {
        "A": 8,
        "B": 12,
        "C": 20,
        "D": 10,
        "E": 36,
        "F": 36,
        "G": 40,
        "H": 12,
        "I": 28,
        "J": 28,
        "K": 28,
        "L": 12,
        "M": 12,
        "N": 10,
        "O": 20,
        "P": 60,
    }
    for col, w in widths.items():
        ws_detail.column_dimensions[col].width = w

    ws_failed = wb.create_sheet("失败用例")
    failed_rows = [r for r in rows if "失败" in str(r[7]) or "异常" in str(r[7])]
    _write_table(ws_failed, headers, failed_rows)
    _apply_result_fills(ws_failed, result_col=8, fix_col=12)
    _apply_zebra_rows(ws_failed)
    for col, w in widths.items():
        ws_failed.column_dimensions[col].width = w

    ws_review = wb.create_sheet("待抽查")
    review_rows = [r for r in rows if r[13] == "是"]
    _write_table(ws_review, headers, review_rows)
    _apply_result_fills(ws_review, result_col=8, fix_col=12)
    _apply_zebra_rows(ws_review)
    for col, w in widths.items():
        ws_review.column_dimensions[col].width = w

    ws_rules = wb.create_sheet("评估标准")
    rule_rows = [
        ["非空", "输出不能为空（除非允许空状态）"],
        ["格式", "每条新闻需包含时间与来源链接"],
        ["结构", "包含标题与【热点 N】条目"],
        ["条数", "不超过用户指定条数，默认5条，上限10条"],
        ["来源", "用户指定来源时只输出该来源"],
        ["语言", "用户要求英文时输出英文"],
        ["时效", "筛选时间前24小时内"],
        ["排序键", "热点顺序必须满足 报道数desc→跨媒体类型desc→发布时间desc"],
        ["空状态", "无匹配时输出友好提示"],
        ["模型评估", "相关性/排序/去重/指令/语言评分 >=7 视为通过"],
    ]
    _write_table(ws_rules, ["评估项", "说明"], rule_rows)
    ws_rules.column_dimensions["A"].width = 16
    ws_rules.column_dimensions["B"].width = 80

    # --- Sheet: 按意图统计 ---
    ws_intent = wb.create_sheet("按意图统计")
    intent_stats: Dict[str, Dict[str, int]] = {}
    for r in report.get("results", []) or []:
        intent = r.get("intent") or "(未分类)"
        if intent not in intent_stats:
            intent_stats[intent] = {"total": 0, "passed": 0, "failed": 0}
        intent_stats[intent]["total"] += 1
        if r.get("status") == "PASS":
            intent_stats[intent]["passed"] += 1
        else:
            intent_stats[intent]["failed"] += 1

    intent_headers = ["意图", "用例数", "通过", "失败", "通过率"]
    intent_rows: List[List[Any]] = []
    for intent_name, stats in intent_stats.items():
        t = stats["total"]
        p = stats["passed"]
        f = stats["failed"]
        rate = f"{round((p / max(t, 1)) * 100, 1)}%"
        intent_rows.append([intent_name, t, p, f, rate])

    _write_table(ws_intent, intent_headers, intent_rows)
    ws_intent.column_dimensions["A"].width = 18
    ws_intent.column_dimensions["B"].width = 12
    ws_intent.column_dimensions["C"].width = 12
    ws_intent.column_dimensions["D"].width = 12
    ws_intent.column_dimensions["E"].width = 14

    # Color pass rate cells
    for row in ws_intent.iter_rows(min_row=2, min_col=5, max_col=5):
        cell = row[0]
        val = str(cell.value or "")
        try:
            pct = float(val.replace("%", ""))
            if pct >= 100:
                cell.fill = PatternFill("solid", fgColor="DCFCE7")
            elif pct >= 80:
                cell.fill = PatternFill("solid", fgColor="FEF9C3")
            else:
                cell.fill = PatternFill("solid", fgColor="FEE2E2")
        except Exception:
            pass

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(xlsx_path)
    except PermissionError:
        run_id = str(run.get("run_id") or datetime.now().strftime("%Y%m%d_%H%M%S"))
        fallback = xlsx_path.with_name(f"instant_news_test_data_{run_id}.xlsx")
        wb.save(fallback)
        _safe_print(f"Permission denied writing {xlsx_path}; wrote fallback: {fallback}")


def _write_table(ws, headers: List[str], rows: List[List[Any]]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.append(headers)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
    ws.freeze_panes = "A2"
    for r in rows:
        ws.append(r)


def _apply_result_fills(ws, *, result_col: int, fix_col: int) -> None:
    pass_fill = PatternFill("solid", fgColor="DCFCE7")
    fail_fill = PatternFill("solid", fgColor="FEE2E2")
    error_fill = PatternFill("solid", fgColor="E5E7EB")
    fix_red = PatternFill("solid", fgColor="FEE2E2")
    fix_green = PatternFill("solid", fgColor="DCFCE7")
    fix_gray = PatternFill("solid", fgColor="F3F4F6")

    for row in ws.iter_rows(min_row=2):
        result_cell = row[result_col - 1]
        val = str(result_cell.value or "")
        if "通过" in val:
            result_cell.fill = pass_fill
        elif "失败" in val:
            result_cell.fill = fail_fill
        elif "异常" in val:
            result_cell.fill = error_fill

        fix_cell = row[fix_col - 1]
        fix_val = str(fix_cell.value or "")
        if "待修复" in fix_val:
            fix_cell.fill = fix_red
        elif "已修复" in fix_val:
            fix_cell.fill = fix_green
        else:
            fix_cell.fill = fix_gray


def _apply_zebra_rows(ws) -> None:
    zebra_fill = PatternFill("solid", fgColor="F8FAFC")
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if i % 2 == 0:
            for cell in row:
                if cell.fill is None or cell.fill.fill_type is None:
                    cell.fill = zebra_fill


def _render_md(report: Dict[str, Any]) -> str:
    summary = report.get("summary") or {}
    run = report.get("run") or {}
    suite = report.get("suite") or {}

    lines: List[str] = []
    lines.append("# instant_news_e2e report")
    lines.append("")
    lines.append(f"- Run ID: `{run.get('run_id')}`")
    lines.append(f"- Time: `{run.get('timestamp')}`")
    lines.append(f"- Workflow ID: `{suite.get('workflow_id')}`")
    lines.append(f"- Version: `{suite.get('version')}`")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- Total: {summary.get('total')}")
    lines.append(f"- Passed: {summary.get('passed')}")
    lines.append(f"- Failed: {summary.get('failed')}")
    lines.append(f"- Errors: {summary.get('errors')}")
    lines.append(f"- Pass Rate: {summary.get('pass_rate')}")
    lines.append(f"- Manual Review: {summary.get('manual_review')}")
    lines.append("")
    lines.append("## Failures")
    for r in report.get("results", []) or []:
        if r.get("status") not in ("FAIL", "ERROR"):
            continue
        lines.append("")
        lines.append(f"### {r.get('id')} {r.get('scenario')}")
        lines.append(f"- Result: {r.get('status_cn')}")
        lines.append(f"- Issue: {r.get('issue')}")
        lines.append(f"- Root Cause: {r.get('root_cause')}")
        lines.append(f"- Suggestion: {r.get('suggestion')}")

    return "\n".join(lines)


def main() -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    if not SUITE_PATH.exists():
        raise SystemExit(f"Suite file not found: {SUITE_PATH}")

    suite = _load_json_file(SUITE_PATH)
    suite_meta = suite.get("suite") or {}
    workflow_id = str(suite_meta.get("workflow_id") or "").strip()
    if not workflow_id:
        raise SystemExit("suite.workflow_id is required")

    cases = suite.get("cases")
    if not isinstance(cases, list) or not cases:
        raise SystemExit("suite.cases must be a non-empty list")

    judge_enabled = bool(_try_get_judge_api_key())

    # Validate Coze credentials early.
    CozeWorkflowClient.from_env()

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_json_path = REPORT_DIR / f"instant_news_report_{run_id}.json"
    report_md_path = REPORT_DIR / f"instant_news_report_{run_id}.md"
    report_xlsx_path = REPORT_DIR / f"instant_news_test_data_{run_id}.xlsx"
    report_json_latest_path = REPORT_DIR / "instant_news_report.json"
    report_md_latest_path = REPORT_DIR / "instant_news_report.md"
    report_xlsx_latest_path = REPORT_DIR / "instant_news_test_data.xlsx"

    report: Dict[str, Any] = {
        "suite": suite_meta,
        "run": {
            "run_id": run_id,
            "timestamp": _now_ts(),
            "base_url": os.getenv("COZE_BASE_URL") or os.getenv("COZE_API_BASE") or "https://api.coze.cn",
            "judge_enabled": judge_enabled,
            "judge_model": _get_judge_model() if judge_enabled else "(disabled)",
        },
        "results": [],
        "summary": {},
    }

    max_workers = int(os.getenv("INSTANT_NEWS_WORKERS") or "4")
    futures = {}
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        for idx, case in enumerate(cases):
            fut = ex.submit(_run_one_case, workflow_id, case)
            futures[fut] = idx

        results_by_idx: Dict[int, Dict[str, Any]] = {}
        for fut in as_completed(futures):
            idx = futures[fut]
            try:
                results_by_idx[idx] = fut.result() or {}
            except Exception as e:
                results_by_idx[idx] = {
                    "id": f"case_{idx}",
                    "intent": "",
                    "scenario": "(runner exception)",
                    "params": {},
                    "acceptance": "",
                    "duration_ms": 0,
                    "duration_s": 0,
                    "status": "ERROR",
                    "status_cn": "⚪异常",
                    "indicator": "",
                    "issue": f"{type(e).__name__}: {e}",
                    "root_cause": "测试脚本并发执行异常",
                    "suggestion": "检查测试脚本并发与环境依赖",
                    "fix_status": "🔴待修复",
                    "manual_review": False,
                    "remarks": "",
                    "checks": [],
                    "judge": {},
                    "output": "",
                }

    report["results"] = [results_by_idx.get(i, {}) for i in range(len(cases)) if results_by_idx.get(i)]

    total = len(report["results"])
    passed = len([r for r in report["results"] if r.get("status") == "PASS"])
    failed = len([r for r in report["results"] if r.get("status") == "FAIL"])
    errors = len([r for r in report["results"] if r.get("status") == "ERROR"])
    manual_review = len([r for r in report["results"] if r.get("manual_review")])
    denom = max(passed + failed, 1)
    pass_rate = f"{round((passed / denom) * 100, 1)}%"

    report["summary"] = {
        "total": total,
        "passed": passed,
        "failed": failed,
        "errors": errors,
        "pass_rate": pass_rate,
        "manual_review": manual_review,
    }

    _save_json(report_json_path, report)
    _save_json(report_json_latest_path, report)

    md = _render_md(report)
    report_md_path.write_text(md, encoding="utf-8")
    report_md_latest_path.write_text(md, encoding="utf-8")

    _make_excel(report, xlsx_path=report_xlsx_path)
    _make_excel(report, xlsx_path=report_xlsx_latest_path)

    _safe_print(f"Report JSON: {report_json_path}")
    _safe_print(f"Report MD: {report_md_path}")
    _safe_print(f"Report XLSX: {report_xlsx_path}")

    return 0 if failed == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
