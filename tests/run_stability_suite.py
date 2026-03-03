"""Run stability suite for instant news workflow.

This script is meant for diagnosing flaky behavior in an end-to-end Coze workflow.

It combines:
  - Round 1: an existing `instant_news_report_*.json` produced by
    `tests/run_instant_news_suite.py`
  - Rounds 2-3: executed now, concurrently across cases

Pass criteria (per your requirement): a case only passes if it passes 3/3 rounds.

Usage:
  python tests/run_stability_suite.py

Optional env vars:
  - STABILITY_ROUND1_REPORT: path to round-1 JSON report
  - INSTANT_NEWS_WORKERS: thread pool size (default: 8)

Outputs (tests/reports/):
  - instant_news_stability_report_<run_id>.json
  - instant_news_stability_report_<run_id>.md
  - instant_news_stability_test_data_<run_id>.xlsx
  Plus stable "latest" files:
  - instant_news_stability_report.json
  - instant_news_stability_report.md
  - instant_news_stability_test_data.xlsx
"""

from __future__ import annotations

import json
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from coze_workflow_client import CozeWorkflowClient

import run_instant_news_suite as base


ROOT = Path(__file__).resolve().parents[1]
SUITE_PATH = Path(__file__).resolve().parent / "instant_news_suite.json"
REPORT_DIR = Path(__file__).resolve().parent / "reports"


def _now_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _load_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _save_json(path: Path, data: Dict[str, Any]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _find_round1_report() -> Tuple[Path, Dict[str, Any]]:
    # Allow explicit override.
    env = (os.getenv("STABILITY_ROUND1_REPORT") or "").strip()
    candidates: List[Path] = []
    if env:
        candidates.append(Path(env))

    # Default: use the specific run from the last conversation if present.
    candidates.append(REPORT_DIR / "instant_news_report_20260207_201524.json")
    # Fallback to the stable latest.
    candidates.append(REPORT_DIR / "instant_news_report.json")

    for p in candidates:
        try:
            if p.exists():
                return p, _load_json(p)
        except Exception:
            continue
    raise SystemExit(
        "Round-1 report not found. Set STABILITY_ROUND1_REPORT or run tests/run_instant_news_suite.py first."
    )


def _round_mark(status: str) -> str:
    if status == "PASS":
        return "✅"
    if status == "ERROR":
        return "⚪"
    return "❌"


def _slim_case_result(r: Dict[str, Any]) -> Dict[str, Any]:
    """Keep only fields needed for stability aggregation/reporting."""

    out = {
        "id": r.get("id"),
        "intent": r.get("intent"),
        "scenario": r.get("scenario"),
        "params": r.get("params") or {},
        "acceptance": r.get("acceptance") or "",
        "duration_s": r.get("duration_s"),
        "duration_ms": r.get("duration_ms"),
        "status": r.get("status"),
        "status_cn": r.get("status_cn"),
        "indicator": r.get("indicator") or "",
        "issue": r.get("issue") or "",
        "root_cause": r.get("root_cause") or "",
        "suggestion": r.get("suggestion") or "",
        "fix_status": r.get("fix_status") or "",
        "manual_review": bool(r.get("manual_review")),
        "remarks": r.get("remarks") or "",
        "debug_url": r.get("debug_url") or "",
        "output_preview": base._truncate(str(r.get("output") or ""), 300),
    }
    return out


def _compute_round_summary(results: Dict[str, Dict[str, Any]], case_ids: List[str]) -> Dict[str, Any]:
    items = [results.get(cid) for cid in case_ids]
    items = [x for x in items if x]
    total = len(items)
    passed = len([r for r in items if r.get("status") == "PASS"])
    failed = len([r for r in items if r.get("status") == "FAIL"])
    errors = len([r for r in items if r.get("status") == "ERROR"])
    denom = max(passed + failed, 1)
    pass_rate = round((passed / denom) * 100.0, 1)
    return {
        "total": total,
        "passed": passed,
        "failed": failed,
        "errors": errors,
        "pass_rate": f"{pass_rate}%",
    }


def _run_live_rounds(
    *,
    workflow_id: str,
    cases: List[Dict[str, Any]],
    round_names: List[str],
    max_workers: int,
) -> Dict[str, Dict[str, Dict[str, Any]]]:
    """Run multiple rounds in one thread pool for maximum throughput."""

    futures: Dict[Any, Tuple[str, str]] = {}
    results: Dict[str, Dict[str, Dict[str, Any]]] = {rn: {} for rn in round_names}

    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        for rn in round_names:
            for case in cases:
                cid = str(case.get("id") or "").strip()
                if not cid:
                    continue
                fut = ex.submit(base._run_one_case, workflow_id, case)
                futures[fut] = (rn, cid)

        for fut in as_completed(futures):
            rn, cid = futures[fut]
            try:
                raw = fut.result() or {}
            except Exception as e:
                raw = {
                    "id": cid,
                    "intent": "",
                    "scenario": "(runner exception)",
                    "params": {},
                    "acceptance": "",
                    "duration_ms": 0,
                    "duration_s": 0,
                    "status": "ERROR",
                    "status_cn": "⚪异常",
                    "indicator": "",
                    "issue": f"{type(e).__name__}: {e}",
                    "root_cause": "测试脚本并发执行异常",
                    "suggestion": "降低并发并重试；检查网络和Coze服务状态",
                    "fix_status": "⚪不适用",
                    "manual_review": False,
                    "remarks": "",
                    "output": "",
                }
            if raw and raw.get("id"):
                results[rn][str(raw.get("id"))] = _slim_case_result(raw)

    return results


def _merge_remarks(items: List[str]) -> str:
    seen: List[str] = []
    for x in items:
        x = (x or "").strip()
        if not x:
            continue
        for part in x.split("；"):
            part = part.strip()
            if part and part not in seen:
                seen.append(part)
    return "；".join(seen)


def _merge_round_field(rounds: List[Dict[str, Any]], field: str) -> str:
    parts: List[str] = []
    for idx, r in enumerate(rounds, start=1):
        if not r or r.get("status") == "PASS":
            continue
        v = str(r.get(field) or "").strip()
        if not v:
            continue
        parts.append(f"R{idx}:{v}")
    return base._merge_text(parts)


def _avg_duration(rounds: List[Dict[str, Any]]) -> Optional[float]:
    vals: List[float] = []
    for r in rounds:
        v = r.get("duration_s")
        try:
            vals.append(float(v))
        except Exception:
            continue
    if not vals:
        return None
    return round(sum(vals) / len(vals), 3)


def _stability_label(passed_rounds: int, total_rounds: int) -> str:
    if passed_rounds == total_rounds:
        return "🟢稳定"
    if passed_rounds == 0:
        return "🔴持续失败"
    return "🟡不稳定"


def _make_excel(report: Dict[str, Any], *, xlsx_path: Path) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "测试汇总"

    run = report.get("run") or {}
    suite = report.get("suite") or {}
    rounds = report.get("rounds") or []
    overall = report.get("overall") or {}

    summary_rows: List[List[Any]] = [
        ["执行时间", run.get("timestamp") or ""],
        ["工作流ID", suite.get("workflow_id") or ""],
        ["测试版本", suite.get("version") or ""],
        ["轮次数", run.get("rounds") or 3],
        ["并发数", run.get("workers") or ""],
        ["Round1来源", run.get("round1_report") or ""],
    ]

    for r in rounds:
        name = r.get("name")
        rid = r.get("run_id") or ""
        s = r.get("summary") or {}
        pr = s.get("pass_rate") or ""
        summary_rows.append([f"{name} Run ID", rid])
        summary_rows.append([f"{name} 通过率", pr])
        summary_rows.append([f"{name} 通过/失败/异常", f"{s.get('passed')}/{s.get('failed')}/{s.get('errors')}"])

    summary_rows.extend(
        [
            ["平均通过率", overall.get("avg_pass_rate") or ""],
            ["总用例数", overall.get("total_cases") or 0],
            ["稳定通过(3/3)", overall.get("stable_pass") or 0],
            ["不稳定/失败(非3/3)", overall.get("unstable_or_fail") or 0],
        ]
    )

    base._write_table(ws_summary, ["指标", "值"], summary_rows)
    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 84

    # Highlight pass-rate rows for quick scanning.
    green = PatternFill("solid", fgColor="DCFCE7")
    yellow = PatternFill("solid", fgColor="FEF9C3")
    red = PatternFill("solid", fgColor="FEE2E2")
    for row in ws_summary.iter_rows(min_row=2, max_col=2):
        k = str(row[0].value or "")
        v = str(row[1].value or "")
        if "通过率" not in k:
            continue
        try:
            pct = float(v.replace("%", ""))
            row[1].fill = green if pct >= 90 else (yellow if pct >= 80 else red)
        except Exception:
            continue

    ws_detail = wb.create_sheet("测试明细")
    headers = [
        "ID",
        "意图",
        "测试场景",
        "关键词",
        "用户输入",
        "预期结果(验收标准)",
        "轮次结果",
        "通过次数",
        "失败次数",
        "稳定性",
        "评估指标(最后一轮)",
        "结果判断(3/3)",
        "问题",
        "归因",
        "建议",
        "解决状态",
        "平均响应时间(秒)",
        "待抽查",
        "备注",
        "输出预览(最后一轮)",
    ]

    rows: List[List[Any]] = []
    for r in report.get("results", []) or []:
        params = r.get("params") or {}
        rows.append(
            [
                r.get("id"),
                r.get("intent"),
                r.get("scenario"),
                params.get("keyword"),
                params.get("raw_query"),
                r.get("acceptance"),
                r.get("round_result"),
                r.get("passed_rounds"),
                r.get("failed_rounds"),
                r.get("stability"),
                r.get("indicator"),
                r.get("status_cn"),
                r.get("issue"),
                r.get("root_cause"),
                r.get("suggestion"),
                r.get("fix_status"),
                r.get("avg_duration_s"),
                "是" if r.get("manual_review") else "否",
                r.get("remarks"),
                r.get("output_preview"),
            ]
        )

    base._write_table(ws_detail, headers, rows)

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws_detail.iter_rows(min_row=2):
        # wide text columns
        for idx in (4, 5, 6, 12, 13, 14, 18, 19):
            row[idx].alignment = wrap

    # Result fills for strict outcome column
    base._apply_result_fills(ws_detail, result_col=12, fix_col=16)
    base._apply_zebra_rows(ws_detail)

    # Stability fills
    stable_fill = PatternFill("solid", fgColor="DCFCE7")
    flaky_fill = PatternFill("solid", fgColor="FEF9C3")
    fail_fill = PatternFill("solid", fgColor="FEE2E2")
    for row in ws_detail.iter_rows(min_row=2):
        cell = row[9]  # stability column (0-based index)
        v = str(cell.value or "")
        if "🟢" in v:
            cell.fill = stable_fill
        elif "🟡" in v:
            cell.fill = flaky_fill
        elif "🔴" in v:
            cell.fill = fail_fill

    widths = {
        "A": 8,
        "B": 12,
        "C": 22,
        "D": 10,
        "E": 36,
        "F": 34,
        "G": 14,
        "H": 10,
        "I": 10,
        "J": 12,
        "K": 42,
        "L": 14,
        "M": 28,
        "N": 26,
        "O": 26,
        "P": 12,
        "Q": 16,
        "R": 10,
        "S": 22,
        "T": 60,
    }
    for col, w in widths.items():
        ws_detail.column_dimensions[col].width = w

    ws_failed = wb.create_sheet("失败用例")
    failed_rows = [r for r in rows if "失败" in str(r[11]) or "异常" in str(r[11])]
    base._write_table(ws_failed, headers, failed_rows)
    base._apply_result_fills(ws_failed, result_col=12, fix_col=16)
    base._apply_zebra_rows(ws_failed)
    for col, w in widths.items():
        ws_failed.column_dimensions[col].width = w

    ws_intent = wb.create_sheet("按意图统计")
    intent_rows: List[List[Any]] = []
    intent_headers = ["意图", "用例数", "稳定通过(3/3)", "不稳定", "持续失败", "平均通过率"]
    for intent, stats in (report.get("intent_stats") or {}).items():
        intent_rows.append(
            [
                intent,
                stats.get("total"),
                stats.get("stable_pass"),
                stats.get("flaky"),
                stats.get("always_fail"),
                stats.get("avg_pass_rate"),
            ]
        )
    base._write_table(ws_intent, intent_headers, intent_rows)
    ws_intent.column_dimensions["A"].width = 18
    ws_intent.column_dimensions["B"].width = 12
    ws_intent.column_dimensions["C"].width = 16
    ws_intent.column_dimensions["D"].width = 12
    ws_intent.column_dimensions["E"].width = 12
    ws_intent.column_dimensions["F"].width = 14
    base._apply_zebra_rows(ws_intent)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(xlsx_path)
    except PermissionError:
        run_id = str(run.get("run_id") or datetime.now().strftime("%Y%m%d_%H%M%S"))
        fallback = xlsx_path.with_name(f"instant_news_stability_test_data_{run_id}.xlsx")
        wb.save(fallback)
        base._safe_print(f"Permission denied writing {xlsx_path}; wrote fallback: {fallback}")


def _render_md(report: Dict[str, Any]) -> str:
    run = report.get("run") or {}
    suite = report.get("suite") or {}
    overall = report.get("overall") or {}
    rounds = report.get("rounds") or []

    lines: List[str] = []
    lines.append("# instant_news_stability report")
    lines.append("")
    lines.append(f"- Run ID: `{run.get('run_id')}`")
    lines.append(f"- Time: `{run.get('timestamp')}`")
    lines.append(f"- Workflow ID: `{suite.get('workflow_id')}`")
    lines.append(f"- Version: `{suite.get('version')}`")
    lines.append(f"- Rounds: `{run.get('rounds')}`")
    lines.append("")
    lines.append("## Rounds")
    for r in rounds:
        s = r.get("summary") or {}
        lines.append(
            f"- {r.get('name')}: pass_rate={s.get('pass_rate')} passed={s.get('passed')} failed={s.get('failed')} errors={s.get('errors')}"
        )
    lines.append("")
    lines.append("## Overall")
    lines.append("")
    lines.append(f"- Avg Pass Rate: {overall.get('avg_pass_rate')}")
    lines.append(f"- Stable Pass (3/3): {overall.get('stable_pass')}/{overall.get('total_cases')}")
    lines.append(f"- Unstable or Fail: {overall.get('unstable_or_fail')}")
    lines.append("")
    lines.append("## Flaky / Failed Cases")
    for r in report.get("results", []) or []:
        if r.get("passed_rounds") == run.get("rounds"):
            continue
        lines.append("")
        lines.append(f"### {r.get('id')} {r.get('scenario')}")
        lines.append(f"- Round: {r.get('round_result')}")
        lines.append(f"- Issue: {r.get('issue')}")
        if r.get("remarks"):
            lines.append(f"- Remarks: {r.get('remarks')}")
    return "\n".join(lines)


def main() -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    if not SUITE_PATH.exists():
        raise SystemExit(f"Suite file not found: {SUITE_PATH}")

    suite = _load_json(SUITE_PATH)
    suite_meta = suite.get("suite") or {}
    workflow_id = str(suite_meta.get("workflow_id") or "").strip()
    if not workflow_id:
        raise SystemExit("suite.workflow_id is required")

    cases = suite.get("cases")
    if not isinstance(cases, list) or not cases:
        raise SystemExit("suite.cases must be a non-empty list")
    case_ids = [str(c.get("id") or "").strip() for c in cases if str(c.get("id") or "").strip()]

    # Validate Coze credentials early.
    CozeWorkflowClient.from_env()

    round1_path, round1_report = _find_round1_report()
    round1_by_id: Dict[str, Dict[str, Any]] = {}
    for r in round1_report.get("results", []) or []:
        if r and r.get("id"):
            round1_by_id[str(r.get("id"))] = _slim_case_result(r)

    workers = int(os.getenv("INSTANT_NEWS_WORKERS") or "8")

    # Run rounds 2 and 3 (interleaved) for maximum throughput.
    t0 = time.time()
    live_run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    live_ts = _now_ts()
    live = _run_live_rounds(
        workflow_id=workflow_id,
        cases=cases,
        round_names=["R2", "R3"],
        max_workers=workers,
    )
    round2_by_id = live.get("R2") or {}
    round3_by_id = live.get("R3") or {}
    elapsed_s = round(time.time() - t0, 3)

    rounds_list = [
        {
            "name": "R1",
            "run_id": (round1_report.get("run") or {}).get("run_id") or "(unknown)",
            "timestamp": (round1_report.get("run") or {}).get("timestamp") or "",
            "source": "existing",
            "summary": _compute_round_summary(round1_by_id, case_ids),
        },
        {
            "name": "R2",
            "run_id": f"{live_run_id}_r2",
            "timestamp": live_ts,
            "source": "live",
            "summary": _compute_round_summary(round2_by_id, case_ids),
        },
        {
            "name": "R3",
            "run_id": f"{live_run_id}_r3",
            "timestamp": live_ts,
            "source": "live",
            "summary": _compute_round_summary(round3_by_id, case_ids),
        },
    ]

    total_rounds = 3
    merged_results: List[Dict[str, Any]] = []
    for case in cases:
        cid = str(case.get("id") or "").strip()
        if not cid:
            continue

        r1 = round1_by_id.get(cid) or {"id": cid, "status": "ERROR", "status_cn": "⚪异常"}
        r2 = round2_by_id.get(cid) or {"id": cid, "status": "ERROR", "status_cn": "⚪异常"}
        r3 = round3_by_id.get(cid) or {"id": cid, "status": "ERROR", "status_cn": "⚪异常"}
        rr = [r1, r2, r3]

        statuses = [str(x.get("status") or "") for x in rr]
        passed_rounds = len([s for s in statuses if s == "PASS"])
        failed_rounds = total_rounds - passed_rounds
        round_result = " ".join([f"R{i+1}{_round_mark(statuses[i])}" for i in range(total_rounds)])

        avg_dur = _avg_duration(rr)

        last = r3 if r3 else (r2 if r2 else r1)

        overall_pass = passed_rounds == total_rounds
        status_cn = "✅通过" if overall_pass else "❌失败"
        fix_status = "⚪不适用" if overall_pass else "🔴待修复"

        merged_results.append(
            {
                "id": cid,
                "intent": str(case.get("intent") or "").strip(),
                "scenario": str(case.get("scenario") or "").strip(),
                "params": case.get("inputs") or {},
                "acceptance": ((case.get("expect") or {}).get("acceptance") or ""),
                "round_result": round_result,
                "passed_rounds": passed_rounds,
                "failed_rounds": failed_rounds,
                "stability": _stability_label(passed_rounds, total_rounds),
                "indicator": str(last.get("indicator") or ""),
                "status_cn": status_cn,
                "issue": _merge_round_field(rr, "issue"),
                "root_cause": _merge_round_field(rr, "root_cause"),
                "suggestion": _merge_round_field(rr, "suggestion"),
                "fix_status": fix_status,
                "avg_duration_s": avg_dur,
                "manual_review": bool(case.get("manual_review")),
                "remarks": _merge_remarks([r.get("remarks") or "" for r in rr]),
                "output_preview": str(last.get("output_preview") or ""),
            }
        )

    # Overall stats
    round_pass_rates: List[float] = []
    for r in rounds_list:
        pr = str((r.get("summary") or {}).get("pass_rate") or "").replace("%", "")
        try:
            round_pass_rates.append(float(pr))
        except Exception:
            continue
    avg_pass_rate = round(sum(round_pass_rates) / max(len(round_pass_rates), 1), 1)

    total_cases = len([c for c in cases if str(c.get("id") or "").strip()])
    stable_pass = len([r for r in merged_results if r.get("passed_rounds") == total_rounds])
    unstable_or_fail = total_cases - stable_pass

    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    report_json_path = REPORT_DIR / f"instant_news_stability_report_{run_id}.json"
    report_md_path = REPORT_DIR / f"instant_news_stability_report_{run_id}.md"
    report_xlsx_path = REPORT_DIR / f"instant_news_stability_test_data_{run_id}.xlsx"
    report_json_latest_path = REPORT_DIR / "instant_news_stability_report.json"
    report_md_latest_path = REPORT_DIR / "instant_news_stability_report.md"
    report_xlsx_latest_path = REPORT_DIR / "instant_news_stability_test_data.xlsx"

    # Intent stats
    intent_stats: Dict[str, Dict[str, Any]] = {}
    for r in merged_results:
        intent = str(r.get("intent") or "(未分类)")
        if intent not in intent_stats:
            intent_stats[intent] = {
                "total": 0,
                "stable_pass": 0,
                "flaky": 0,
                "always_fail": 0,
                "total_passes": 0,
            }
        intent_stats[intent]["total"] += 1
        intent_stats[intent]["total_passes"] += int(r.get("passed_rounds") or 0)
        if r.get("passed_rounds") == total_rounds:
            intent_stats[intent]["stable_pass"] += 1
        elif r.get("passed_rounds") == 0:
            intent_stats[intent]["always_fail"] += 1
        else:
            intent_stats[intent]["flaky"] += 1

    for intent, stats in intent_stats.items():
        t = stats.get("total") or 0
        total_passes = stats.get("total_passes") or 0
        rate = round((total_passes / max(t * total_rounds, 1)) * 100.0, 1)
        stats["avg_pass_rate"] = f"{rate}%"

    report: Dict[str, Any] = {
        "suite": suite_meta,
        "run": {
            "run_id": run_id,
            "timestamp": _now_ts(),
            "rounds": total_rounds,
            "workers": workers,
            "round1_report": str(round1_path),
            "elapsed_s": elapsed_s,
        },
        "rounds": rounds_list,
        "overall": {
            "avg_pass_rate": f"{avg_pass_rate}%",
            "total_cases": total_cases,
            "stable_pass": stable_pass,
            "unstable_or_fail": unstable_or_fail,
        },
        "intent_stats": intent_stats,
        "results": merged_results,
    }

    _save_json(report_json_path, report)
    _save_json(report_json_latest_path, report)

    md = _render_md(report)
    report_md_path.write_text(md, encoding="utf-8")
    report_md_latest_path.write_text(md, encoding="utf-8")

    _make_excel(report, xlsx_path=report_xlsx_path)
    _make_excel(report, xlsx_path=report_xlsx_latest_path)

    base._safe_print(f"Stability Report JSON: {report_json_path}")
    base._safe_print(f"Stability Report MD: {report_md_path}")
    base._safe_print(f"Stability Report XLSX: {report_xlsx_path}")

    return 0 if unstable_or_fail == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
