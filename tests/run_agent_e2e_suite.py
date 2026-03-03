"""Run Agent E2E suite against production Coze bot chat API.

Usage:
  python tests/run_agent_e2e_suite.py

This runner validates five dimensions:
- D1 intent routing
- D2 parameter completeness
- D3 workflow execution (tool output)
- D4 response quality (1-5 score)
- D5 interaction quality (1-5 score)

Important:
- This script calls real Coze API.
- Do NOT run without explicit approval if call quota is constrained.
"""

from __future__ import annotations

import json
import os
import statistics
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from tests.coze_chat_client import CozeChatClient, pick_primary_call
except Exception:
    from coze_chat_client import CozeChatClient, pick_primary_call


SUITE_PATH = Path(__file__).resolve().parent / 'agent_e2e_suite.json'
REPORT_DIR = Path(__file__).resolve().parent / 'reports' / 'agent_e2e'

STRICT_KEYWORD_SKILLS = <REDACTED>


def _safe_print(s: str) -> None:
    try:
        print(s)
    except UnicodeEncodeError:
        print(s.encode('utf-8', errors='replace').decode('utf-8', errors='replace'))


def _safe_json_loads(text: str) -> Optional[Any]:
    if not isinstance(text, str):
        return None
    s = text.strip()
    if not s:
        return None
    try:
        return json.loads(s)
    except Exception:
        return None


def _now_ts() -> str:
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def _bool_env(name: str, default: bool) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in ('1', 'true', 'yes', 'on')


def _normalize_text(s: Any) -> str:
    return str(s or '').strip()


def _normalize_action(s: Any) -> str:
    text = _normalize_text(s)
    mapping = {
        '查询': '查询订阅',
        '查询订阅': '查询订阅',
        '修改': '修改订阅',
        '修改订阅': '修改订阅',
        '新增订阅': '修改订阅',
        '取消': '取消订阅',
        '取消订阅': '取消订阅',
        '退订': '取消订阅',
    }
    return mapping.get(text, text)


def _normalize_keyword(s: Any) -> str:
    text = _normalize_text(s)
    lower = text.lower()
    if lower in ('ai', '人工智能', '大模型', 'gpt', 'llm'):
        return 'AI'
    if lower in ('财经', '金融', '股票', '投资', 'finance'):
        return '财经'
    if lower in ('科技', '技术', 'tech'):
        return '科技'
    return text


def _to_compact_json(obj: Any) -> str:
    try:
        return json.dumps(obj, ensure_ascii=False, sort_keys=True)
    except Exception:
        return str(obj)


def _percentile(values: List[int], p: int) -> Optional[int]:
    if not values:
        return None
    data = sorted(values)
    idx = int(round((p / 100.0) * (len(data) - 1)))
    idx = max(0, min(len(data) - 1, idx))
    return data[idx]


def _contains_error_marker(text: str) -> bool:
    s = _normalize_text(text).lower()
    if not s:
        return False
    markers = [
        'exception',
        'traceback',
        'error',
        'failed',
        'failure',
        '超时',
        '异常',
        '失败',
    ]
    return any(m in s for m in markers)


def _extract_tool_outputs(messages: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    outputs: List[Dict[str, Any]] = []
    for msg in messages:
        if not isinstance(msg, dict):
            continue
        msg_type = _normalize_text(msg.get('type'))
        if msg_type not in ('tool_output', 'tool_response', 'function_response'):
            continue
        raw_content = msg.get('content')
        payload: Any = raw_content
        text = ''
        if isinstance(raw_content, str):
            text = raw_content
            parsed = _safe_json_loads(raw_content)
            if parsed is not None:
                payload = parsed
        elif isinstance(raw_content, dict):
            payload = raw_content
            text = _to_compact_json(raw_content)
        elif isinstance(raw_content, list):
            payload = raw_content
            text = _to_compact_json(raw_content)
        else:
            text = _normalize_text(raw_content)

        name = _normalize_text(msg.get('name'))
        if not name and isinstance(payload, dict):
            name = _normalize_text(payload.get('name') or payload.get('tool') or payload.get('skill'))

        outputs.append(
            {
                'type': msg_type,
                'name': name,
                'payload': payload,
                'text': text,
            }
        )
    return outputs


def _tool_output_ok(item: Dict[str, Any]) -> bool:
    payload = item.get('payload')
    text = _normalize_text(item.get('text'))
    if isinstance(payload, dict):
        success = payload.get('success')
        if success is False:
            return False
        code = payload.get('code')
        if code is not None:
            code_text = _normalize_text(code).lower()
            if code_text and code_text not in ('0', 'ok', 'success'):
                return False
        status = _normalize_text(payload.get('status')).lower()
        if status and status in ('failed', 'error', 'timeout'):
            return False
        err_msg = _normalize_text(payload.get('error') or payload.get('message'))
        if err_msg and _contains_error_marker(err_msg):
            return False
        return True
    if text and _contains_error_marker(text):
        return False
    return bool(text)


def _get_actual_param(params: Dict[str, Any], key: str) -> str:
    alias_map = {
        'keyword': ['keyword', 'new_keyword', 'input_keyword'],
        'raw_query': ['raw_query'],
        'new_email': ['new_email', 'email', 'input_email', 'ex_email'],
        'action': ['action'],
        'input_user_id': ['input_user_id', 'user_id'],
    }
    keys = alias_map.get(key, [key])
    for k in keys:
        if k in params and params[k] is not None:
            return _normalize_text(params[k])
    return ''


def _match_expected_param(
    key: str,
    expected: Any,
    params: Dict[str, Any],
    *,
    actual_skill: str = '',
) -> bool:
    actual = _get_actual_param(params, key)

    def _match_one(exp: Any) -> bool:
        if key == 'keyword':
            # send-mail skills require strict keyword matching (case-sensitive).
            if _normalize_text(actual_skill) in STRICT_KEYWORD_SKILLS:
                return actual == _normalize_text(exp)
            return _normalize_keyword(actual) == _normalize_keyword(exp)
        if key == 'action':
            return _normalize_action(actual) == _normalize_action(exp)
        if key in ('new_email', 'input_user_id'):
            return actual.lower() == _normalize_text(exp).lower()
        return actual == _normalize_text(exp)

    if isinstance(expected, list):
        return any(_match_one(e) for e in expected)
    return _match_one(expected)


def _derive_actual_behavior(actual_skill: str) -> str:
    return '调用工作流' if actual_skill else '未调用工作流'


def _check_include_exclude(answer: str, include_words: List[Any], exclude_words: List[Any]) -> Tuple[bool, int, int, List[str]]:
    text = _normalize_text(answer)
    include = [_normalize_text(w) for w in include_words if _normalize_text(w)]
    exclude = [_normalize_text(w) for w in exclude_words if _normalize_text(w)]
    include_hit = sum(1 for w in include if w in text)
    forbidden_hit = [w for w in exclude if w in text]
    behavior_ok = True
    if include and include_hit == 0:
        behavior_ok = False
    if forbidden_hit:
        behavior_ok = False
    return behavior_ok, include_hit, len(include), forbidden_hit


def _score_d4(answer: str, include_hit: int, include_total: int, forbidden_hit: List[str]) -> int:
    text = _normalize_text(answer)
    if not text:
        return 1
    if _contains_error_marker(text):
        return 1
    score = 3
    if len(text) < 16:
        score = 2
    if len(text) >= 40:
        score += 1
    if text.endswith(('。', '！', '？', '.', '!', '?')):
        score += 1
    if include_total > 0 and include_hit == include_total:
        score += 1
    if forbidden_hit:
        score = min(score, 2)
    return max(1, min(5, score))


def _score_d5(
    *,
    expected_behavior: str,
    actual_skill: str,
    include_hit: int,
    include_total: int,
    forbidden_hit: List[str],
) -> int:
    if expected_behavior in ('追问', '拒绝', '直接回复') and actual_skill:
        return 1
    score = 3
    if include_total > 0:
        if include_hit == 0:
            score = 1
        elif include_hit < include_total:
            score = 3
        else:
            score = 4
    if forbidden_hit:
        return 1
    if expected_behavior in ('追问', '拒绝') and include_total > 0 and include_hit == include_total:
        score += 1
    return max(1, min(5, score))


def _evaluate_case(
    case: Dict[str, Any],
    *,
    actual_skill: str,
    actual_params: Dict[str, Any],
    final_answer: str,
    expected_user_id: str,
    tool_outputs: List[Dict[str, Any]],
) -> Dict[str, Any]:
    expected = case.get('预期结果') or {}
    dimensions = case.get('评测维度') or {}

    expected_behavior = _normalize_text(expected.get('预期行为'))
    expected_skill = _normalize_text(expected.get('预期技能'))
    expected_action = _normalize_text(expected.get('预期动作'))
    expected_params = expected.get('预期参数') or {}
    include_words = expected.get('回复应包含') or []
    exclude_words = expected.get('回复不应包含') or []

    failures: List[str] = []

    # D1: intent routing
    d1_ok = True
    if expected_behavior == '调用工作流':
        if not actual_skill:
            d1_ok = False
            failures.append('预期应调用工作流，但实际未调用')
        elif expected_skill and actual_skill != expected_skill:
            d1_ok = False
            failures.append(f'技能不匹配: 预期={expected_skill}, 实际={actual_skill}')
    else:
        if actual_skill:
            d1_ok = False
            failures.append(f'预期不调用工作流，但实际调用了 {actual_skill}')

    # D2: parameter completeness
    d2_ok = True
    if isinstance(expected_params, dict) and expected_params:
        for key, exp in expected_params.items():
            if not _match_expected_param(key, exp, actual_params, actual_skill=actual_skill):
                d2_ok = False
                failures.append(f'参数不匹配: {key}, 预期={exp}, 实际={_get_actual_param(actual_params, key)}')
    if expected_action:
        if not _match_expected_param('action', expected_action, actual_params, actual_skill=actual_skill):
            d2_ok = False
            failures.append(
                f"动作不匹配: 预期={expected_action}, 实际={_get_actual_param(actual_params, 'action')}"
            )
    if expected_behavior == '调用工作流' and bool(dimensions.get('D2_参数完整')):
        # get_instant_news_using no longer requires input_user_id.
        # Keep strict user_id checks for workflow skills that still require it.
        skill_for_user_id_check = expected_skill or actual_skill
        if skill_for_user_id_check in ('subscribe_using', 'send_mail_using', 'send_mail_cache_using'):
            actual_user_id = _get_actual_param(actual_params, 'input_user_id')
            if not actual_user_id:
                d2_ok = False
                failures.append('调用参数中未识别到 input_user_id/user_id')
            elif actual_user_id != expected_user_id:
                d2_ok = False
                failures.append(f'user_id 不匹配: 预期={expected_user_id}, 实际={actual_user_id}')

    # D3: workflow execution
    d3_ok = True
    if expected_behavior == '调用工作流' and bool(dimensions.get('D3_工作流执行')):
        if not tool_outputs:
            d3_ok = False
            failures.append('调用工作流后未捕获到 tool_output/tool_response')
        else:
            ok_outputs = [o for o in tool_outputs if _tool_output_ok(o)]
            if not ok_outputs:
                d3_ok = False
                failures.append('工作流返回疑似失败（tool_output含错误标记）')

    # shared behavior checks
    behavior_ok, include_hit, include_total, forbidden_hit = _check_include_exclude(
        final_answer,
        include_words,
        exclude_words,
    )
    if not behavior_ok:
        if include_total > 0 and include_hit == 0:
            failures.append(f'回复未命中期望关键词: {include_words}')
        if forbidden_hit:
            failures.append(f'回复命中禁用关键词: {forbidden_hit}')

    # D4 / D5 scoring
    d4_score = _score_d4(final_answer, include_hit, include_total, forbidden_hit)
    d4_ok = d4_score >= 3
    d5_score = _score_d5(
        expected_behavior=expected_behavior,
        actual_skill=actual_skill,
        include_hit=include_hit,
        include_total=include_total,
        forbidden_hit=forbidden_hit,
    )
    d5_ok = d5_score >= 3 and behavior_ok

    active_checks: List[bool] = []
    if bool(dimensions.get('D1_意图路由')):
        active_checks.append(d1_ok)
    if bool(dimensions.get('D2_参数完整')):
        active_checks.append(d2_ok)
    if bool(dimensions.get('D3_工作流执行')):
        active_checks.append(d3_ok)
    if bool(dimensions.get('D4_回复质量')):
        active_checks.append(d4_ok)
    if bool(dimensions.get('D5_交互规范')):
        active_checks.append(d5_ok)

    overall_pass = all(active_checks) if active_checks else False

    return {
        'D1通过': d1_ok,
        'D2通过': d2_ok,
        'D3通过': d3_ok,
        'D4评分': d4_score,
        'D4通过': d4_ok,
        'D5评分': d5_score,
        'D5通过': d5_ok,
        '行为正确': behavior_ok,
        '是否通过': overall_pass,
        '失败原因': '；'.join(failures),
        '预期行为': expected_behavior,
        '实际行为': _derive_actual_behavior(actual_skill),
    }


def _rate(num: int, den: int) -> Optional[float]:
    if den <= 0:
        return None
    return round((num / den) * 100, 2)


def _calc_summary(results: List[Dict[str, Any]], thresholds: Dict[str, Any]) -> Dict[str, Any]:
    total = len(results)
    passed = sum(1 for r in results if r.get('是否通过'))
    failed = total - passed

    d1_pool = [r for r in results if r.get('评测维度', {}).get('D1_意图路由')]
    d2_pool = [r for r in results if r.get('评测维度', {}).get('D2_参数完整')]
    d3_pool = [r for r in results if r.get('评测维度', {}).get('D3_工作流执行')]
    d4_pool = [r for r in results if r.get('评测维度', {}).get('D4_回复质量')]
    d5_pool = [r for r in results if r.get('评测维度', {}).get('D5_交互规范')]

    d4_scores = [int(r.get('D4评分')) for r in d4_pool if isinstance(r.get('D4评分'), int)]
    d5_scores = [int(r.get('D5评分')) for r in d5_pool if isinstance(r.get('D5评分'), int)]

    first_values: List[int] = []
    total_values: List[int] = []
    for r in results:
        first_v = r.get('首字响应ms')
        total_v = r.get('总耗时ms')
        if isinstance(first_v, int):
            first_values.append(first_v)
        if isinstance(total_v, int):
            total_values.append(total_v)

    single_pool = [r for r in results if r.get('对话类型') == '单轮']
    multi_pool = [r for r in results if r.get('对话类型') == '多轮']

    summary = {
        '总用例': total,
        '通过': passed,
        '失败': failed,
        '通过率': _rate(passed, total),
        'D1_意图路由': _rate(sum(1 for r in d1_pool if r.get('D1通过')), len(d1_pool)),
        'D2_参数完整': _rate(sum(1 for r in d2_pool if r.get('D2通过')), len(d2_pool)),
        'D3_工作流执行': _rate(sum(1 for r in d3_pool if r.get('D3通过')), len(d3_pool)),
        'D4_回复质量_平均分': round(statistics.mean(d4_scores), 2) if d4_scores else None,
        'D5_交互规范': _rate(sum(1 for r in d5_pool if r.get('D5通过')), len(d5_pool)),
        'D5_交互规范_平均分': round(statistics.mean(d5_scores), 2) if d5_scores else None,
        '首字响应ms': {
            '均值': round(statistics.mean(first_values), 2) if first_values else None,
            'P50': _percentile(first_values, 50),
            'P90': _percentile(first_values, 90),
        },
        '总耗时ms': {
            '均值': round(statistics.mean(total_values), 2) if total_values else None,
            'P50': _percentile(total_values, 50),
            'P90': _percentile(total_values, 90),
        },
        '单轮通过率': _rate(sum(1 for r in single_pool if r.get('是否通过')), len(single_pool)),
        '多轮通过率': _rate(sum(1 for r in multi_pool if r.get('是否通过')), len(multi_pool)),
    }

    gate_rows = {
        '整体通过率': _rate(passed, total),
        'D1_意图路由': summary.get('D1_意图路由'),
        'D2_参数完整': summary.get('D2_参数完整'),
        'D3_工作流执行': summary.get('D3_工作流执行'),
        'D4_回复质量_平均分': summary.get('D4_回复质量_平均分'),
        'D5_交互规范': summary.get('D5_交互规范'),
    }
    gate: Dict[str, Dict[str, Any]] = {}
    for key, actual in gate_rows.items():
        threshold = thresholds.get(key)
        status = '未知'
        if isinstance(actual, (int, float)) and isinstance(threshold, (int, float)):
            status = '达标' if actual >= threshold else '未达标'
        gate[key] = {
            'actual': actual,
            'threshold': threshold,
            'status': status,
        }
    summary['验收门槛'] = gate
    return summary


def _render_md(report: Dict[str, Any]) -> str:
    run = report.get('运行信息') or {}
    summary = report.get('汇总') or {}
    results = report.get('结果') or []

    lines: List[str] = []
    lines.append('# Agent E2E 测试报告')
    lines.append('')
    lines.append(f"- Run ID: `{run.get('run_id')}`")
    lines.append(f"- 时间: `{run.get('timestamp')}`")
    lines.append(f"- Bot ID: `{run.get('bot_id')}`")
    lines.append(f"- 流式首响: `{run.get('use_stream')}`")
    lines.append('')
    lines.append('## 汇总')
    lines.append('')
    for k in [
        '总用例',
        '通过',
        '失败',
        '通过率',
        'D1_意图路由',
        'D2_参数完整',
        'D3_工作流执行',
        'D4_回复质量_平均分',
        'D5_交互规范',
        'D5_交互规范_平均分',
        '单轮通过率',
        '多轮通过率',
    ]:
        lines.append(f'- {k}: {summary.get(k)}')

    gate = summary.get('验收门槛') or {}
    lines.append('')
    lines.append('## 验收门槛')
    lines.append('')
    for key, row in gate.items():
        lines.append(f"- {key}: 实际={row.get('actual')} / 门槛={row.get('threshold')} / 状态={row.get('status')}")

    failed = [r for r in results if not r.get('是否通过')]
    if failed:
        lines.append('')
        lines.append('## 失败用例')
        lines.append('')
        for r in failed:
            lines.append(f"### {r.get('用例ID')} {r.get('测试场景')}")
            lines.append(f"- 对话类型: {r.get('对话类型')} 会话组: {r.get('会话组ID') or '-'} 轮次: {r.get('轮次')}")
            lines.append(f"- 用户输入: {r.get('用户输入')}")
            lines.append(f"- 预期技能/动作: {r.get('预期结果', {}).get('预期技能')} / {r.get('预期结果', {}).get('预期动作')}")
            lines.append(f"- 实际技能: {r.get('实际技能')}")
            lines.append(f"- D1/D2/D3: {r.get('D1通过')} / {r.get('D2通过')} / {r.get('D3通过')}")
            lines.append(f"- D4/D5: {r.get('D4评分')} / {r.get('D5评分')}")
            lines.append(f"- 失败原因: {r.get('失败原因')}")
            lines.append('')
    return '\n'.join(lines)


def _write_table(ws, headers: List[str], rows: List[List[Any]]) -> None:
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1F4E78')
        c.alignment = Alignment(vertical='top', horizontal='center', wrap_text=True)
    for r in rows:
        ws.append(r)


def _apply_zebra_rows(ws) -> None:
    fill_even = PatternFill('solid', fgColor='F8FAFC')
    fill_odd = PatternFill('solid', fgColor='FFFFFF')
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = fill_even if idx % 2 == 0 else fill_odd
        for cell in row:
            cell.fill = fill


def _apply_pass_fail_fills(ws, pass_col: int) -> None:
    pass_fill = PatternFill('solid', fgColor='DCFCE7')
    fail_fill = PatternFill('solid', fgColor='FEE2E2')
    for row in ws.iter_rows(min_row=2):
        cell = row[pass_col - 1]
        val = str(cell.value or '')
        if val in ('True', 'TRUE', 'true', '通过'):
            cell.fill = pass_fill
        elif val in ('False', 'FALSE', 'false', '失败'):
            cell.fill = fail_fill


def _make_summary_sheet(wb: Workbook, report: Dict[str, Any]) -> None:
    ws = wb.active
    ws.title = '测试汇总'
    run = report.get('运行信息') or {}
    summary = report.get('汇总') or {}
    gate = summary.get('验收门槛') or {}

    rows: List[List[Any]] = [
        ['执行时间', run.get('timestamp')],
        ['Run ID', run.get('run_id')],
        ['Bot ID', run.get('bot_id')],
        ['流式首响', run.get('use_stream')],
        ['总用例', summary.get('总用例')],
        ['通过', summary.get('通过')],
        ['失败', summary.get('失败')],
        ['通过率', summary.get('通过率')],
        ['D1_意图路由', summary.get('D1_意图路由')],
        ['D2_参数完整', summary.get('D2_参数完整')],
        ['D3_工作流执行', summary.get('D3_工作流执行')],
        ['D4_回复质量_平均分', summary.get('D4_回复质量_平均分')],
        ['D5_交互规范', summary.get('D5_交互规范')],
        ['D5_交互规范_平均分', summary.get('D5_交互规范_平均分')],
        ['首字响应ms', _to_compact_json(summary.get('首字响应ms'))],
        ['总耗时ms', _to_compact_json(summary.get('总耗时ms'))],
        ['单轮通过率', summary.get('单轮通过率')],
        ['多轮通过率', summary.get('多轮通过率')],
        ['', ''],
        ['验收指标', '实际值', '门槛', '状态'],
    ]

    _write_table(ws, ['指标', '值'], [[r[0], r[1]] for r in rows if len(r) >= 2 and r[0] != '验收指标'])
    start_row = ws.max_row + 2
    ws.cell(row=start_row, column=1, value='验收指标').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=start_row, column=2, value='实际值').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=start_row, column=3, value='门槛').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=start_row, column=4, value='状态').font = Font(bold=True, color='FFFFFF')
    for c in range(1, 5):
        ws.cell(row=start_row, column=c).fill = PatternFill('solid', fgColor='1F4E78')
        ws.cell(row=start_row, column=c).alignment = Alignment(vertical='top', horizontal='center', wrap_text=True)

    row = start_row + 1
    for key in ['整体通过率', 'D1_意图路由', 'D2_参数完整', 'D3_工作流执行', 'D4_回复质量_平均分', 'D5_交互规范']:
        g = gate.get(key) or {}
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=g.get('actual'))
        ws.cell(row=row, column=3, value=g.get('threshold'))
        ws.cell(row=row, column=4, value=g.get('status'))
        row += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.freeze_panes = 'A2'


def _make_detail_sheet(wb: Workbook, report: Dict[str, Any]) -> None:
    ws = wb.create_sheet('测试明细')
    headers = [
        '用例ID',
        '意图分类',
        '二级意图',
        '测试场景',
        '对话类型',
        '会话组ID',
        '轮次',
        '用户输入',
        '预期行为',
        '预期技能',
        '预期动作',
        '预期参数',
        '实际行为',
        '实际技能',
        '实际参数',
        'tool_output摘要',
        '首字响应ms',
        '总耗时ms',
        'D1通过',
        'D2通过',
        'D3通过',
        'D4评分',
        'D5评分',
        '是否通过',
        '失败原因',
        '修复状态',
        '备注',
    ]

    rows: List[List[Any]] = []
    for r in report.get('结果', []) or []:
        exp = r.get('预期结果') or {}
        tool_outputs = r.get('tool_outputs') or []
        tool_preview = ''
        if tool_outputs:
            first = tool_outputs[0]
            tool_preview = _normalize_text(first.get('name')) or _normalize_text(first.get('type'))
            text = _normalize_text(first.get('text'))
            if text:
                tool_preview = f'{tool_preview}: {text[:120]}'
        rows.append(
            [
                r.get('用例ID'),
                r.get('意图分类'),
                r.get('二级意图'),
                r.get('测试场景'),
                r.get('对话类型'),
                r.get('会话组ID'),
                r.get('轮次'),
                r.get('用户输入'),
                exp.get('预期行为'),
                exp.get('预期技能'),
                exp.get('预期动作'),
                _to_compact_json(exp.get('预期参数') or {}),
                r.get('实际行为'),
                r.get('实际技能'),
                _to_compact_json(r.get('实际参数') or {}),
                tool_preview,
                r.get('首字响应ms'),
                r.get('总耗时ms'),
                r.get('D1通过'),
                r.get('D2通过'),
                r.get('D3通过'),
                r.get('D4评分'),
                r.get('D5评分'),
                r.get('是否通过'),
                r.get('失败原因'),
                r.get('修复状态'),
                r.get('备注'),
            ]
        )

    _write_table(ws, headers, rows)
    wrap = Alignment(wrap_text=True, vertical='top')
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    widths = {
        'A': 10,
        'B': 12,
        'C': 14,
        'D': 22,
        'E': 9,
        'F': 10,
        'G': 7,
        'H': 32,
        'I': 10,
        'J': 18,
        'K': 12,
        'L': 30,
        'M': 10,
        'N': 18,
        'O': 32,
        'P': 40,
        'Q': 10,
        'R': 10,
        'S': 8,
        'T': 8,
        'U': 8,
        'V': 8,
        'W': 8,
        'X': 8,
        'Y': 40,
        'Z': 12,
        'AA': 20,
    }
    for c, w in widths.items():
        ws.column_dimensions[c].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'
    _apply_zebra_rows(ws)
    _apply_pass_fail_fills(ws, pass_col=24)


def _append_section(ws, title: str, headers: List[str], rows: List[List[Any]]) -> None:
    ws.append([title])
    title_row = ws.max_row
    ws.cell(row=title_row, column=1).font = Font(bold=True)
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E78')
        cell.alignment = Alignment(vertical='top', horizontal='center', wrap_text=True)
    for r in rows:
        ws.append(r)
    ws.append([])


def _make_standard_sheet(wb: Workbook, report: Dict[str, Any]) -> None:
    ws = wb.create_sheet('验收标准')
    suite_dims = report.get('评测维度') or {}
    thresholds = report.get('验收门槛定义') or {}

    rows_a: List[List[Any]] = []
    for key in ['D1_意图路由', 'D2_参数完整', 'D3_工作流执行', 'D4_回复质量', 'D5_交互规范']:
        d = suite_dims.get(key) or {}
        rows_a.append([key, d.get('权重'), d.get('评判方式'), d.get('通过标准'), d.get('说明')])
    _append_section(ws, '表A：维度定义与权重', ['维度', '权重', '评判方式', '通过标准', '说明'], rows_a)

    rows_b = [
        [5, '信息完整、表达清晰、无错误、可直接使用'],
        [4, '信息较完整，存在轻微措辞或结构瑕疵'],
        [3, '基本可用，但缺少关键信息或可读性一般'],
        [2, '质量较差，信息缺失明显或表达混乱'],
        [1, '不可用：空回复、严重错误、明显幻觉'],
    ]
    _append_section(ws, '表B：D4 回复质量评分量表（1-5）', ['分值', '判定标准'], rows_b)

    rows_c = [
        [5, '追问/拒绝精准且语气友好，关键词命中充分'],
        [4, '交互方向正确，关键词覆盖较好'],
        [3, '交互基本正确但信息不充分'],
        [2, '交互方向偏差明显'],
        [1, '严重违规：应追问却调用工作流，或拒绝失当'],
    ]
    _append_section(ws, '表C：D5 交互规范评分量表（1-5）', ['分值', '判定标准'], rows_c)

    rows_d = [
        ['单轮-调用工作流', 'D1 && D2 && D3 && D4>=3'],
        ['单轮-追问/拒绝/直接回复', 'D1 && D5>=3'],
        ['多轮-中间补槽轮', 'D1 && D5>=3'],
        ['多轮-最终执行轮', 'D1 && D2 && D3 && D4>=3'],
        ['send_mail 场景', '以 tool_output 成功判定 D3，不要求邮箱实达自动校验'],
    ]
    _append_section(ws, '表D：场景通过判定规则', ['场景类型', '通过条件'], rows_d)

    rows_e = []
    for key in ['整体通过率', '稳定性通过率', 'D1_意图路由', 'D2_参数完整', 'D3_工作流执行', 'D4_回复质量_平均分', 'D5_交互规范']:
        rows_e.append([key, thresholds.get(key)])
    _append_section(ws, '表E：验收门槛', ['指标', '门槛'], rows_e)

    rows_f = [
        ['D1_意图路由', '30', '路由错误会导致后续全部偏离，失败代价最高'],
        ['D2_参数完整', '20', '路径正确但参数错误会导致业务结果错误'],
        ['D3_工作流执行', '20', 'E2E 需要区分调度正确与下游执行失败'],
        ['D4_回复质量', '15', '功能正确之外衡量用户体验，影响次于路由与执行'],
        ['D5_交互规范', '15', '追问与拒绝质量影响多轮可持续性与安全性'],
    ]
    _append_section(ws, '表F：权重分配依据（必须记录）', ['维度', '权重', '分配理由'], rows_f)

    rows_g = [
        ['面向 C 端体验优先', 'D4 +5, D3 -5'],
        ['多轮补槽占比高', 'D5 +5, D4 -5'],
        ['下游工作流不稳定', 'D3 +5, D2 -5'],
        ['纯路由场景（无下游）', 'D1 +5, D3 -5'],
    ]
    _append_section(ws, '表G：权重调整指引', ['场景', '建议调整'], rows_g)

    wrap = Alignment(wrap_text=True, vertical='top')
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            cell.alignment = wrap
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 48
    ws.column_dimensions['E'].width = 48


def _classify_failure(case: Dict[str, Any]) -> Dict[str, str]:
    reason = _normalize_text(case.get('失败原因'))
    if '异常:' in reason:
        return {
            '失败类型': '执行异常',
            '诊断': '运行时异常，通常为网络/API/超时问题。',
            '建议': '检查 token、网络与超时配置。',
        }
    if '技能不匹配' in reason or '预期应调用工作流' in reason or '预期不调用工作流' in reason:
        return {
            '失败类型': '意图路由错误',
            '诊断': '模型未进入正确技能路径。',
            '建议': '强化意图边界与反例。',
        }
    if '参数不匹配' in reason or '动作不匹配' in reason or 'user_id' in reason:
        return {
            '失败类型': '参数提取错误',
            '诊断': '技能已命中但参数不完整或值错误。',
            '建议': '补充槽位示例并校验归一化映射。',
        }
    if 'tool_output' in reason or '工作流返回疑似失败' in reason:
        return {
            '失败类型': '工作流执行失败',
            '诊断': '调度成功但下游执行异常或结果不可用。',
            '建议': '检查 workflow 节点日志与返回结构。',
        }
    if '回复未命中期望关键词' in reason or '回复命中禁用关键词' in reason:
        return {
            '失败类型': '交互/文案不达标',
            '诊断': '回复内容未覆盖关键提示词或触发禁用词。',
            '建议': '优化追问模板与拒绝文案。',
        }
    return {
        '失败类型': '未分类',
        '诊断': '未命中既有分类，需人工复核。',
        '建议': '结合原始消息和 tool_output 进一步定位。',
    }


def _make_failures_sheet(wb: Workbook, report: Dict[str, Any]) -> None:
    ws = wb.create_sheet('失败案例分析')
    headers = [
        '用例ID',
        '测试场景',
        '对话类型',
        '预期技能',
        '实际技能',
        'D1/D2/D3',
        'D4/D5',
        '失败类型',
        '失败原因',
        '诊断',
        '建议',
        '修复状态',
    ]
    rows: List[List[Any]] = []
    for case in report.get('结果', []) or []:
        if case.get('是否通过'):
            continue
        cls = _classify_failure(case)
        expected = case.get('预期结果') or {}
        rows.append(
            [
                case.get('用例ID'),
                case.get('测试场景'),
                case.get('对话类型'),
                expected.get('预期技能'),
                case.get('实际技能'),
                f"{case.get('D1通过')}/{case.get('D2通过')}/{case.get('D3通过')}",
                f"{case.get('D4评分')}/{case.get('D5评分')}",
                cls.get('失败类型'),
                case.get('失败原因'),
                cls.get('诊断'),
                cls.get('建议'),
                case.get('修复状态') or '待处理',
            ]
        )

    _write_table(ws, headers, rows)
    wrap = Alignment(wrap_text=True, vertical='top')
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    widths = {
        'A': 10,
        'B': 24,
        'C': 10,
        'D': 18,
        'E': 18,
        'F': 12,
        'G': 10,
        'H': 16,
        'I': 36,
        'J': 34,
        'K': 34,
        'L': 12,
    }
    for c, w in widths.items():
        ws.column_dimensions[c].width = w
    ws.auto_filter.ref = ws.dimensions
    _apply_zebra_rows(ws)


def _make_iteration_sheet(wb: Workbook, report: Dict[str, Any]) -> None:
    ws = wb.create_sheet('迭代追踪')
    headers = ['用例ID', '测试场景', '基线', '第1轮修复', '第2轮回归', '当前状态', '变更记录']
    rows: List[List[Any]] = []
    for case in report.get('结果', []) or []:
        baseline = 'PASS' if case.get('是否通过') else 'FAIL'
        current_status = '✅稳定' if case.get('是否通过') else '❌未解决'
        rows.append(
            [
                case.get('用例ID'),
                case.get('测试场景'),
                baseline,
                '',
                '',
                current_status,
                '',
            ]
        )
    _write_table(ws, headers, rows)
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 34
    ws.auto_filter.ref = ws.dimensions
    _apply_zebra_rows(ws)


def _make_excel(report: Dict[str, Any], xlsx_path: Path) -> None:
    wb = Workbook()
    _make_summary_sheet(wb, report)
    _make_detail_sheet(wb, report)
    _make_standard_sheet(wb, report)
    _make_failures_sheet(wb, report)
    _make_iteration_sheet(wb, report)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def main() -> int:
    try:
        stdout_reconfigure = getattr(sys.stdout, 'reconfigure', None)
        if callable(stdout_reconfigure):
            stdout_reconfigure(encoding='utf-8', errors='replace')
        stderr_reconfigure = getattr(sys.stderr, 'reconfigure', None)
        if callable(stderr_reconfigure):
            stderr_reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

    if not SUITE_PATH.exists():
        raise SystemExit(f'Suite file not found: {SUITE_PATH}')

    suite = json.loads(SUITE_PATH.read_text(encoding='utf-8'))
    meta = suite.get('测试集') or {}
    default_cfg = suite.get('默认配置') or {}
    thresholds = suite.get('验收门槛') or {}
    cases = suite.get('用例')
    if not isinstance(cases, list) or not cases:
        raise SystemExit('agent_e2e_suite.json 中的 用例 必须为非空数组')

    bot_id = (
        (os.getenv('COZE_AGENT_E2E_BOT_ID') or '').strip()
        or _normalize_text(meta.get('bot_id'))
        or (os.getenv('COZE_BOT_ID') or '').strip()
    )
    if not bot_id:
        raise SystemExit('未找到 bot_id，请设置 COZE_AGENT_E2E_BOT_ID 或在 suite 中提供 bot_id')

    default_user_id = (
        (os.getenv('AGENT_E2E_CONTEXT_USER_ID') or '').strip()
        or _normalize_text(default_cfg.get('测试user_id'))
        or 'agent_e2e_test_user_001'
    )
    use_stream = _bool_env('AGENT_E2E_USE_STREAM', bool(default_cfg.get('是否使用流式首响', True)))
    timeout_s = int(os.getenv('AGENT_E2E_TIMEOUT_S') or '120')
    max_polls = int(os.getenv('AGENT_E2E_MAX_POLLS') or '20')
    poll_interval_s = float(os.getenv('AGENT_E2E_POLL_INTERVAL_S') or '0.8')
    start_index = int(os.getenv('AGENT_E2E_START_INDEX') or '1')
    max_cases = int(os.getenv('AGENT_E2E_MAX_CASES') or '0')
    stop_on_exception = _bool_env('AGENT_E2E_STOP_ON_EXCEPTION', False)

    if start_index < 1:
        start_index = 1
    cases = cases[start_index - 1 :]
    if max_cases > 0:
        cases = cases[:max_cases]

    client = CozeChatClient.from_env()
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    run_id = datetime.now().strftime('%Y%m%d_%H%M%S')
    report_json_path = REPORT_DIR / f'agent_e2e_report_{run_id}.json'
    report_md_path = REPORT_DIR / f'agent_e2e_report_{run_id}.md'
    report_xlsx_path = REPORT_DIR / f'agent_e2e_test_data_{run_id}.xlsx'
    report_json_latest_path = REPORT_DIR / 'agent_e2e_report.json'
    report_md_latest_path = REPORT_DIR / 'agent_e2e_report.md'
    report_xlsx_latest_path = REPORT_DIR / 'agent_e2e_test_data.xlsx'

    report: Dict[str, Any] = {
        '测试集': meta,
        '评测维度': suite.get('评测维度') or {},
        '验收门槛定义': thresholds,
        '运行信息': {
            'run_id': run_id,
            'timestamp': _now_ts(),
            'base_url': os.getenv('COZE_BASE_URL') or os.getenv('COZE_API_BASE') or 'https://api.coze.cn',
            'bot_id': bot_id,
            'use_stream': use_stream,
            'timeout_s': timeout_s,
            'max_polls': max_polls,
            'poll_interval_s': poll_interval_s,
            'start_index': start_index,
            'max_cases': max_cases,
            'stop_on_exception': stop_on_exception,
            'default_user_id': default_user_id,
        },
        '结果': [],
        '汇总': {},
    }

    _safe_print(f'Running agent e2e suite ({len(cases)} cases) at {_now_ts()}')
    _safe_print(f'Bot ID: {bot_id} | stream={use_stream} | timeout_s={timeout_s}')
    _safe_print(
        f'Range: start_index={start_index}, max_cases={max_cases}, stop_on_exception={stop_on_exception}'
    )

    conv_map: Dict[str, Dict[str, Any]] = {}
    aborted = False
    abort_reason = ''

    for idx, case in enumerate(cases, start=1):
        case = dict(case)
        case_id = _normalize_text(case.get('用例ID')) or f'CASE-{idx:03d}'
        dialog_type = _normalize_text(case.get('对话类型')) or '单轮'
        group_id = _normalize_text(case.get('会话组ID'))
        user_input = _normalize_text(case.get('用户输入'))

        if not user_input:
            case_result = dict(case)
            case_result.update(
                {
                    '实际技能': '',
                    '实际参数': {},
                    '实际回复': '',
                    'tool_outputs': [],
                    '首字响应ms': None,
                    '总耗时ms': 0,
                    '会话ID': '',
                    'chat_id': '',
                    'D1通过': False,
                    'D2通过': False,
                    'D3通过': False,
                    'D4评分': 1,
                    'D5评分': 1,
                    '行为正确': False,
                    '实际行为': '未调用工作流',
                    '是否通过': False,
                    '失败原因': '用户输入为空',
                }
            )
            report['结果'].append(case_result)
            _safe_print(f'FAIL {case_id} 输入为空')
            continue

        conversation_id: Optional[str] = None
        user_id = default_user_id
        if dialog_type == '多轮':
            if not group_id:
                case_result = dict(case)
                case_result.update(
                    {
                        '实际技能': '',
                        '实际参数': {},
                        '实际回复': '',
                        'tool_outputs': [],
                        '首字响应ms': None,
                        '总耗时ms': 0,
                        '会话ID': '',
                        'chat_id': '',
                        'D1通过': False,
                        'D2通过': False,
                        'D3通过': False,
                        'D4评分': 1,
                        'D5评分': 1,
                        '行为正确': False,
                        '实际行为': '未调用工作流',
                        '是否通过': False,
                        '失败原因': '多轮用例缺少会话组ID',
                    }
                )
                report['结果'].append(case_result)
                _safe_print(f'FAIL {case_id} 多轮缺少会话组ID')
                continue
            if group_id in conv_map:
                conversation_id = conv_map[group_id]['conversation_id']
                user_id = conv_map[group_id]['user_id']
            else:
                try:
                    created_conversation_id = client.create_conversation(bot_id=bot_id)
                    conversation_id = created_conversation_id
                    conv_map[group_id] = {
                        'conversation_id': created_conversation_id,
                        'user_id': user_id,
                    }
                    _safe_print(f'  [multi-turn] Created conversation {conversation_id} for group {group_id}')
                except Exception as e:
                    case_result = dict(case)
                    case_result.update(
                        {
                            '实际技能': '',
                            '实际参数': {},
                            '实际回复': '',
                            'tool_outputs': [],
                            '首字响应ms': None,
                            '总耗时ms': 0,
                            '会话ID': '',
                            'chat_id': '',
                            'D1通过': False,
                            'D2通过': False,
                            'D3通过': False,
                            'D4评分': 1,
                            'D5评分': 1,
                            '行为正确': False,
                            '实际行为': '未调用工作流',
                            '是否通过': False,
                            '失败原因': f'异常: 创建会话失败: {type(e).__name__}: {e}',
                        }
                    )
                    report['结果'].append(case_result)
                    _safe_print(f'FAIL {case_id} 创建会话失败: {e}')
                    if stop_on_exception:
                        aborted = True
                        abort_reason = f'{case_id}: create_conversation failed: {e}'
                        break
                    continue

        try:
            run = client.chat_once(
                bot_id=bot_id,
                message=user_input,
                user_id=user_id,
                conversation_id=conversation_id,
                use_stream=use_stream,
                max_polls=max_polls,
                poll_interval_s=poll_interval_s,
                timeout_s=timeout_s,
            )
            if run.status and run.status != 'completed':
                raise RuntimeError(f'chat status is {run.status}')

            actual_skill, actual_params = pick_primary_call(
                function_calls=run.function_calls,
                stub_calls=run.stub_calls,
            )
            tool_outputs = _extract_tool_outputs(run.messages)
            eval_result = _evaluate_case(
                case,
                actual_skill=actual_skill,
                actual_params=actual_params,
                final_answer=run.final_answer,
                expected_user_id=user_id,
                tool_outputs=tool_outputs,
            )
            case_result = dict(case)
            case_result.update(
                {
                    '实际技能': actual_skill,
                    '实际参数': actual_params,
                    '实际回复': run.final_answer,
                    'tool_outputs': tool_outputs,
                    '首字响应ms': run.first_response_ms,
                    '总耗时ms': run.total_ms,
                    '会话ID': run.conversation_id,
                    'chat_id': run.chat_id,
                    'function_calls': run.function_calls,
                    'stub_calls': run.stub_calls,
                }
            )
            case_result.update(eval_result)
        except Exception as e:
            case_result = dict(case)
            case_result.update(
                {
                    '实际技能': '',
                    '实际参数': {},
                    '实际回复': '',
                    'tool_outputs': [],
                    '首字响应ms': None,
                    '总耗时ms': 0,
                    '会话ID': conversation_id or '',
                    'chat_id': '',
                    'D1通过': False,
                    'D2通过': False,
                    'D3通过': False,
                    'D4评分': 1,
                    'D5评分': 1,
                    '行为正确': False,
                    '实际行为': '未调用工作流',
                    '是否通过': False,
                    '失败原因': f'异常: {type(e).__name__}: {e}',
                }
            )
            if stop_on_exception:
                report['结果'].append(case_result)
                aborted = True
                abort_reason = f'{case_id}: {type(e).__name__}: {e}'
                _safe_print(f'ABORT {case_id} due to exception: {type(e).__name__}: {e}')
                break

        report['结果'].append(case_result)
        status = 'PASS' if case_result.get('是否通过') else 'FAIL'
        _safe_print(
            f"{status} {case_id} {case_result.get('测试场景')} "
            f"(D4={case_result.get('D4评分')} D5={case_result.get('D5评分')} "
            f"first={case_result.get('首字响应ms')}ms total={case_result.get('总耗时ms')}ms)"
        )

    if aborted:
        report['运行信息']['是否中止'] = True
        report['运行信息']['中止原因'] = abort_reason
    else:
        report['运行信息']['是否中止'] = False

    report['汇总'] = _calc_summary(report['结果'], thresholds)

    json_content = json.dumps(report, ensure_ascii=False, indent=2)
    report_json_path.write_text(json_content, encoding='utf-8')
    report_json_latest_path.write_text(json_content, encoding='utf-8')

    md_content = _render_md(report)
    report_md_path.write_text(md_content, encoding='utf-8')
    report_md_latest_path.write_text(md_content, encoding='utf-8')

    _make_excel(report, report_xlsx_path)
    _make_excel(report, report_xlsx_latest_path)

    _safe_print(f'Report JSON: {report_json_path}')
    _safe_print(f'Report MD: {report_md_path}')
    _safe_print(f'Report XLSX: {report_xlsx_path}')
    _safe_print(f'Latest JSON: {report_json_latest_path}')
    _safe_print(f'Latest MD: {report_md_latest_path}')
    _safe_print(f'Latest XLSX: {report_xlsx_latest_path}')

    if aborted:
        return 2
    return 0 if int(report['汇总'].get('失败') or 0) == 0 else 1


if __name__ == '__main__':
    raise SystemExit(main())
