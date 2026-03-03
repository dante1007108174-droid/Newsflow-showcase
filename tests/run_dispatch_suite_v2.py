"""Run dispatch-model v2 test suite against Coze bot chat API.

Usage:
  python tests/run_dispatch_suite_v2.py

This runner validates:
- Intent routing correctness (which skill/workflow is called)
- Parameter extraction correctness (from STUB response/function_call)
- Follow-up/guardrail behavior (ask clarification, reject, fallback)
- First-response latency (stream mode)

Important:
- This script calls real Coze API.
- By default it uses stream mode to measure first response latency.
"""

from __future__ import annotations

import json
import os
import statistics
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from tests.coze_chat_client import CozeChatClient, pick_primary_call
except Exception:
    from coze_chat_client import CozeChatClient, pick_primary_call


SUITE_PATH = Path(__file__).resolve().parent / "dispatch_suite_v2.json"
REPORT_DIR = Path(__file__).resolve().parent / "reports" / "dispatch_model_v2"

STRICT_KEYWORD_SKILLS = <REDACTED>


def _safe_print(s: str) -> None:
    try:
        print(s)
    except UnicodeEncodeError:
        print(s.encode("utf-8", errors="replace").decode("utf-8", errors="replace"))


def _now_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _bool_env(name: str, default: bool) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in ("1", "true", "yes", "on")


def _normalize_text(s: Any) -> str:
    return str(s or "").strip()


def _normalize_action(s: Any) -> str:
    text = _normalize_text(s)
    mapping = {
        "查询": "查询订阅",
        "查询订阅": "查询订阅",
        "修改": "修改订阅",
        "修改订阅": "修改订阅",
        "新增订阅": "修改订阅",
        "取消": "取消订阅",
        "取消订阅": "取消订阅",
        "退订": "取消订阅",
    }
    return mapping.get(text, text)


def _normalize_keyword(s: Any) -> str:
    text = _normalize_text(s)
    lower = text.lower()
    if lower in ("ai", "人工智能", "大模型", "gpt", "llm"):
        return "AI"
    if lower in ("财经", "金融", "股票", "投资", "finance"):
        return "财经"
    if lower in ("科技", "技术", "tech"):
        return "科技"
    return text


def _to_compact_json(obj: Any) -> str:
    try:
        return json.dumps(obj, ensure_ascii=False, sort_keys=True)
    except Exception:
        return str(obj)


def _percentile(values: List[int], p: int) -> Optional[int]:
    if not values:
        return None
    data = sorted(values)
    idx = int(round((p / 100.0) * (len(data) - 1)))
    idx = max(0, min(len(data) - 1, idx))
    return data[idx]


def _get_actual_param(params: Dict[str, Any], key: str) -> str:
    alias_map = {
        "keyword": ["keyword", "new_keyword", "input_keyword"],
        "new_email": ["new_email", "email", "input_email", "ex_email"],
        "action": ["action"],
        "input_user_id": ["input_user_id", "user_id"],
    }
    keys = alias_map.get(key, [key])
    for k in keys:
        if k in params and params[k] is not None:
            return _normalize_text(params[k])
    return ""


def _match_expected_param(
    key: str,
    expected: Any,
    params: Dict[str, Any],
    *,
    actual_skill: str = "",
) -> bool:
    actual = _get_actual_param(params, key)

    def _match_one(exp: Any) -> bool:
        if key == "keyword":
            # v2 rule: send-mail keyword must preserve exact value, including lowercase "ai".
            if _normalize_text(actual_skill) in STRICT_KEYWORD_SKILLS:
                return actual == _normalize_text(exp)
            return _normalize_keyword(actual) == _normalize_keyword(exp)
        if key == "action":
            return _normalize_action(actual) == _normalize_action(exp)
        if key in ("new_email", "input_user_id"):
            return actual.lower() == _normalize_text(exp).lower()
        return actual == _normalize_text(exp)

    if isinstance(expected, list):
        return any(_match_one(e) for e in expected)
    return _match_one(expected)


def _derive_actual_behavior(actual_skill: str) -> str:
    return "调用工作流" if actual_skill else "未调用工作流"


def _evaluate_case(
    case: Dict[str, Any],
    *,
    actual_skill: str,
    actual_params: Dict[str, Any],
    final_answer: str,
    expected_user_id: str,
) -> Dict[str, Any]:
    expected = case.get("预期结果") or {}
    metrics = case.get("指标") or {}

    expected_behavior = _normalize_text(expected.get("预期行为"))
    expected_skill = _normalize_text(expected.get("预期技能"))
    expected_action = _normalize_text(expected.get("预期动作"))
    expected_params = expected.get("预期参数") or {}
    include_words = expected.get("回复应包含") or []
    exclude_words = expected.get("回复不应包含") or []

    failures: List[str] = []

    # 1) Intent/routing check
    intent_ok = True
    if expected_behavior == "调用工作流":
        if not actual_skill:
            intent_ok = False
            failures.append("预期应调用工作流，但实际未调用")
        elif expected_skill and actual_skill != expected_skill:
            intent_ok = False
            failures.append(f"技能不匹配: 预期={expected_skill}, 实际={actual_skill}")
    else:
        if actual_skill:
            intent_ok = False
            failures.append(f"预期不调用工作流，但实际调用了 {actual_skill}")

    # 2) Parameter checks
    params_ok = True
    if isinstance(expected_params, dict) and expected_params:
        for k, v in expected_params.items():
            if not _match_expected_param(k, v, actual_params, actual_skill=actual_skill):
                params_ok = False
                failures.append(
                    f"参数不匹配: {k}, 预期={v}, 实际={_get_actual_param(actual_params, k)}"
                )

    if expected_action:
        if not _match_expected_param(
            "action", expected_action, actual_params, actual_skill=actual_skill
        ):
            params_ok = False
            failures.append(
                f"动作不匹配: 预期={expected_action}, 实际={_get_actual_param(actual_params, 'action')}"
            )

    # 3) user_id checks (only when enabled)
    user_id_ok = True
    if bool(metrics.get("纳入user_id提取")):
        actual_user_id = _get_actual_param(actual_params, "input_user_id")
        if not actual_user_id:
            user_id_ok = False
            failures.append("未在调用参数中识别到 user_id")
        elif actual_user_id != expected_user_id:
            user_id_ok = False
            failures.append(
                f"user_id 不匹配: 预期={expected_user_id}, 实际={actual_user_id}"
            )

    # 4) Behavior checks
    behavior_ok = True
    answer = _normalize_text(final_answer)

    if isinstance(include_words, list) and include_words:
        if not any(_normalize_text(w) and _normalize_text(w) in answer for w in include_words):
            behavior_ok = False
            failures.append(f"回复未命中期望关键词: {include_words}")

    if isinstance(exclude_words, list) and exclude_words:
        hit = [w for w in exclude_words if _normalize_text(w) and _normalize_text(w) in answer]
        if hit:
            behavior_ok = False
            failures.append(f"回复命中禁用关键词: {hit}")

    overall_pass = intent_ok and params_ok and user_id_ok and behavior_ok

    return {
        "意图正确": intent_ok,
        "参数正确": params_ok,
        "user_id正确": user_id_ok,
        "行为正确": behavior_ok,
        "是否通过": overall_pass,
        "失败原因": "；".join(failures),
        "预期行为": expected_behavior,
        "实际行为": _derive_actual_behavior(actual_skill),
    }


def _calc_summary(results: List[Dict[str, Any]]) -> Dict[str, Any]:
    total = len(results)
    passed = sum(1 for r in results if r.get("是否通过"))
    failed = total - passed

    def _rate(num: int, den: int) -> Optional[float]:
        if den <= 0:
            return None
        return round((num / den) * 100, 2)

    # Intent
    intent_pool = [r for r in results if r.get("指标", {}).get("纳入意图识别")]
    intent_ok = sum(1 for r in intent_pool if r.get("意图正确"))

    # Params
    params_pool = [r for r in results if r.get("指标", {}).get("纳入参数提取")]
    params_ok = sum(1 for r in params_pool if r.get("参数正确"))

    # user_id
    uid_pool = [r for r in results if r.get("指标", {}).get("纳入user_id提取")]
    uid_ok = sum(1 for r in uid_pool if r.get("user_id正确"))

    # Follow-up
    follow_pool = [r for r in results if r.get("指标", {}).get("纳入缺失追问")]
    follow_ok = sum(1 for r in follow_pool if r.get("行为正确"))

    # Guardrail/fallback
    guard_pool = [r for r in results if r.get("指标", {}).get("纳入拒绝兜底")]
    guard_ok = sum(1 for r in guard_pool if r.get("行为正确"))

    # Mis-call rate
    mis_pool = [r for r in results if r.get("指标", {}).get("纳入误调用率")]
    mis_count = sum(
        1
        for r in mis_pool
        if r.get("预期结果", {}).get("预期行为") != "调用工作流"
        and bool(r.get("实际技能"))
    )

    # Latency
    first_values: List[int] = []
    total_values: List[int] = []
    for r in results:
        first_v = r.get("首字响应ms")
        total_v = r.get("总耗时ms")
        if isinstance(first_v, int):
            first_values.append(first_v)
        if isinstance(total_v, int):
            total_values.append(total_v)

    single_pool = [r for r in results if r.get("对话类型") == "单轮"]
    multi_pool = [r for r in results if r.get("对话类型") == "多轮"]

    return {
        "总用例": total,
        "通过": passed,
        "失败": failed,
        "通过率": _rate(passed, total),
        "意图识别准确率": _rate(intent_ok, len(intent_pool)),
        "参数提取准确率": _rate(params_ok, len(params_pool)),
        "user_id提取准确率": _rate(uid_ok, len(uid_pool)),
        "缺失追问能力": _rate(follow_ok, len(follow_pool)),
        "拒绝兜底能力": _rate(guard_ok, len(guard_pool)),
        "误调用率": _rate(mis_count, len(mis_pool)),
        "首字响应ms": {
            "均值": round(statistics.mean(first_values), 2) if first_values else None,
            "P50": _percentile(first_values, 50),
            "P90": _percentile(first_values, 90),
        },
        "总耗时ms": {
            "均值": round(statistics.mean(total_values), 2) if total_values else None,
            "P50": _percentile(total_values, 50),
            "P90": _percentile(total_values, 90),
        },
        "单轮通过率": _rate(sum(1 for r in single_pool if r.get("是否通过")), len(single_pool)),
        "多轮通过率": _rate(sum(1 for r in multi_pool if r.get("是否通过")), len(multi_pool)),
    }


def _render_md(report: Dict[str, Any]) -> str:
    run = report.get("运行信息") or {}
    summary = report.get("汇总") or {}
    results = report.get("结果") or []

    lines: List[str] = []
    lines.append("# 调度模型测试报告（v2）")
    lines.append("")
    lines.append(f"- Run ID: `{run.get('run_id')}`")
    lines.append(f"- 时间: `{run.get('timestamp')}`")
    lines.append(f"- Bot ID: `{run.get('bot_id')}`")
    lines.append(f"- 流式首响: `{run.get('use_stream')}`")
    lines.append("")
    lines.append("## 汇总")
    lines.append("")
    for k in [
        "总用例",
        "通过",
        "失败",
        "通过率",
        "意图识别准确率",
        "参数提取准确率",
        "user_id提取准确率",
        "缺失追问能力",
        "拒绝兜底能力",
        "误调用率",
        "单轮通过率",
        "多轮通过率",
    ]:
        lines.append(f"- {k}: {summary.get(k)}")

    first = summary.get("首字响应ms") or {}
    total = summary.get("总耗时ms") or {}
    lines.append(f"- 首字响应ms: 均值={first.get('均值')} P50={first.get('P50')} P90={first.get('P90')}")
    lines.append(f"- 总耗时ms: 均值={total.get('均值')} P50={total.get('P50')} P90={total.get('P90')}")
    lines.append("")

    failed = [r for r in results if not r.get("是否通过")]
    if failed:
        lines.append("## 失败用例")
        lines.append("")
        for r in failed:
            lines.append(f"### {r.get('用例ID')} {r.get('测试场景')}")
            lines.append(f"- 对话类型: {r.get('对话类型')} 会话组: {r.get('会话组ID') or '-'} 轮次: {r.get('轮次')}")
            lines.append(f"- 用户输入: {r.get('用户输入')}")
            lines.append(f"- 预期技能/动作: {r.get('预期结果', {}).get('预期技能')} / {r.get('预期结果', {}).get('预期动作')}")
            lines.append(f"- 实际技能: {r.get('实际技能')}")
            lines.append(f"- 实际参数: `{_to_compact_json(r.get('实际参数') or {})}`")
            lines.append(f"- 失败原因: {r.get('失败原因')}")
            lines.append("")

    return "\n".join(lines)


def _write_table(ws, headers: List[str], rows: List[List[Any]]) -> None:
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(vertical="top", horizontal="center", wrap_text=True)
    for r in rows:
        ws.append(r)


def _apply_zebra_rows(ws) -> None:
    fill_even = PatternFill("solid", fgColor="F8FAFC")
    fill_odd = PatternFill("solid", fgColor="FFFFFF")
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = fill_even if idx % 2 == 0 else fill_odd
        for cell in row:
            cell.fill = fill


def _apply_pass_fail_fills(ws, *, pass_col: int) -> None:
    pass_fill = PatternFill("solid", fgColor="DCFCE7")
    fail_fill = PatternFill("solid", fgColor="FEE2E2")
    for row in ws.iter_rows(min_row=2):
        cell = row[pass_col - 1]
        val = str(cell.value or "")
        if val in ("True", "TRUE", "true", "通过"):
            cell.fill = pass_fill
        elif val in ("False", "FALSE", "false", "失败"):
            cell.fill = fail_fill


def _make_rules_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("评判规则")
    headers = ["规则类别", "规则名称", "判定逻辑", "备注"]
    rows: List[List[Any]] = [
        [
            "总通过判定",
            "单用例通过",
            "是否通过 = 意图正确 AND 参数正确 AND user_id正确 AND 行为正确",
            "四项任一不满足即失败",
        ],
        [
            "意图识别",
            "调用类意图",
            "预期行为=调用工作流 时，要求 实际技能非空，且与预期技能一致",
            "技能名会做归一化映射",
        ],
        [
            "意图识别",
            "非调用类意图",
            "预期行为!=调用工作流 时，要求 实际技能为空",
            "用于追问/兜底/拒绝类",
        ],
        [
            "参数提取",
            "参数匹配",
            "逐个校验预期参数；支持别名映射(keyword/new_keyword/input_keyword, input_user_id/user_id等)",
            "action 会归一化；keyword 在 send_mail(_cache)_using 下做严格匹配（ai 大小写敏感）",
        ],
        [
            "参数提取",
            "action匹配",
            "预期动作存在时，额外校验 action 字段",
            "例如 查询/修改/取消 会归一化到 查询订阅/修改订阅/取消订阅",
        ],
        [
            "user_id",
            "user_id匹配",
            "当纳入user_id提取=true 时，要求实际参数中的 input_user_id/user_id 与注入值一致",
            "默认注入值见运行信息.default_user_id",
        ],
        [
            "行为校验",
            "应包含",
            "回复应包含列表：命中任一即通过",
            "用于追问/兜底文案检查",
        ],
        [
            "行为校验",
            "不应包含",
            "回复不应包含列表：命中任一即失败",
            "例如 [STUB]/异常堆栈等",
        ],
        [
            "误调用率",
            "误调用定义",
            "纳入误调用率=true 且 预期行为!=调用工作流，但实际技能非空 => 记1次误调用",
            "用于衡量不该调度时的误触发",
        ],
        [
            "首字响应",
            "首字响应ms",
            "流式模式下记录 assistant 首个文本delta 到达时间；若未抓到则回退总耗时",
            "汇总给出均值/P50/P90",
        ],
        [
            "回复提取",
            "最终回复优先级",
            "优先取 type=answer 的文本；忽略 generate_answer_finish 等收尾信号",
            "避免把完成事件误判为用户可见回复",
        ],
        [
            "汇总指标",
            "口径说明",
            "各准确率仅在对应‘纳入*’标记为true的用例池内计算",
            "避免不同类型case互相污染分母",
        ],
    ]

    _write_table(ws, headers, rows)
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    widths = {
        "A": 14,
        "B": 16,
        "C": 78,
        "D": 36,
    }
    for c, w in widths.items():
        ws.column_dimensions[c].width = w
    ws.auto_filter.ref = ws.dimensions
    _apply_zebra_rows(ws)


def _classify_failure(case: Dict[str, Any]) -> Dict[str, str]:
    reason = _normalize_text(case.get("失败原因"))
    expected = case.get("预期结果") or {}
    expected_behavior = _normalize_text(expected.get("预期行为"))
    actual_skill = _normalize_text(case.get("实际技能"))
    actual_reply = _normalize_text(case.get("实际回复"))

    if "异常:" in reason:
        return {
            "失败类型": "执行异常",
            "诊断": "评测执行过程中出现API/网络/解析异常。",
            "建议": "检查 API Token、网络稳定性、超时设置与消息结构解析。",
            "是否疑似评测器问题": "否",
        }

    if "预期应调用工作流" in reason or "技能不匹配" in reason:
        return {
            "失败类型": "意图路由错误",
            "诊断": "模型未路由到预期技能，或路由到了错误技能。",
            "建议": "强化意图识别提示词与边界示例，加入易混淆意图反例。",
            "是否疑似评测器问题": "否",
        }

    if "预期不调用工作流" in reason:
        return {
            "失败类型": "误调用工作流",
            "诊断": "本应追问/兜底/拒绝，却触发了工作流调用。",
            "建议": "在系统提示词中提高“非任务场景禁止调用技能”的优先级。",
            "是否疑似评测器问题": "否",
        }

    if "参数不匹配" in reason or "动作不匹配" in reason:
        return {
            "失败类型": "参数提取错误",
            "诊断": "技能调用正确，但关键参数提取或归一化有偏差。",
            "建议": "补充槽位提取示例；统一 keyword/action 同义词映射。",
            "是否疑似评测器问题": "否",
        }

    if "user_id" in reason:
        return {
            "失败类型": "user_id提取错误",
            "诊断": "调用链路未携带或错误携带注入的 user_id。",
            "建议": "检查 System Context 提取规则与参数映射(input_user_id/user_id)。",
            "是否疑似评测器问题": "否",
        }

    if "回复未命中期望关键词" in reason:
        # 该场景在兜底/追问中最常见：如果拿到的是 finish 信号而非自然语言，则属于评测口径问题。
        if "generate_answer_finish" in actual_reply:
            return {
                "失败类型": "回复提取问题",
                "诊断": "记录到的是 Coze 完成信号，而非用户可见的自然语言 answer。",
                "建议": "优先提取 type=answer 的文本；忽略 verbose/finish 消息。",
                "是否疑似评测器问题": "是",
            }
        if expected_behavior in ("追问", "直接回复", "拒绝") and not actual_skill:
            return {
                "失败类型": "兜底文案不达标",
                "诊断": "未触发工作流，但回复内容未满足兜底/追问关键词要求。",
                "建议": "放宽关键词口径为“命中任一+禁词”，并加强兜底模板示例。",
                "是否疑似评测器问题": "否",
            }
        return {
            "失败类型": "回复内容不达标",
            "诊断": "回复存在但未命中预期关键信息。",
            "建议": "优化提示词中的回复模板与关键术语覆盖。",
            "是否疑似评测器问题": "否",
        }

    return {
        "失败类型": "未分类",
        "诊断": "未命中既定分类规则，需人工复核。",
        "建议": "查看测试明细中的原始失败原因与上下文。",
        "是否疑似评测器问题": "否",
    }


def _make_failures_sheet(wb: Workbook, report: Dict[str, Any]) -> None:
    ws = wb.create_sheet("失败案例分析")
    failed_cases = [r for r in (report.get("结果") or []) if not r.get("是否通过")]

    headers = [
        "用例ID",
        "意图分类",
        "二级意图",
        "测试场景",
        "预期行为",
        "实际行为",
        "实际技能",
        "失败类型",
        "失败原因(原始)",
        "诊断",
        "建议",
        "是否疑似评测器问题",
    ]

    rows: List[List[Any]] = []
    for case in failed_cases:
        cls = _classify_failure(case)
        expected = case.get("预期结果") or {}
        rows.append(
            [
                case.get("用例ID"),
                case.get("意图分类"),
                case.get("二级意图"),
                case.get("测试场景"),
                expected.get("预期行为"),
                case.get("实际行为"),
                case.get("实际技能"),
                cls.get("失败类型"),
                case.get("失败原因"),
                cls.get("诊断"),
                cls.get("建议"),
                cls.get("是否疑似评测器问题"),
            ]
        )

    _write_table(ws, headers, rows)
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    widths = {
        "A": 10,
        "B": 12,
        "C": 14,
        "D": 24,
        "E": 10,
        "F": 10,
        "G": 18,
        "H": 16,
        "I": 40,
        "J": 40,
        "K": 42,
        "L": 16,
    }
    for c, w in widths.items():
        ws.column_dimensions[c].width = w
    ws.auto_filter.ref = ws.dimensions
    _apply_zebra_rows(ws)


def _make_excel(report: Dict[str, Any], xlsx_path: Path) -> None:
    wb = Workbook()
    ws_summary = wb.active
    if ws_summary is None:
        ws_summary = wb.create_sheet("测试汇总")
    ws_summary.title = "测试汇总"

    run = report.get("运行信息") or {}
    summary = report.get("汇总") or {}

    summary_rows = [
        ["执行时间", run.get("timestamp")],
        ["Run ID", run.get("run_id")],
        ["Bot ID", run.get("bot_id")],
        ["流式首响", run.get("use_stream")],
        ["总用例", summary.get("总用例")],
        ["通过", summary.get("通过")],
        ["失败", summary.get("失败")],
        ["通过率", summary.get("通过率")],
        ["意图识别准确率", summary.get("意图识别准确率")],
        ["参数提取准确率", summary.get("参数提取准确率")],
        ["user_id提取准确率", summary.get("user_id提取准确率")],
        ["缺失追问能力", summary.get("缺失追问能力")],
        ["拒绝兜底能力", summary.get("拒绝兜底能力")],
        ["误调用率", summary.get("误调用率")],
        ["首字响应ms", _to_compact_json(summary.get("首字响应ms"))],
        ["总耗时ms", _to_compact_json(summary.get("总耗时ms"))],
        ["单轮通过率", summary.get("单轮通过率")],
        ["多轮通过率", summary.get("多轮通过率")],
    ]
    _write_table(ws_summary, ["指标", "值"], summary_rows)
    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 80
    ws_summary.freeze_panes = "A2"

    ws_detail = wb.create_sheet("测试明细")
    headers = [
        "用例ID",
        "意图分类",
        "二级意图",
        "测试场景",
        "对话类型",
        "会话组ID",
        "轮次",
        "用户输入",
        "预期行为",
        "预期技能",
        "预期动作",
        "预期参数",
        "实际行为",
        "实际技能",
        "实际参数",
        "首字响应ms",
        "总耗时ms",
        "意图正确",
        "参数正确",
        "user_id正确",
        "行为正确",
        "是否通过",
        "失败原因",
        "备注",
    ]

    rows: List[List[Any]] = []
    for r in report.get("结果", []) or []:
        exp = r.get("预期结果") or {}
        rows.append(
            [
                r.get("用例ID"),
                r.get("意图分类"),
                r.get("二级意图"),
                r.get("测试场景"),
                r.get("对话类型"),
                r.get("会话组ID"),
                r.get("轮次"),
                r.get("用户输入"),
                exp.get("预期行为"),
                exp.get("预期技能"),
                exp.get("预期动作"),
                _to_compact_json(exp.get("预期参数") or {}),
                r.get("实际行为"),
                r.get("实际技能"),
                _to_compact_json(r.get("实际参数") or {}),
                r.get("首字响应ms"),
                r.get("总耗时ms"),
                r.get("意图正确"),
                r.get("参数正确"),
                r.get("user_id正确"),
                r.get("行为正确"),
                r.get("是否通过"),
                r.get("失败原因"),
                r.get("备注"),
            ]
        )

    _write_table(ws_detail, headers, rows)
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws_detail.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    widths = {
        "A": 10,
        "B": 12,
        "C": 14,
        "D": 20,
        "E": 10,
        "F": 10,
        "G": 8,
        "H": 36,
        "I": 10,
        "J": 18,
        "K": 14,
        "L": 30,
        "M": 10,
        "N": 18,
        "O": 36,
        "P": 10,
        "Q": 10,
        "R": 10,
        "S": 10,
        "T": 12,
        "U": 10,
        "V": 10,
        "W": 38,
        "X": 20,
    }
    for c, w in widths.items():
        ws_detail.column_dimensions[c].width = w

    ws_detail.auto_filter.ref = ws_detail.dimensions
    ws_detail.freeze_panes = "A2"
    _apply_zebra_rows(ws_detail)
    _apply_pass_fail_fills(ws_detail, pass_col=22)

    _make_rules_sheet(wb)
    _make_failures_sheet(wb, report)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def main() -> int:
    try:
        stdout_reconfigure = getattr(sys.stdout, "reconfigure", None)
        if callable(stdout_reconfigure):
            stdout_reconfigure(encoding="utf-8", errors="replace")
        stderr_reconfigure = getattr(sys.stderr, "reconfigure", None)
        if callable(stderr_reconfigure):
            stderr_reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    if not SUITE_PATH.exists():
        raise SystemExit(f"Suite file not found: {SUITE_PATH}")

    suite = json.loads(SUITE_PATH.read_text(encoding="utf-8"))
    meta = suite.get("测试集") or {}
    cases = suite.get("用例")
    if not isinstance(cases, list) or not cases:
        raise SystemExit("dispatch_suite_v2.json 中的 用例 必须为非空数组")

    bot_id = (
        (os.getenv("COZE_DISPATCH_TEST_BOT_ID") or "").strip()
        or _normalize_text(meta.get("bot_id"))
        or (os.getenv("COZE_BOT_ID") or "").strip()
    )
    if not bot_id:
        raise SystemExit("未找到测试 bot_id，请设置 COZE_DISPATCH_TEST_BOT_ID 或在 suite 中提供 bot_id")

    default_cfg = suite.get("默认配置") or {}
    default_user_id = (
        (os.getenv("DISPATCH_CONTEXT_USER_ID") or "").strip()
        or _normalize_text(default_cfg.get("测试user_id"))
        or "dispatch_test_user_001"
    )
    use_stream = _bool_env("DISPATCH_USE_STREAM", bool(default_cfg.get("是否使用流式首响", True)))
    timeout_s = int(os.getenv("DISPATCH_TIMEOUT_S") or "120")
    max_polls = int(os.getenv("DISPATCH_MAX_POLLS") or "20")
    poll_interval_s = float(os.getenv("DISPATCH_POLL_INTERVAL_S") or "0.8")
    start_index = int(os.getenv("DISPATCH_START_INDEX") or "1")
    max_cases = int(os.getenv("DISPATCH_MAX_CASES") or "0")
    stop_on_exception = _bool_env("DISPATCH_STOP_ON_EXCEPTION", False)

    if start_index < 1:
        start_index = 1

    # 1-based start index for safer resume runs.
    cases = cases[start_index - 1 :]

    if max_cases > 0:
        cases = cases[:max_cases]

    client = CozeChatClient.from_env()
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_json_path = REPORT_DIR / f"dispatch_report_{run_id}.json"
    report_md_path = REPORT_DIR / f"dispatch_report_{run_id}.md"
    report_xlsx_path = REPORT_DIR / f"dispatch_test_data_{run_id}.xlsx"
    report_json_latest_path = REPORT_DIR / "dispatch_report.json"
    report_md_latest_path = REPORT_DIR / "dispatch_report.md"
    report_xlsx_latest_path = REPORT_DIR / "dispatch_test_data.xlsx"

    report: Dict[str, Any] = {
        "测试集": meta,
        "运行信息": {
            "run_id": run_id,
            "timestamp": _now_ts(),
            "base_url": os.getenv("COZE_BASE_URL") or os.getenv("COZE_API_BASE") or "https://api.coze.cn",
            "bot_id": bot_id,
            "use_stream": use_stream,
            "timeout_s": timeout_s,
            "max_polls": max_polls,
            "poll_interval_s": poll_interval_s,
            "start_index": start_index,
            "max_cases": max_cases,
            "stop_on_exception": stop_on_exception,
            "default_user_id": default_user_id,
        },
        "结果": [],
        "汇总": {},
    }

    _safe_print(f"Running dispatch suite ({len(cases)} cases) at {_now_ts()}")
    _safe_print(f"Bot ID: {bot_id} | stream={use_stream} | timeout_s={timeout_s}")
    _safe_print(
        f"Range: start_index={start_index}, max_cases={max_cases}, stop_on_exception={stop_on_exception}"
    )

    conv_map: Dict[str, Dict[str, Any]] = {}
    aborted = False
    abort_reason = ""

    for idx, case in enumerate(cases, start=1):
        case = dict(case)
        case_id = _normalize_text(case.get("用例ID")) or f"CASE-{idx:03d}"
        dialog_type = _normalize_text(case.get("对话类型")) or "单轮"
        group_id = _normalize_text(case.get("会话组ID"))
        user_input = _normalize_text(case.get("用户输入"))

        if not user_input:
            case_result = dict(case)
            case_result.update(
                {
                    "实际技能": "",
                    "实际参数": {},
                    "实际回复": "",
                    "首字响应ms": None,
                    "总耗时ms": 0,
                    "会话ID": "",
                    "chat_id": "",
                    "意图正确": False,
                    "参数正确": False,
                    "user_id正确": False,
                    "行为正确": False,
                    "实际行为": "未调用工作流",
                    "是否通过": False,
                    "失败原因": "用户输入为空",
                }
            )
            report["结果"].append(case_result)
            _safe_print(f"FAIL {case_id} 输入为空")
            continue

        conversation_id: Optional[str] = None
        user_id = default_user_id
        if dialog_type == "多轮":
            if not group_id:
                case_result = dict(case)
                case_result.update(
                    {
                        "实际技能": "",
                        "实际参数": {},
                        "实际回复": "",
                        "首字响应ms": None,
                        "总耗时ms": 0,
                        "会话ID": "",
                        "chat_id": "",
                        "意图正确": False,
                        "参数正确": False,
                        "user_id正确": False,
                        "行为正确": False,
                        "实际行为": "未调用工作流",
                        "是否通过": False,
                        "失败原因": "多轮用例缺少会话组ID",
                    }
                )
                report["结果"].append(case_result)
                _safe_print(f"FAIL {case_id} 多轮缺少会话组ID")
                continue
            if group_id in conv_map:
                conversation_id = conv_map[group_id]["conversation_id"]
                user_id = conv_map[group_id]["user_id"]
            else:
                # First turn of a multi-turn group: pre-create conversation
                # so all subsequent turns share the same conversation context.
                try:
                    created_conversation_id = client.create_conversation(bot_id=bot_id)
                    conversation_id = created_conversation_id
                    conv_map[group_id] = {
                        "conversation_id": created_conversation_id,
                        "user_id": user_id,
                    }
                    _safe_print(f"  [multi-turn] Created conversation {conversation_id} for group {group_id}")
                except Exception as e:
                    case_result = dict(case)
                    case_result.update(
                        {
                            "实际技能": "",
                            "实际参数": {},
                            "实际回复": "",
                            "首字响应ms": None,
                            "总耗时ms": 0,
                            "会话ID": "",
                            "chat_id": "",
                            "意图正确": False,
                            "参数正确": False,
                            "user_id正确": False,
                            "行为正确": False,
                            "实际行为": "未调用工作流",
                            "是否通过": False,
                            "失败原因": f"异常: 创建会话失败: {type(e).__name__}: {e}",
                        }
                    )
                    report["结果"].append(case_result)
                    _safe_print(f"FAIL {case_id} 创建会话失败: {e}")
                    if stop_on_exception:
                        aborted = True
                        abort_reason = f"{case_id}: create_conversation failed: {e}"
                        break
                    continue

        try:
            run = client.chat_once(
                bot_id=bot_id,
                message=user_input,
                user_id=user_id,
                conversation_id=conversation_id,
                use_stream=use_stream,
                max_polls=max_polls,
                poll_interval_s=poll_interval_s,
                timeout_s=timeout_s,
            )
            if run.status and run.status != "completed":
                raise RuntimeError(f"chat status is {run.status}")
            actual_skill, actual_params = pick_primary_call(
                function_calls=run.function_calls,
                stub_calls=run.stub_calls,
            )
            eval_result = _evaluate_case(
                case,
                actual_skill=actual_skill,
                actual_params=actual_params,
                final_answer=run.final_answer,
                expected_user_id=user_id,
            )
            case_result = dict(case)
            case_result.update(
                {
                    "实际技能": actual_skill,
                    "实际参数": actual_params,
                    "实际回复": run.final_answer,
                    "首字响应ms": run.first_response_ms,
                    "总耗时ms": run.total_ms,
                    "会话ID": run.conversation_id,
                    "chat_id": run.chat_id,
                    "function_calls": run.function_calls,
                    "stub_calls": run.stub_calls,
                }
            )
            case_result.update(eval_result)
        except Exception as e:
            case_result = dict(case)
            case_result.update(
                {
                    "实际技能": "",
                    "实际参数": {},
                    "实际回复": "",
                    "首字响应ms": None,
                    "总耗时ms": 0,
                    "会话ID": conversation_id or "",
                    "chat_id": "",
                    "意图正确": False,
                    "参数正确": False,
                    "user_id正确": False,
                    "行为正确": False,
                    "实际行为": "未调用工作流",
                    "是否通过": False,
                    "失败原因": f"异常: {type(e).__name__}: {e}",
                }
            )
            if stop_on_exception:
                report["结果"].append(case_result)
                aborted = True
                abort_reason = f"{case_id}: {type(e).__name__}: {e}"
                _safe_print(f"ABORT {case_id} due to exception: {type(e).__name__}: {e}")
                break

        report["结果"].append(case_result)
        status = "PASS" if case_result.get("是否通过") else "FAIL"
        _safe_print(
            f"{status} {case_id} {case_result.get('测试场景')} "
            f"(first={case_result.get('首字响应ms')}ms total={case_result.get('总耗时ms')}ms)"
        )

    if aborted:
        report["运行信息"]["是否中止"] = True
        report["运行信息"]["中止原因"] = abort_reason
    else:
        report["运行信息"]["是否中止"] = False

    report["汇总"] = _calc_summary(report["结果"])

    json_content = json.dumps(report, ensure_ascii=False, indent=2)
    report_json_path.write_text(json_content, encoding="utf-8")
    report_json_latest_path.write_text(json_content, encoding="utf-8")

    md_content = _render_md(report)
    report_md_path.write_text(md_content, encoding="utf-8")
    report_md_latest_path.write_text(md_content, encoding="utf-8")

    _make_excel(report, report_xlsx_path)
    _make_excel(report, report_xlsx_latest_path)

    _safe_print(f"Report JSON: {report_json_path}")
    _safe_print(f"Report MD: {report_md_path}")
    _safe_print(f"Report XLSX: {report_xlsx_path}")
    _safe_print(f"Latest JSON: {report_json_latest_path}")
    _safe_print(f"Latest MD: {report_md_latest_path}")
    _safe_print(f"Latest XLSX: {report_xlsx_latest_path}")

    if aborted:
        return 2
    return 0 if int(report["汇总"].get("失败") or 0) == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
