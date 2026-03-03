"""Run send_mail_using cache/rate-limit workflow suite and write reports.

Usage:
  python tests/run_send_mail_cache_suite.py

Outputs (tests/reports/send_mail_cache/):
  - send_mail_cache_report_<run_id>.json
  - send_mail_cache_report_<run_id>.md
  - send_mail_cache_test_data_<run_id>.xlsx
  Plus stable "latest" files:
  - send_mail_cache_report.json
  - send_mail_cache_report.md
  - send_mail_cache_test_data.xlsx
"""

from __future__ import annotations

import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from coze_workflow_client import CozeWorkflowClient


SUITE_PATH = Path(__file__).resolve().parent / "send_mail_cache_suite.json"
REPORT_DIR = Path(__file__).resolve().parent / "reports" / "send_mail_cache"


def _safe_print(s: str) -> None:
    try:
        print(s)
    except UnicodeEncodeError:
        print(s.encode("utf-8", errors="replace").decode("utf-8", errors="replace"))


def _now_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _rand_token(n: int = 8) -> str:
    import random
    import string

    alphabet = string.ascii_lowercase + string.digits
    return "".join(random.choice(alphabet) for _ in range(n))


def _make_context() -> Dict[str, str]:
    tag = datetime.now().strftime("%Y%m%d_%H%M%S")
    real_email = (os.getenv("SEND_MAIL_CACHE_TEST_EMAIL") or "demo@example.com").strip()
    subscribed_user_id = (os.getenv("SEND_MAIL_CACHE_SUBSCRIBED_USER_ID") or "111").strip()
    return {
        "$REAL_EMAIL": real_email,
        "$SUBSCRIBED_USER_ID": subscribed_user_id,
        "$SM01_USER_ID": f"autotest_cache_sm01_{tag}_{_rand_token()}",
        "$UNSUBSCRIBED_USER_ID": f"autotest_cache_u1_{tag}_{_rand_token()}",
        "$UNSUBSCRIBED_USER_ID_2": f"autotest_cache_u2_{tag}_{_rand_token()}",
        "$SM05_USER_ID": f"autotest_cache_sm05_{tag}_{_rand_token()}",
        "$SM06_USER_ID": f"autotest_cache_sm06_{tag}_{_rand_token()}",
        "$SM08_USER_ID": f"autotest_cache_sm08_{tag}_{_rand_token()}",
        "$NO_CACHE_KEYWORD": f"nocache_{tag}_{_rand_token(6)}",
    }


def _render_template(value: str, ctx: Dict[str, str]) -> str:
    if not isinstance(value, str):
        return value
    out = value
    # Replace longer placeholders first to avoid prefix collisions
    # e.g. "$CACHE01_EMAIL" vs "$CACHE01_EMAIL_UPPER".
    for k in sorted(ctx.keys(), key=len, reverse=True):
        out = out.replace(k, ctx[k])
    return out


def _safe_json_load(raw_text: object) -> Optional[Any]:
    if not isinstance(raw_text, str):
        return None
    raw = raw_text.strip()
    if not raw:
        return None
    cleaned = re.sub(r"^```json\s*", "", raw, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```\s*$", "", cleaned)
    try:
        return json.loads(cleaned)
    except Exception:
        # Try a best-effort substring parse for embedded JSON
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start >= 0 and end > start:
            try:
                return json.loads(cleaned[start : end + 1])
            except Exception:
                return None
        return None


def _infer_allowed(msg: str, hinted_allowed: Optional[bool]) -> Optional[bool]:
    if hinted_allowed is not None:
        return hinted_allowed

    text = (msg or "").strip()
    if not text:
        return None

    deny_patterns = (
        "未找到您的邮箱",
        "请提供邮箱",
        "发送次数已达上限",
        "次数已达上限",
        "rate limit",
    )
    if any(p in text for p in deny_patterns):
        return False

    allow_patterns = (
        "发送成功",
        "邮件正在发送中",
        "已发送",
        "请稍等一分钟",
    )
    if any(p in text for p in allow_patterns):
        return True

    return None


def _extract_payload(raw_output: str) -> Tuple[Optional[Dict[str, Any]], List[str]]:
    notes: List[str] = []
    obj = _safe_json_load(raw_output)
    if obj is None:
        msg = str(raw_output or "").strip()
        if not msg:
            return None, ["empty_output"]
        return {
            "msg": msg,
            "allowed": _infer_allowed(msg, None),
            "email": "",
            "keyword": "",
        }, ["plain_text_output"]

    def unwrap(o: Any) -> Optional[Dict[str, Any]]:
        if isinstance(o, dict):
            # Prefer canonical payload wrappers first, because some branches may expose
            # an intermediate progress message while Output/data carries the final result.
            for k in ("Output", "output", "data", "result", "content"):
                v = o.get(k)
                if isinstance(v, dict):
                    notes.append(f"wrapped_{k}_dict")
                    inner = unwrap(v)
                    if inner is not None:
                        return inner
                if isinstance(v, str):
                    notes.append(f"wrapped_{k}_str")
                    inner_obj = _safe_json_load(v)
                    inner = unwrap(inner_obj) if inner_obj is not None else None
                    if inner is not None:
                        return inner

            # Fallback to branch keys.
            for key, allowed in (
                ("输出_3", True),
                ("输出_2", True),
                ("输出_1", False),
                ("output_3", True),
                ("output_2", True),
                ("output_1", False),
            ):
                v = o.get(key)
                if isinstance(v, str) and v.strip():
                    notes.append(f"wrapped_{key}_msg")
                    return {"allowed": _infer_allowed(v, allowed), "msg": v}

        if isinstance(o, str) and o.strip():
            return {"allowed": _infer_allowed(o, None), "msg": o}

        if isinstance(o, dict) and any(k in o for k in ("allowed", "email", "keyword", "msg")):
            msg = str(o.get("msg") or o.get("message") or "").strip()
            allowed_val = o.get("allowed")
            hinted_allowed: Optional[bool] = None
            if isinstance(allowed_val, bool):
                hinted_allowed = allowed_val
            elif isinstance(allowed_val, str) and allowed_val.strip().lower() in ("true", "false"):
                hinted_allowed = allowed_val.strip().lower() == "true"

            o["allowed"] = _infer_allowed(msg, hinted_allowed)
            o["msg"] = msg
            return o

        if isinstance(o, dict):
            # Last fallback: take first non-empty string field as message.
            for v in o.values():
                if isinstance(v, str) and v.strip():
                    notes.append("fallback_first_string")
                    return {"allowed": _infer_allowed(v, None), "msg": v}

        return None

    extracted = unwrap(obj)
    if extracted is None:
        msg = str(raw_output or "").strip()
        if msg:
            notes.append("fallback_raw_text")
            extracted = {
                "allowed": _infer_allowed(msg, None),
                "msg": msg,
                "email": "",
                "keyword": "",
            }
        else:
            notes.append("missing_fields")
    return extracted, notes


def _extract_remaining(payload: Dict[str, Any], msg: str) -> Optional[int]:
    for key in ("remaining", "remaining_count", "remain"):
        v = payload.get(key)
        if isinstance(v, int):
            return v
        if isinstance(v, str) and v.isdigit():
            return int(v)

    if isinstance(msg, str):
        # Examples we want to match:
        # - "今日剩余次数: 99"
        # - "remaining: 12"
        m = re.search(r"(?:剩余|remain|remaining)[^0-9]{0,16}(\d+)", msg, re.IGNORECASE)
        if m:
            return int(m.group(1))
    return None


def _inputs_to_str(inputs: Dict[str, Any]) -> str:
    return "input_email={}; input_keyword={}; input_user_id={}".format(
        inputs.get("input_email", ""),
        inputs.get("input_keyword", ""),
        inputs.get("input_user_id", ""),
    )


def _expected_to_str(expected: Dict[str, Any]) -> str:
    parts: List[str] = []
    if "allowed" in expected:
        parts.append(f"allowed={expected.get('allowed')}")
    if isinstance(expected.get("any_contains"), list) and expected.get("any_contains"):
        parts.append("命中任一文案")
    if expected.get("remaining_present"):
        parts.append("remaining存在")
    if expected.get("remaining_less_than_case"):
        parts.append(f"remaining < {expected.get('remaining_less_than_case')}")
    return "；".join(parts) if parts else ""


def _assert_case(
    expected: Dict[str, Any],
    payload: Optional[Dict[str, Any]],
    case_state: Dict[str, Any],
) -> Tuple[bool, List[str], Dict[str, Any]]:
    errors: List[str] = []
    if payload is None:
        return False, ["missing payload"], {}

    allowed_val = payload.get("allowed")
    if isinstance(allowed_val, bool):
        allowed = allowed_val
    elif isinstance(allowed_val, str) and allowed_val.strip().lower() in ("true", "false"):
        allowed = allowed_val.strip().lower() == "true"
    else:
        allowed = _infer_allowed(str(payload.get("msg") or payload.get("message") or ""), None)

    msg = payload.get("msg") or payload.get("message") or ""
    email = payload.get("email") or ""
    keyword = payload.get("keyword") or ""
    remaining = _extract_remaining(payload, str(msg))

    for forbidden in ("Exception", "Traceback"):
        if forbidden in str(msg):
            errors.append(f"msg contains forbidden text: {forbidden}")

    if "allowed" in expected:
        if allowed is None:
            errors.append("missing allowed in output")
        elif allowed != expected.get("allowed"):
            errors.append(f"allowed mismatch: got {allowed}")

    any_contains = expected.get("any_contains")
    if isinstance(any_contains, list) and any_contains:
        if not any((str(s) in str(msg)) for s in any_contains):
            errors.append(f"msg missing any expected substring: {any_contains}")

    none_contains = expected.get("none_contains")
    if isinstance(none_contains, list) and none_contains:
        present = [s for s in none_contains if str(s) in str(msg)]
        if present:
            errors.append(f"msg contains forbidden substrings: {present}")

    if expected.get("remaining_present") and remaining is None:
        errors.append("remaining missing")

    compare_case = expected.get("remaining_less_than_case")
    if isinstance(compare_case, str) and compare_case:
        base = case_state.get(compare_case)
        if remaining is None or base is None:
            errors.append(f"remaining compare failed ({compare_case} missing)")
        elif remaining >= base:
            errors.append(f"remaining not decreased vs {compare_case}: {remaining} >= {base}")

    extracted = {
        "allowed": allowed,
        "msg": msg,
        "email": email,
        "keyword": keyword,
        "remaining": remaining,
    }
    return len(errors) == 0, errors, extracted


def _call_workflow(
    client: CozeWorkflowClient,
    workflow_id: str,
    params: Dict[str, Any],
) -> Tuple[str, int, Optional[str]]:
    t0 = time.time()
    # Some workflows return a placeholder like "{{output}}" in the sync response.
    # Use async+poll to reliably fetch the final output from run history.
    res = client.run_workflow(
        workflow_id,
        params,
        is_async=True,
        timeout_s=180,
        poll_interval_s=1.0,
    )
    dt_ms = int((time.time() - t0) * 1000)
    return res.output or "", dt_ms, res.debug_url


def _apply_zebra_rows(ws) -> None:
    zebra_fill = PatternFill("solid", fgColor="F8FAFC")
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if i % 2 == 0:
            for cell in row:
                if cell.fill is None or cell.fill.fill_type is None:
                    cell.fill = zebra_fill


def _apply_table_borders(ws) -> None:
    border_color = "E2E8F0"
    thin = Side(style="thin", color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border


def _write_table(ws, headers: List[str], rows: List[List[Any]]) -> None:
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.append(headers)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
    ws.freeze_panes = "A2"
    for r in rows:
        ws.append(r)


def _apply_result_fills(ws, *, actual_col: int) -> None:
    pass_fill = PatternFill("solid", fgColor="DCFCE7")
    fail_fill = PatternFill("solid", fgColor="FEE2E2")
    for row in ws.iter_rows(min_row=2):
        cell = row[actual_col - 1]
        val = str(cell.value or "")
        if "✅" in val or "PASS" in val:
            cell.fill = pass_fill
        elif "❌" in val or "FAIL" in val:
            cell.fill = fail_fill


def _make_excel(report: Dict[str, Any], xlsx_path: Path) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "测试汇总"

    summary = report.get("summary") or {}
    run = report.get("run") or {}
    suite = report.get("suite") or {}
    results = report.get("results") or []

    summary_rows = [
        ["执行时间", run.get("timestamp") or ""],
        ["工作流ID", suite.get("workflow_id") or ""],
        ["Run ID", run.get("run_id") or ""],
        ["总用例数(实际执行)", summary.get("total") or 0],
        ["通过", summary.get("passed") or 0],
        ["失败", summary.get("failed") or 0],
        ["API Base URL", run.get("base_url") or ""],
        ["API 调用上限", run.get("max_calls") or ""],
        ["API 调用统计", json.dumps((run.get("call_stats") or {}), ensure_ascii=False)],
    ]
    _write_table(ws_summary, ["指标", "值"], summary_rows)
    label_fill = PatternFill("solid", fgColor="EEF2FF")
    for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row, min_col=1, max_col=1):
        row[0].fill = label_fill
        row[0].font = Font(bold=True, color="1E293B")

    ws_summary.column_dimensions["A"].width = 18
    ws_summary.column_dimensions["B"].width = 80
    for row in ws_summary.iter_rows(min_row=2):
        row[0].alignment = Alignment(horizontal="left", vertical="center")
        row[1].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws_detail = wb.create_sheet("测试明细")
    headers = [
        "意图",
        "测试场景",
        "用户输入",
        "预期结果",
        "实际结果",
        "评分维度",
        "评分指标",
        "失败原因",
        "修复状态",
        "备注",
    ]

    rows: List[List[Any]] = []
    for r in results:
        rows.append(
            [
                r.get("intent"),
                r.get("scenario"),
                r.get("inputs_str"),
                r.get("expected_str"),
                r.get("actual_str"),
                r.get("dimension"),
                r.get("criteria"),
                r.get("failure_reason"),
                r.get("fix_status"),
                r.get("note"),
            ]
        )

    _write_table(ws_detail, headers, rows)

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws_detail.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    for r in range(2, ws_detail.max_row + 1):
        ws_detail.row_dimensions[r].height = 40

    widths = {
        "A": 12,
        "B": 28,
        "C": 52,
        "D": 44,
        "E": 34,
        "F": 14,
        "G": 24,
        "H": 24,
        "I": 12,
        "J": 22,
    }
    for col, w in widths.items():
        ws_detail.column_dimensions[col].width = w

    ws_detail.auto_filter.ref = ws_detail.dimensions
    _apply_zebra_rows(ws_detail)
    _apply_result_fills(ws_detail, actual_col=5)
    _apply_table_borders(ws_summary)
    _apply_table_borders(ws_detail)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(xlsx_path)
    except PermissionError:
        fallback = xlsx_path.with_name(xlsx_path.stem + "_fallback.xlsx")
        wb.save(fallback)


def _render_md(report: Dict[str, Any]) -> str:
    summary = report.get("summary") or {}
    run = report.get("run") or {}
    suite = report.get("suite") or {}

    lines: List[str] = []
    lines.append("# send_mail_cache_using test report")
    lines.append("")
    lines.append(f"- Run ID: `{run.get('run_id')}`")
    lines.append(f"- Time: `{run.get('timestamp')}`")
    lines.append(f"- Workflow ID: `{suite.get('workflow_id')}`")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- Total: {summary.get('total')}")
    lines.append(f"- Passed: {summary.get('passed')}")
    lines.append(f"- Failed: {summary.get('failed')}")
    lines.append("")

    if summary.get("failed"):
        lines.append("## Failures")
        for r in report.get("results", []) or []:
            if r.get("passed"):
                continue
            lines.append("")
            lines.append(f"### {r.get('id')} {r.get('scenario')}")
            lines.append(f"- Errors: {r.get('errors')}")
            if r.get("debug_url"):
                lines.append(f"- Debug: {r.get('debug_url')}")

    return "\n".join(lines)


def main() -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    if not SUITE_PATH.exists():
        raise SystemExit(f"Suite file not found: {SUITE_PATH}")

    suite = json.loads(SUITE_PATH.read_text(encoding="utf-8"))
    suite_meta = suite.get("suite") or {}
    workflow_id = str(suite_meta.get("workflow_id") or "").strip()
    if not workflow_id:
        raise SystemExit("suite.workflow_id is required")

    cases = suite.get("cases")
    if not isinstance(cases, list) or not cases:
        raise SystemExit("suite.cases must be a non-empty list")

    default_max_calls = len(cases) * 20
    max_calls = int(os.getenv("COZE_MAX_CALLS") or str(default_max_calls))
    client = CozeWorkflowClient.from_env(max_calls=max_calls)

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_json_path = REPORT_DIR / f"send_mail_cache_report_{run_id}.json"
    report_md_path = REPORT_DIR / f"send_mail_cache_report_{run_id}.md"
    report_xlsx_path = REPORT_DIR / f"send_mail_cache_test_data_{run_id}.xlsx"
    report_json_latest_path = REPORT_DIR / "send_mail_cache_report.json"
    report_md_latest_path = REPORT_DIR / "send_mail_cache_report.md"
    report_xlsx_latest_path = REPORT_DIR / "send_mail_cache_test_data.xlsx"

    report: Dict[str, Any] = {
        "suite": suite_meta,
        "run": {
            "run_id": run_id,
            "timestamp": _now_ts(),
            "base_url": os.getenv("COZE_BASE_URL")
            or os.getenv("COZE_API_BASE")
            or "https://api.coze.cn",
            "max_calls": max_calls,
        },
        "results": [],
        "summary": {},
    }

    ctx = _make_context()
    case_state: Dict[str, Any] = {}

    passed = 0
    failed = 0

    _safe_print(f"Running send_mail_cache_using suite ({len(cases)} cases) at {_now_ts()}")
    _safe_print(f"[Quota] API call limit set to {max_calls}")

    for idx, case in enumerate(cases):
        case_id = str(case.get("id") or f"case_{idx}")
        intent = str(case.get("intent") or "")
        scenario = str(case.get("scenario") or "")
        dimension = str(case.get("dimension") or "")
        expected = case.get("expected") or {}
        inputs = case.get("inputs") or {}

        # Render placeholders
        rendered_inputs = {
            k: _render_template(str(v or ""), ctx) for k, v in inputs.items()
        }

        params = {
            "input_email": rendered_inputs.get("input_email", ""),
            "input_keyword": rendered_inputs.get("input_keyword", ""),
            "input_user_id": rendered_inputs.get("input_user_id", ""),
        }

        try:
            output, dt_ms, debug_url = _call_workflow(client, workflow_id, params)
            payload, notes = _extract_payload(output)
            passed_case, errors, extracted = _assert_case(expected, payload, case_state)
        except Exception as e:
            output = ""
            dt_ms = 0
            debug_url = None
            payload = None
            notes = []
            passed_case = False
            errors = [f"exception: {type(e).__name__}: {e}"]
            extracted = {}

        if passed_case and extracted.get("remaining") is not None:
            case_state[case_id] = extracted.get("remaining")

        status_cn = "✅通过" if passed_case else "❌失败"
        actual_str = "；".join(
            [
                status_cn,
                f"allowed={extracted.get('allowed')}",
                f"remaining={extracted.get('remaining')}",
                f"email={extracted.get('email')}",
                f"keyword={extracted.get('keyword')}",
                f"msg={extracted.get('msg')}",
                f"duration_ms={dt_ms}",
            ]
        )

        note_parts: List[str] = [f"id={case_id}"]
        if debug_url:
            note_parts.append(f"debug_url={debug_url}")
        note = "；".join(note_parts)

        result = {
            "id": case_id,
            "intent": intent,
            "scenario": scenario,
            "dimension": dimension,
            "inputs": params,
            "inputs_str": _inputs_to_str(params),
            "expected": expected,
            "expected_str": _expected_to_str(expected),
            "actual_str": actual_str,
            "criteria": case.get("expected_criteria")
            or case.get("criteria")
            or _expected_to_str(expected),
            "failure_reason": "；".join(errors) if errors else "",
            "fix_status": "",
            "note": note,
            "output": output,
            "payload": payload,
            "notes": notes,
            "duration_ms": dt_ms,
            "passed": passed_case,
            "errors": errors,
            "debug_url": debug_url,
        }
        report["results"].append(result)

        if passed_case:
            passed += 1
            _safe_print(f"PASS {case_id} {scenario} ({dt_ms}ms)")
        else:
            failed += 1
            _safe_print(f"FAIL {case_id} {scenario} ({dt_ms}ms)")
            _safe_print(f"  errors: {errors}")

        if idx < len(cases) - 1:
            time.sleep(1.0)

    report["run"]["call_stats"] = client.get_call_stats()

    report["summary"] = {
        "total": passed + failed,
        "passed": passed,
        "failed": failed,
    }

    report_json_content = json.dumps(report, ensure_ascii=False, indent=2)
    report_json_path.write_text(report_json_content, encoding="utf-8")
    report_json_latest_path.write_text(report_json_content, encoding="utf-8")

    report_md_content = _render_md(report)
    report_md_path.write_text(report_md_content, encoding="utf-8")
    report_md_latest_path.write_text(report_md_content, encoding="utf-8")

    _make_excel(report, report_xlsx_path)
    _make_excel(report, report_xlsx_latest_path)

    _safe_print(f"Report JSON: {report_json_path}")
    _safe_print(f"Report MD: {report_md_path}")
    _safe_print(f"Report XLSX: {report_xlsx_path}")

    return 0 if failed == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
