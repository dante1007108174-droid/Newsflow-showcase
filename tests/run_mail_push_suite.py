"""Run mail-push generation workflow suite and write reports.

This suite targets the *mail push generation* workflow that returns JSON:
  {"top_news": [...], "quick_news": [...]}

Usage:
  python tests/run_mail_push_suite.py

Outputs (tests/reports/):
  - mail_push_report_<run_id>.json
  - mail_push_report_<run_id>.md
  - mail_push_test_data_<run_id>.xlsx
  Plus stable "latest" files:
  - mail_push_report.json
  - mail_push_report.md
  - mail_push_test_data.xlsx

Notes:
  - Most scoring is deterministic and aligned with the workflow prompt.
  - "Sorting by news value" is intentionally left for manual judging.
"""

from __future__ import annotations

import io
import json
import os
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from coze_workflow_client import CozeWorkflowClient


ROOT = Path(__file__).resolve().parents[1]
SUITE_PATH = Path(__file__).resolve().parent / "mail_push_suite.json"
REPORT_DIR = Path(__file__).resolve().parent / "reports"


def _safe_print(s: str) -> None:
    try:
        print(s)
    except UnicodeEncodeError:
        print(s.encode("utf-8", errors="replace").decode("utf-8", errors="replace"))


def _now_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _load_json_file(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _save_json(path: Path, data: Dict[str, Any]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _truncate(s: str, n: int) -> str:
    s = s or ""
    return s if len(s) <= n else (s[: n - 1] + "…")


def _is_infra_exception(e: Exception) -> bool:
    msg = f"{type(e).__name__}: {e}".lower()
    return any(
        k in msg
        for k in (
            "sslerror",
            "certificate_verify_failed",
            "hostname mismatch",
            "connectionerror",
            "readtimeout",
            "connecttimeout",
            "proxyerror",
            "temporary failure",
            "name resolution",
        )
    )


def _safe_json_load(text: object) -> Tuple[Optional[Any], str]:
    if not isinstance(text, str):
        return None, "not_str"
    raw = text.strip()
    has_fence = "```" in raw
    cleaned = re.sub(r"^```json\s*", "", raw, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```\s*$", "", cleaned)
    try:
        return json.loads(cleaned), ("fence" if has_fence else "")
    except Exception:
        return None, "parse_error"


def _normalize_llm_payload(raw_text: object) -> Tuple[Optional[Dict[str, Any]], List[str]]:
    """Best-effort extract {top_news, quick_news} JSON from raw payload."""

    obj, meta = _safe_json_load(raw_text)
    notes: List[str] = []
    if meta == "fence":
        notes.append("has_markdown_fence")

    def try_extract(o: Any) -> Optional[Dict[str, Any]]:
        if isinstance(o, dict) and isinstance(o.get("top_news"), list) and isinstance(o.get("quick_news"), list):
            return o

        # Coze workflow may wrap the final answer as {"Output": "..."} (capital O)
        if isinstance(o, dict) and "Output" in o:
            inner = o.get("Output")
            if isinstance(inner, dict):
                notes.append("wrapped_Output_dict")
                return try_extract(inner)
            if isinstance(inner, str):
                notes.append("wrapped_Output_str")
                inner_obj, inner_meta = _safe_json_load(inner)
                if inner_meta == "fence":
                    notes.append("inner_has_markdown_fence")
                return try_extract(inner_obj)

        # Coze may wrap as {"content_type": 1, "data": "..."}
        if isinstance(o, dict) and "data" in o and isinstance(o.get("data"), str):
            notes.append("wrapped_content_data")
            inner_obj, inner_meta = _safe_json_load(o.get("data"))
            if inner_meta == "fence":
                notes.append("inner_has_markdown_fence")
            return try_extract(inner_obj)

        # Coze may wrap as {"output": "..."}
        if isinstance(o, dict) and "output" in o:
            inner = o.get("output")
            if isinstance(inner, dict):
                notes.append("wrapped_output_dict")
                return try_extract(inner)
            if isinstance(inner, str):
                notes.append("wrapped_output_str")
                inner_obj, inner_meta = _safe_json_load(inner)
                if inner_meta == "fence":
                    notes.append("inner_has_markdown_fence")
                return try_extract(inner_obj)
        if isinstance(o, str):
            inner_obj, inner_meta = _safe_json_load(o)
            if inner_meta == "fence":
                notes.append("inner_has_markdown_fence")
            return try_extract(inner_obj)
        return None

    extracted = try_extract(obj)
    if extracted is None:
        if meta == "parse_error":
            notes.append("json_parse_error")
        else:
            notes.append("missing_top_quick")
    return extracted, notes


def _parse_dt(s: object) -> Optional[datetime]:
    if not isinstance(s, str):
        return None
    s = s.strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None


def _bucket_by_rate(rate: float, table: List[Tuple[float, int]]) -> int:
    for min_rate, score in table:
        if rate >= min_rate:
            return score
    return table[-1][1]


def _score_json(parsed: Optional[Dict[str, Any]], notes: List[str]) -> int:
    if parsed is None:
        return 0
    # Prompt requires pure JSON (no markdown fence). Penalize but don't fail hard.
    has_fence = any("fence" in n for n in notes)
    return 8 if has_fence else 10


def _score_structure(parsed: Optional[Dict[str, Any]]) -> int:
    if not isinstance(parsed, dict):
        return 0
    top = parsed.get("top_news")
    quick = parsed.get("quick_news")
    if not isinstance(top, list) or not isinstance(quick, list):
        return 0

    top_req = ["title", "tag", "time", "summary", "source", "link"]
    q_req = ["title", "time", "source", "link"]

    def missing_ratio(items: List[Any], req: List[str]) -> float:
        if not items:
            return 0.0
        miss = 0
        total = len(items) * len(req)
        for it in items:
            if not isinstance(it, dict):
                miss += len(req)
                continue
            for k in req:
                v = it.get(k)
                if v is None or (isinstance(v, str) and not v.strip()):
                    miss += 1
        return miss / total if total else 0.0

    r = max(missing_ratio(top, top_req), missing_ratio(quick, q_req))
    if r == 0:
        base = 10
    elif r <= 0.05:
        base = 8
    elif r <= 0.15:
        base = 6
    elif r <= 0.30:
        base = 4
    else:
        base = 2

    qty_pen = 0
    if len(top) > 8:
        qty_pen += 1
    if len(quick) > 10:
        qty_pen += 1
    return max(0, base - qty_pen)


def _topic_tokens() -> Dict[str, List[str]]:
    return {
        "ai": [
            "ai",
            "人工智能",
            "大模型",
            "llm",
            "gpt",
            "openai",
            "anthropic",
            "deepseek",
            "gemini",
            "chatgpt",
            "agent",
            "智能体",
            "多模态",
        ],
        "tech": [
            "芯片",
            "半导体",
            "手机",
            "硬件",
            "操作系统",
            "os",
            "iphone",
            "android",
            "华为",
            "苹果",
            "小米",
            "三星",
            "英伟达",
            "nvidia",
            "amd",
            "intel",
            "gpu",
            "服务器",
            "电动车",
            "特斯拉",
        ],
        "finance": [
            "股",
            "a股",
            "港股",
            "美股",
            "央行",
            "降息",
            "利率",
            "cpi",
            "gdp",
            "ipo",
            "融资",
            "并购",
            "财报",
            "营收",
            "利润",
            "汇率",
            "美元",
            "人民币",
            "通胀",
            "基金",
            "债",
        ],
    }


def _score_relevance(parsed: Optional[Dict[str, Any]], topic: str) -> int:
    if not isinstance(parsed, dict):
        return 0
    top = parsed.get("top_news")
    if not isinstance(top, list) or not top:
        return 4

    tokens = _topic_tokens()
    tset = tokens.get(topic, tokens["tech"])

    def hit(text: object) -> bool:
        s = str(text or "").lower()
        return any(tok in s for tok in tset)

    hits = 0
    total = 0
    for it in top:
        total += 1
        if not isinstance(it, dict):
            continue
        if hit(it.get("title")) or hit(it.get("summary")) or hit(it.get("tag")):
            hits += 1
    rate = hits / total if total else 0.0
    return _bucket_by_rate(rate, [(0.9, 10), (0.75, 8), (0.6, 6), (0.4, 4), (0.0, 0)])


def _score_time_order_proxy(parsed: Optional[Dict[str, Any]]) -> int:
    """Proxy check: whether top_news times are non-increasing.

    Prompt requires 'news value' ordering and only uses time as tie-breaker.
    This metric is *not* the final sorting score; it's a quick sanity signal.
    """

    if not isinstance(parsed, dict):
        return 0
    top = parsed.get("top_news")
    if not isinstance(top, list) or len(top) < 2:
        return 8

    times = [_parse_dt(it.get("time")) if isinstance(it, dict) else None for it in top]
    if all(t is None for t in times):
        return 0

    inversions = 0
    pairs = 0
    for i in range(len(times) - 1):
        a, b = times[i], times[i + 1]
        if a is None or b is None:
            continue
        pairs += 1
        if a < b:
            inversions += 1
    if pairs == 0:
        return 4
    inv_rate = inversions / pairs
    if inv_rate == 0:
        return 10
    if inv_rate <= 0.25:
        return 8
    if inv_rate <= 0.5:
        return 6
    if inv_rate <= 0.8:
        return 4
    return 0


def _score_tag(parsed: Optional[Dict[str, Any]], topic: str) -> int:
    if not isinstance(parsed, dict):
        return 0
    top = parsed.get("top_news")
    if not isinstance(top, list) or not top:
        return 5

    # Prompt asks for specific domain tags (2-6 chars), and bans wide terms.
    banned_exact = {"ai", "人工智能", "科技", "技术", "财经", "经济"}

    bad = 0.0
    total = 0
    for it in top:
        total += 1
        if not isinstance(it, dict):
            bad += 1
            continue
        tag = str(it.get("tag") or "").strip()
        if not tag:
            bad += 1
            continue
        if tag.lower() in banned_exact:
            bad += 1
            continue

        # Prefer 2-6; penalize but do not hard fail outside.
        if len(tag) < 2 or len(tag) > 10:
            bad += 1
        elif len(tag) > 6:
            bad += 0.5

        # Topic-specific extra bans (e.g., finance shouldn't tag "财经").
        if topic == "finance" and tag in ("财经", "经济"):
            bad += 1

    good_rate = 1 - (bad / total if total else 1)
    return _bucket_by_rate(good_rate, [(0.9, 10), (0.75, 8), (0.6, 6), (0.4, 4), (0.0, 0)])


def _summary_char_len(s: str) -> int:
    # Approximate "字数" as non-whitespace codepoints.
    return len(re.sub(r"\s+", "", s or ""))


def _score_summary(parsed: Optional[Dict[str, Any]]) -> int:
    if not isinstance(parsed, dict):
        return 0
    top = parsed.get("top_news")
    if not isinstance(top, list) or not top:
        return 5

    def ok(s: object) -> bool:
        text = str(s or "").strip()
        if not text:
            return False
        if "暂无" in text or "N/A" in text:
            return False
        n = _summary_char_len(text)
        # Ideal is 100-120; allow a wider window but score indirectly via ratio.
        if n < 80 or n > 180:
            return False
        if text[-1] not in "。！？.!?":
            return False
        return True

    def ideal(s: object) -> bool:
        text = str(s or "").strip()
        if not ok(text):
            return False
        n = _summary_char_len(text)
        return 100 <= n <= 120

    ok_n = 0
    ideal_n = 0
    total = 0
    for it in top:
        total += 1
        if not isinstance(it, dict):
            continue
        summ = it.get("summary")
        if ok(summ):
            ok_n += 1
        if ideal(summ):
            ideal_n += 1

    ok_rate = ok_n / total if total else 0.0
    ideal_rate = ideal_n / total if total else 0.0
    base = _bucket_by_rate(ok_rate, [(0.8, 10), (0.6, 8), (0.4, 6), (0.2, 4), (0.0, 0)])
    # If many summaries are not in the ideal range, slightly cap the score.
    if base >= 8 and ideal_rate < 0.4:
        base = 6
    return base


def _ai_overlap_rate(parsed: Optional[Dict[str, Any]]) -> float:
    if not isinstance(parsed, dict):
        return 0.0
    top = parsed.get("top_news")
    if not isinstance(top, list) or not top:
        return 0.0
    ai_tokens = _topic_tokens()["ai"]

    def is_ai(it: Any) -> bool:
        if not isinstance(it, dict):
            return False
        s = (str(it.get("title") or "") + " " + str(it.get("summary") or "") + " " + str(it.get("tag") or "")).lower()
        return any(tok in s for tok in ai_tokens)

    hits = sum(1 for it in top if is_ai(it))
    return hits / max(len(top), 1)


def _call_workflow(client: CozeWorkflowClient, workflow_id: str, params: Dict[str, Any]) -> Tuple[str, int, Optional[str]]:
    """Call workflow exactly ONCE. No retries to prevent duplicate runs.

    IMPORTANT: This function makes exactly 1 POST request to trigger the workflow.
    If it fails, the exception propagates and the test case fails.
    This prevents the 12x API over-call issue we've seen before.
    """
    t0 = time.time()
    # Use async mode so the POST returns quickly; we poll run history.
    res = client.run_workflow(workflow_id, params, is_async=True, timeout_s=300)
    dt_ms = int((time.time() - t0) * 1000)
    return res.output or "", dt_ms, res.debug_url


def _run_one_case(client: CozeWorkflowClient, workflow_id: str, case: Dict[str, Any]) -> Dict[str, Any]:
    """Run a single case using the provided client.

    The client tracks API calls across all cases to enforce quota limits.
    This prevents runaway API usage (e.g., 114 calls for 9 test cases).
    """

    case_id = str(case.get("id") or "").strip()
    intent = str(case.get("intent") or "").strip()
    scenario = str(case.get("scenario") or "").strip()
    topic = str(case.get("topic") or "").strip() or "tech"
    inputs = case.get("inputs") or {}

    keyword = str(inputs.get("keyword") or "").strip()
    if not case_id or not keyword:
        return {}

    t0 = time.time()
    debug_url: Optional[str] = None
    try:
        # Use the provided client (with quota tracking)
        # Call workflow exactly once with the required parameters
        # NO fallback retry - if it fails, the test case fails
        params = {"keyword": keyword}
        output, dt_ms, debug_url = _call_workflow(client, workflow_id, params)

        duration_s = round(dt_ms / 1000.0, 3)
        extracted, notes = _normalize_llm_payload(output)

        s_json = _score_json(extracted, notes)
        s_struct = _score_structure(extracted)
        s_rel = _score_relevance(extracted, topic)
        s_tag = _score_tag(extracted, topic)
        s_sum = _score_summary(extracted)
        time_sort_proxy = _score_time_order_proxy(extracted)

        # Manual sorting score (news value ordering) is filled after the run.
        sorting_manual: Optional[int] = None

        # Partial score excludes sorting (max 80).
        partial_100 = (s_json * 0.2 + s_struct * 0.2 + s_rel * 0.2 + s_tag * 0.1 + s_sum * 0.1) * 10

        remarks: List[str] = []
        if "has_markdown_fence" in notes or "inner_has_markdown_fence" in notes:
            remarks.append("contains_markdown_fence")
        if extracted is None:
            remarks.append("json_extract_failed")
        if duration_s > 40:
            remarks.append("slow")
        if debug_url:
            remarks.append(f"debug_url={debug_url}")

        # Digest for manual judging.
        digest_top: List[Dict[str, Any]] = []
        if isinstance(extracted, dict) and isinstance(extracted.get("top_news"), list):
            for it in extracted.get("top_news")[:8]:
                if not isinstance(it, dict):
                    continue
                digest_top.append(
                    {
                        "time": it.get("time"),
                        "title": it.get("title"),
                        "tag": it.get("tag"),
                        "source": it.get("source"),
                    }
                )

        tech_ai_overlap = _ai_overlap_rate(extracted) if topic == "tech" else None

        # Status is pending manual sorting judgment unless JSON is unrecoverable.
        status = "FAIL" if extracted is None else "PENDING"
        status_cn = "❌失败" if status == "FAIL" else "⏳待评(排序)"

        return {
            "id": case_id,
            "intent": intent,
            "scenario": scenario,
            "topic": topic,
            "params": {"keyword": keyword},
            "duration_ms": dt_ms,
            "duration_s": duration_s,
            "status": status,
            "status_cn": status_cn,
            "scores": {
                "json": s_json,
                "structure": s_struct,
                "relevance": s_rel,
                "sorting_manual": sorting_manual,
                "time_sort_proxy": time_sort_proxy,
                "tag": s_tag,
                "summary": s_sum,
            },
            "partial_score_100": round(partial_100, 1),
            "notes": notes,
            "remarks": "；".join([x for x in remarks if x]),
            "tech_ai_overlap": tech_ai_overlap,
            "digest": {"top_news": digest_top},
            "output": output,
        }
    except Exception as e:
        duration_s = round((time.time() - t0), 3)
        status = "ERROR" if _is_infra_exception(e) else "FAIL"
        status_cn = "⚪异常" if status == "ERROR" else "❌失败"
        remarks = []
        if debug_url:
            remarks.append(f"debug_url={debug_url}")
        return {
            "id": case_id,
            "intent": intent,
            "scenario": scenario,
            "topic": topic,
            "params": {"keyword": keyword},
            "duration_ms": int(duration_s * 1000),
            "duration_s": duration_s,
            "status": status,
            "status_cn": status_cn,
            "scores": {},
            "partial_score_100": 0,
            "notes": [],
            "remarks": "；".join(remarks),
            "tech_ai_overlap": None,
            "digest": {"top_news": []},
            "output": "",
            "error": f"{type(e).__name__}: {e}",
        }


def _write_table(ws, headers: List[str], rows: List[List[Any]]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.append(headers)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
    ws.freeze_panes = "A2"
    for r in rows:
        ws.append(r)


def _apply_zebra_rows(ws) -> None:
    zebra_fill = PatternFill("solid", fgColor="F8FAFC")
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if i % 2 == 0:
            for cell in row:
                if cell.fill is None or cell.fill.fill_type is None:
                    cell.fill = zebra_fill


def _apply_result_fills(ws, *, result_col: int) -> None:
    pass_fill = PatternFill("solid", fgColor="DCFCE7")
    fail_fill = PatternFill("solid", fgColor="FEE2E2")
    pending_fill = PatternFill("solid", fgColor="FEF9C3")
    error_fill = PatternFill("solid", fgColor="E5E7EB")
    for row in ws.iter_rows(min_row=2):
        cell = row[result_col - 1]
        val = str(cell.value or "")
        if "通过" in val or "PASS" in val:
            cell.fill = pass_fill
        elif "待评" in val:
            cell.fill = pending_fill
        elif "失败" in val:
            cell.fill = fail_fill
        elif "异常" in val:
            cell.fill = error_fill


def _make_excel(report: Dict[str, Any], *, xlsx_path: Path) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "测试汇总"

    summary = report.get("summary") or {}
    run = report.get("run") or {}
    suite = report.get("suite") or {}

    # Derive a couple of helpful rollups (if manual scores were applied).
    results = report.get("results") or []
    partials = []
    finals = []
    for r in results:
        try:
            partials.append(float(r.get("partial_score_100") or 0.0))
        except Exception:
            pass
        if r.get("final_score_100") is not None:
            try:
                finals.append(float(r.get("final_score_100") or 0.0))
            except Exception:
                pass
    avg_partial = round(sum(partials) / max(len(partials), 1), 1) if partials else ""
    avg_final = round(sum(finals) / max(len(finals), 1), 1) if finals else ""

    summary_rows = [
        ["执行时间", run.get("timestamp") or ""],
        ["工作流ID", suite.get("workflow_id") or ""],
        ["测试版本", suite.get("version") or ""],
        ["总用例数", summary.get("total") or 0],
        ["通过", summary.get("passed") or 0],
        ["失败", summary.get("failed") or 0],
        ["异常", summary.get("errors") or 0],
        ["待评(排序)", summary.get("pending") or 0],
        ["通过率", summary.get("pass_rate") or ""],
        ["平均响应时间(秒)", summary.get("avg_duration_s") or ""],
        ["平均部分得分(缺排序)", avg_partial],
        ["平均总得分(含排序)", avg_final],
        ["科技主题AI重叠(均值)", summary.get("tech_ai_overlap_avg") or ""],
        ["通过线", str(suite.get("pass_threshold") or "")],
    ]
    _write_table(ws_summary, ["指标", "值"], summary_rows)
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 70

    ws_detail = wb.create_sheet("测试明细")
    headers = [
        "ID",
        "主题",
        "测试场景",
        "keyword",
        "结果",
        "响应时间(秒)",
        "JSON格式(10)",
        "结构完整(10)",
        "相关性(10)",
        "排序(人工,10)",
        "时间顺序代理(10)",
        "Tag质量(10)",
        "摘要质量(10)",
        "部分得分(缺排序,100)",
        "总得分(含排序,100)",
        "是否通过",
        "科技AI重叠率",
        "备注",
        "top_news预览",
    ]

    rows: List[List[Any]] = []
    for r in report.get("results", []) or []:
        sc = r.get("scores") or {}
        final_score_100 = r.get("final_score_100")
        passed = r.get("passed")
        top_preview = " | ".join(
            [
                _truncate(f"{it.get('time') or ''} {it.get('title') or ''}", 60)
                for it in (r.get("digest") or {}).get("top_news")
                if isinstance(it, dict)
            ][:5]
        )
        rows.append(
            [
                r.get("id"),
                r.get("topic"),
                r.get("scenario"),
                (r.get("params") or {}).get("keyword"),
                r.get("status_cn"),
                r.get("duration_s"),
                sc.get("json"),
                sc.get("structure"),
                sc.get("relevance"),
                sc.get("sorting_manual"),
                sc.get("time_sort_proxy"),
                sc.get("tag"),
                sc.get("summary"),
                r.get("partial_score_100"),
                final_score_100 if final_score_100 is not None else "",
                "✅" if passed is True else ("❌" if passed is False else ""),
                ("{:.0%}".format(r.get("tech_ai_overlap")) if isinstance(r.get("tech_ai_overlap"), float) else ""),
                r.get("remarks"),
                top_preview,
            ]
        )

    _write_table(ws_detail, headers, rows)
    _apply_result_fills(ws_detail, result_col=5)
    _apply_zebra_rows(ws_detail)

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws_detail.iter_rows(min_row=2):
        # Wrap scenario/remarks/preview so the sheet stays readable.
        for idx in (3, 18, 19):
            row[idx - 1].alignment = wrap

    widths = {
        "A": 12,
        "B": 10,
        "C": 22,
        "D": 10,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 14,
        "K": 14,
        "L": 12,
        "M": 12,
        "N": 18,
        "O": 18,
        "P": 10,
        "Q": 14,
        "R": 26,
        "S": 70,
    }
    for col, w in widths.items():
        ws_detail.column_dimensions[col].width = w

    # Topic stability summary (by topic, 3 runs)
    ws_topic = wb.create_sheet("主题统计")
    topic_stats: Dict[str, Dict[str, Any]] = {}
    for r in report.get("results", []) or []:
        t = r.get("topic") or "(unknown)"
        if t not in topic_stats:
            topic_stats[t] = {"total": 0, "fail": 0, "pending": 0, "avg_partial": []}
        topic_stats[t]["total"] += 1
        if r.get("status") == "FAIL":
            topic_stats[t]["fail"] += 1
        elif r.get("status") == "PENDING":
            topic_stats[t]["pending"] += 1
        try:
            topic_stats[t]["avg_partial"].append(float(r.get("partial_score_100") or 0.0))
        except Exception:
            pass

    topic_rows: List[List[Any]] = []
    for t, st in sorted(topic_stats.items(), key=lambda x: x[0]):
        partials: List[float] = st.get("avg_partial") or []
        avg_partial = round(sum(partials) / max(len(partials), 1), 1)
        topic_rows.append([t, st.get("total"), st.get("fail"), st.get("pending"), avg_partial])
    _write_table(ws_topic, ["主题", "用例数", "失败", "待评(排序)", "平均部分得分"], topic_rows)
    for col, w in {"A": 10, "B": 10, "C": 10, "D": 14, "E": 16}.items():
        ws_topic.column_dimensions[col].width = w
    _apply_zebra_rows(ws_topic)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(xlsx_path)
    except PermissionError:
        run_id = str((report.get("run") or {}).get("run_id") or datetime.now().strftime("%Y%m%d_%H%M%S"))
        fallback = xlsx_path.with_name(f"mail_push_test_data_{run_id}_fallback.xlsx")
        wb.save(fallback)
        _safe_print(f"Permission denied writing {xlsx_path}; wrote fallback: {fallback}")


def _render_md(report: Dict[str, Any]) -> str:
    summary = report.get("summary") or {}
    run = report.get("run") or {}
    suite = report.get("suite") or {}

    lines: List[str] = []
    lines.append("# mail_push_generation report")
    lines.append("")
    lines.append(f"- Run ID: `{run.get('run_id')}`")
    lines.append(f"- Time: `{run.get('timestamp')}`")
    lines.append(f"- Workflow ID: `{suite.get('workflow_id')}`")
    lines.append(f"- Version: `{suite.get('version')}`")
    lines.append(f"- Pass Threshold: `{suite.get('pass_threshold')}`")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- Total: {summary.get('total')}")
    lines.append(f"- Failed: {summary.get('failed')}")
    lines.append(f"- Errors: {summary.get('errors')}")
    lines.append(f"- Pending (manual sorting judge): {summary.get('pending')}")
    lines.append(f"- Avg Duration (s): {summary.get('avg_duration_s')}")
    lines.append(f"- Tech AI Overlap (avg): {summary.get('tech_ai_overlap_avg')}")
    lines.append("")
    lines.append("## Notes")
    for n in (suite.get("notes") or []):
        lines.append(f"- {n}")
    lines.append("")
    lines.append("## Failures")
    for r in report.get("results", []) or []:
        if r.get("status") not in ("FAIL", "ERROR"):
            continue
        lines.append("")
        lines.append(f"### {r.get('id')} {r.get('scenario')}")
        lines.append(f"- Result: {r.get('status_cn')}")
        if r.get("error"):
            lines.append(f"- Error: {r.get('error')}")
        if r.get("remarks"):
            lines.append(f"- Remarks: {r.get('remarks')}")
    return "\n".join(lines)


def main() -> int:
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
    except Exception:
        pass

    if not SUITE_PATH.exists():
        raise SystemExit(f"Suite file not found: {SUITE_PATH}")

    suite = _load_json_file(SUITE_PATH)
    suite_meta = suite.get("suite") or {}
    workflow_id = str(suite_meta.get("workflow_id") or "").strip()
    if not workflow_id:
        raise SystemExit("suite.workflow_id is required")

    cases = suite.get("cases")
    if not isinstance(cases, list) or not cases:
        raise SystemExit("suite.cases must be a non-empty list")

    # Validate Coze credentials early and set call quota.
    # Each case: 1 POST trigger + ~100-150 GET polls (async workflow takes ~2 min).
    # 150 calls per case is a safe upper bound.
    max_calls = len(cases) * 150
    client = CozeWorkflowClient.from_env(max_calls=max_calls)
    print(f"[Quota] API call limit set to {max_calls} for {len(cases)} test cases")
    print(f"[Quota] Set COZE_MAX_CALLS env var to override")

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_json_path = REPORT_DIR / f"mail_push_report_{run_id}.json"
    report_md_path = REPORT_DIR / f"mail_push_report_{run_id}.md"
    report_xlsx_path = REPORT_DIR / f"mail_push_test_data_{run_id}.xlsx"
    report_json_latest_path = REPORT_DIR / "mail_push_report.json"
    report_md_latest_path = REPORT_DIR / "mail_push_report.md"
    report_xlsx_latest_path = REPORT_DIR / "mail_push_test_data.xlsx"

    report: Dict[str, Any] = {
        "suite": suite_meta,
        "run": {
            "run_id": run_id,
            "timestamp": _now_ts(),
            "base_url": os.getenv("COZE_BASE_URL") or os.getenv("COZE_API_BASE") or "https://api.coze.cn",
            "workers": 1,
        },
        "results": [],
        "summary": {},
    }

    # Serial execution: run one case at a time to avoid runaway API calls.
    results_by_idx: Dict[int, Dict[str, Any]] = {}
    for idx, case in enumerate(cases):
        case_id = str(case.get("id") or f"case_{idx}")
        _safe_print(f"\n[{idx+1}/{len(cases)}] Running {case_id} ...")
        try:
            results_by_idx[idx] = _run_one_case(client, workflow_id, case) or {}
        except Exception as e:
            results_by_idx[idx] = {
                "id": case_id,
                "intent": "",
                "scenario": "(runner exception)",
                "topic": "",
                "params": {},
                "duration_ms": 0,
                "duration_s": 0,
                "status": "ERROR",
                "status_cn": "⚪异常",
                "scores": {},
                "partial_score_100": 0,
                "notes": [],
                "remarks": "",
                "tech_ai_overlap": None,
                "digest": {"top_news": []},
                "output": "",
                "error": f"{type(e).__name__}: {e}",
            }
        # Print progress
        r = results_by_idx[idx]
        _safe_print(f"  -> {r.get('status_cn', '?')} | {r.get('duration_s', 0)}s | partial={r.get('partial_score_100', 0)}")

    # Report final call stats
    stats = client.get_call_stats()
    _safe_print(f"[Stats] Total API calls: {stats['total_calls']} (limit: {stats['max_calls']})")

    report["results"] = [results_by_idx.get(i, {}) for i in range(len(cases)) if results_by_idx.get(i)]

    total = len(report["results"])
    failed = len([r for r in report["results"] if r.get("status") == "FAIL"])
    errors = len([r for r in report["results"] if r.get("status") == "ERROR"])
    pending = len([r for r in report["results"] if r.get("status") == "PENDING"])

    durations = [float(r.get("duration_s") or 0.0) for r in report["results"] if r.get("duration_s") is not None]
    avg_duration_s = round(sum(durations) / max(len(durations), 1), 3)

    tech_overlap = [r.get("tech_ai_overlap") for r in report["results"] if isinstance(r.get("tech_ai_overlap"), float)]
    tech_ai_overlap_avg = round(sum(tech_overlap) / max(len(tech_overlap), 1), 3) if tech_overlap else ""

    report["summary"] = {
        "total": total,
        "failed": failed,
        "errors": errors,
        "pending": pending,
        "avg_duration_s": avg_duration_s,
        "tech_ai_overlap_avg": ("{:.0%}".format(tech_ai_overlap_avg) if isinstance(tech_ai_overlap_avg, float) else ""),
    }

    _save_json(report_json_path, report)
    _save_json(report_json_latest_path, report)

    md = _render_md(report)
    report_md_path.write_text(md, encoding="utf-8")
    report_md_latest_path.write_text(md, encoding="utf-8")

    _make_excel(report, xlsx_path=report_xlsx_path)
    _make_excel(report, xlsx_path=report_xlsx_latest_path)

    _safe_print(f"Report JSON: {report_json_path}")
    _safe_print(f"Report MD: {report_md_path}")
    _safe_print(f"Report XLSX: {report_xlsx_path}")

    # Exit code: keep non-zero only for hard failures (JSON extract fail / infra errors).
    return 0 if (failed == 0 and errors == 0) else 1


if __name__ == "__main__":
    raise SystemExit(main())
